#!/usr/bin/env python3
"""
Generador del dashboard de la Porra Mundial 2026 (Reveni).

Lee el Excel de administración (Porra_Admin_v4_EN.xlsx) y produce un único
fichero HTML autocontenido con analíticas estilo "quién es el más rebelde",
"quién piensa como quién", estilo de apuesta, favoritos, etc.

Reejecutable: cuando se rellenen resultados/eliminatorias en el Excel, basta
con volver a correr el script y el HTML se regenera (incluyendo aciertos en
directo si ya hay resultados).

Uso:
    python3 generate_dashboard.py [ruta_xlsx] [ruta_salida_html]
"""

from __future__ import annotations

import json
import re
import sys
import statistics
from collections import Counter, defaultdict
from datetime import date as _date, datetime, timedelta
from pathlib import Path

import openpyxl

# --------------------------------------------------------------------------
# Configuración / constantes
# --------------------------------------------------------------------------

DEFAULT_XLSX = "Porra_Admin_v5_EN.xlsx"
DEFAULT_OUT = "index.html"  # raíz del sitio de GitHub Pages
LOGO_SVG_PATH = "reveni-logo.svg"

N_PARTICIPANTS_MAX = 30          # huecos P1..P30 en el Excel
RAW_FIRST_HOME_COL = 3           # columna C = primer "home" de P1
NAME_ROW = 6                     # fila con los nombres reales
GROUP_MATCH_ROWS = range(7, 79)  # 72 partidos de fase de grupos
QUALIFIER_ROWS = range(80, 116)  # 1st/2nd/3rd de cada grupo (A..L)
THIRDS_ROWS = range(117, 125)    # 8 mejores terceros

# W-numbers for knockout bracket feeders (must match JS KO_WBASE).
KO_WBASE = {"R32": 72, "R16": 88, "QF": 96, "SF": 100}

KNOCKOUT_ROUNDS = [
    {"key": "r32", "code": "R32", "label_es": "Dieciseisavos", "label_en": "Round of 32", "first_row": 126, "matches": 16, "advance_points": 1},
    {"key": "r16", "code": "R16", "label_es": "Octavos", "label_en": "Round of 16", "first_row": 175, "matches": 8, "advance_points": 2},
    {"key": "qf", "code": "QF", "label_es": "Cuartos", "label_en": "Quarter-finals", "first_row": 200, "matches": 4, "advance_points": 4},
    {"key": "sf", "code": "SF", "label_es": "Semifinales", "label_en": "Semi-finals", "first_row": 213, "matches": 2, "advance_points": 6},
]
FINAL_MATCHES = [
    {"key": "third_place_match", "code": "3P", "label_es": "Tercer puesto", "label_en": "Third-place match", "score_row": 220, "penalty_row": 221},
    {"key": "final", "code": "FINAL", "label_es": "Final", "label_en": "Final", "score_row": 223, "penalty_row": 224},
]
SPECIAL_OUTRIGHT_ROWS = {
    "third_place": {"row": 222, "label_es": "Tercer puesto", "label_en": "Third place", "points": 4},
    "champion": {"row": 225, "label_es": "Campeón", "label_en": "Champion", "points": 12},
    "runner_up": {"row": 226, "label_es": "Subcampeón", "label_en": "Runner-up", "points": 8},
}
AWARD_ROWS = {
    "top_scorer": {"row": 228, "label_es": "Máximo goleador", "label_en": "Top scorer", "points": 8},
    "ballon_dor": {"row": 229, "label_es": "Balón de Oro", "label_en": "Ballon d'Or", "points": 8},
}

# Calendario oficial FIFA descargado de la página de fixtures el 2026-06-28.
# kickoff_et guarda el instante base en Eastern Time; el dashboard lo muestra
# como hora peninsular española (ES) o británica (EN), igual que la fase de grupos.
KNOCKOUT_MATCH_SCHEDULE = {
    # Code order follows the knockout template copied into Raw data.
    "R32-M1": {"home": "Germany", "away": "Paraguay", "venue": "Boston Stadium", "city": "Boston", "kickoff_et": "2026-06-29T16:30"},
    "R32-M2": {"home": "France", "away": "Sweden", "venue": "New York/New Jersey Stadium", "city": "New Jersey", "kickoff_et": "2026-06-30T17:00"},
    "R32-M3": {"home": "South Africa", "away": "Canada", "venue": "Los Angeles Stadium", "city": "Los Angeles", "kickoff_et": "2026-06-28T15:00"},
    "R32-M4": {"home": "Netherlands", "away": "Morocco", "venue": "Monterrey Stadium", "city": "Monterrey", "kickoff_et": "2026-06-29T21:00"},
    "R32-M5": {"home": "Portugal", "away": "Croatia", "venue": "Toronto Stadium", "city": "Toronto", "kickoff_et": "2026-07-02T19:00"},
    "R32-M6": {"home": "Spain", "away": "Austria", "venue": "Los Angeles Stadium", "city": "Los Angeles", "kickoff_et": "2026-07-02T15:00"},
    "R32-M7": {"home": "USA", "away": "Bosnia and Herzegovina", "venue": "San Francisco Bay Area Stadium", "city": "San Francisco Bay Area", "kickoff_et": "2026-07-01T20:00"},
    "R32-M8": {"home": "Belgium", "away": "Senegal", "venue": "Seattle Stadium", "city": "Seattle", "kickoff_et": "2026-07-01T16:00"},
    "R32-M9": {"home": "Brazil", "away": "Japan", "venue": "Houston Stadium", "city": "Houston", "kickoff_et": "2026-06-29T13:00"},
    "R32-M10": {"home": "Côte d'Ivoire", "away": "Norway", "venue": "Dallas Stadium", "city": "Dallas", "kickoff_et": "2026-06-30T13:00"},
    "R32-M11": {"home": "Mexico", "away": "Ecuador", "venue": "Mexico City Stadium", "city": "Mexico City", "kickoff_et": "2026-06-30T21:00"},
    "R32-M12": {"home": "England", "away": "Congo DR", "venue": "Atlanta Stadium", "city": "Atlanta", "kickoff_et": "2026-07-01T12:00"},
    "R32-M13": {"home": "Argentina", "away": "Cabo Verde", "venue": "Miami Stadium", "city": "Miami", "kickoff_et": "2026-07-03T18:00"},
    "R32-M14": {"home": "Australia", "away": "Egypt", "venue": "Dallas Stadium", "city": "Dallas", "kickoff_et": "2026-07-03T14:00"},
    "R32-M15": {"home": "Switzerland", "away": "Algeria", "venue": "BC Place Vancouver", "city": "Vancouver", "kickoff_et": "2026-07-02T23:00"},
    "R32-M16": {"home": "Colombia", "away": "Ghana", "venue": "Kansas City Stadium", "city": "Kansas City", "kickoff_et": "2026-07-03T21:30"},
    "R16-M1": {"home": "W73", "away": "W74", "venue": "Philadelphia Stadium", "city": "Philadelphia", "kickoff_et": "2026-07-04T17:00"},
    "R16-M2": {"home": "W75", "away": "W76", "venue": "Houston Stadium", "city": "Houston", "kickoff_et": "2026-07-04T13:00"},
    "R16-M3": {"home": "W77", "away": "W78", "venue": "Dallas Stadium", "city": "Dallas", "kickoff_et": "2026-07-06T15:00"},
    "R16-M4": {"home": "W79", "away": "W80", "venue": "Seattle Stadium", "city": "Seattle", "kickoff_et": "2026-07-06T20:00"},
    "R16-M5": {"home": "W81", "away": "W82", "venue": "New York/New Jersey Stadium", "city": "New Jersey", "kickoff_et": "2026-07-05T16:00"},
    "R16-M6": {"home": "W83", "away": "W84", "venue": "Mexico City Stadium", "city": "Mexico City", "kickoff_et": "2026-07-05T20:00"},
    "R16-M7": {"home": "W85", "away": "W86", "venue": "Atlanta Stadium", "city": "Atlanta", "kickoff_et": "2026-07-07T12:00"},
    "R16-M8": {"home": "W87", "away": "W88", "venue": "BC Place Vancouver", "city": "Vancouver", "kickoff_et": "2026-07-07T16:00"},
    "QF-M1": {"home": "W89", "away": "W90", "venue": "Boston Stadium", "city": "Boston", "kickoff_et": "2026-07-09T16:00"},
    "QF-M2": {"home": "W91", "away": "W92", "venue": "Los Angeles Stadium", "city": "Los Angeles", "kickoff_et": "2026-07-10T15:00"},
    "QF-M3": {"home": "W93", "away": "W94", "venue": "Miami Stadium", "city": "Miami", "kickoff_et": "2026-07-11T17:00"},
    "QF-M4": {"home": "W95", "away": "W96", "venue": "Kansas City Stadium", "city": "Kansas City", "kickoff_et": "2026-07-11T21:00"},
    "SF-M1": {"home": "W97", "away": "W98", "venue": "Dallas Stadium", "city": "Dallas", "kickoff_et": "2026-07-14T15:00"},
    "SF-M2": {"home": "W99", "away": "W100", "venue": "Atlanta Stadium", "city": "Atlanta", "kickoff_et": "2026-07-15T15:00"},
    "3P": {"home": "RU101", "away": "RU102", "venue": "Miami Stadium", "city": "Miami", "kickoff_et": "2026-07-18T17:00"},
    "FINAL": {"home": "W101", "away": "W102", "venue": "New York/New Jersey Stadium", "city": "New Jersey", "kickoff_et": "2026-07-19T15:00"},
}

FIFA_TEAM_ALIASES = {
    "Bosnia and Herzegovina": "Bosnia-Herz.",
    "Cabo Verde": "Cape Verde",
    "Congo DR": "DR Congo",
    "Côte d'Ivoire": "Ivory Coast",
    "Morrocco": "Morocco",
}

MATCH_DATES = {
    "GA-M1": "2026-06-11", "GA-M2": "2026-06-11",
    "GA-M3": "2026-06-18", "GA-M4": "2026-06-18",
    "GA-M5": "2026-06-24", "GA-M6": "2026-06-24",
    "GB-M1": "2026-06-12", "GB-M2": "2026-06-13",
    "GB-M3": "2026-06-18", "GB-M4": "2026-06-18",
    "GB-M5": "2026-06-24", "GB-M6": "2026-06-24",
    "GC-M1": "2026-06-13", "GC-M2": "2026-06-13",
    "GC-M3": "2026-06-19", "GC-M4": "2026-06-19",
    "GC-M5": "2026-06-24", "GC-M6": "2026-06-24",
    "GD-M1": "2026-06-12", "GD-M2": "2026-06-13",
    "GD-M3": "2026-06-19", "GD-M4": "2026-06-19",
    "GD-M5": "2026-06-25", "GD-M6": "2026-06-25",
    "GE-M1": "2026-06-14", "GE-M2": "2026-06-14",
    "GE-M3": "2026-06-20", "GE-M4": "2026-06-20",
    "GE-M5": "2026-06-25", "GE-M6": "2026-06-25",
    "GF-M1": "2026-06-14", "GF-M2": "2026-06-14",
    "GF-M3": "2026-06-20", "GF-M4": "2026-06-20",
    "GF-M5": "2026-06-25", "GF-M6": "2026-06-25",
    "GG-M1": "2026-06-15", "GG-M2": "2026-06-15",
    "GG-M3": "2026-06-21", "GG-M4": "2026-06-21",
    "GG-M5": "2026-06-26", "GG-M6": "2026-06-26",
    "GH-M1": "2026-06-15", "GH-M2": "2026-06-15",
    "GH-M3": "2026-06-21", "GH-M4": "2026-06-21",
    "GH-M5": "2026-06-26", "GH-M6": "2026-06-26",
    "GI-M1": "2026-06-16", "GI-M2": "2026-06-16",
    "GI-M3": "2026-06-22", "GI-M4": "2026-06-22",
    "GI-M5": "2026-06-26", "GI-M6": "2026-06-26",
    "GJ-M1": "2026-06-16", "GJ-M2": "2026-06-16",
    "GJ-M3": "2026-06-22", "GJ-M4": "2026-06-22",
    "GJ-M5": "2026-06-27", "GJ-M6": "2026-06-27",
    "GK-M1": "2026-06-17", "GK-M2": "2026-06-17",
    "GK-M3": "2026-06-23", "GK-M4": "2026-06-23",
    "GK-M5": "2026-06-27", "GK-M6": "2026-06-27",
    "GL-M1": "2026-06-17", "GL-M2": "2026-06-17",
    "GL-M3": "2026-06-23", "GL-M4": "2026-06-23",
    "GL-M5": "2026-06-27", "GL-M6": "2026-06-27",
}

# Horario real de saque del Mundial 2026 (fuente: calendario oficial), en hora
# del Este de EE. UU. (ET, UTC-4 en junio). Se muestra convertido a hora
# peninsular española (CEST, UTC+2) = ET + 6h. Las fechas de agrupación por
# jornada viven en MATCH_DATES; aquí solo está el instante de saque.
MATCH_KICKOFF_ET = {
    "GA-M1": "2026-06-11T15:00", "GA-M2": "2026-06-11T22:00",
    "GA-M3": "2026-06-18T12:00", "GA-M4": "2026-06-18T21:00",
    "GA-M5": "2026-06-24T21:00", "GA-M6": "2026-06-24T21:00",
    "GB-M1": "2026-06-12T15:00", "GB-M2": "2026-06-13T15:00",
    "GB-M3": "2026-06-18T15:00", "GB-M4": "2026-06-18T18:00",
    "GB-M5": "2026-06-24T15:00", "GB-M6": "2026-06-24T15:00",
    "GC-M1": "2026-06-13T18:00", "GC-M2": "2026-06-13T21:00",
    "GC-M3": "2026-06-19T18:00", "GC-M4": "2026-06-19T20:30",
    "GC-M5": "2026-06-24T18:00", "GC-M6": "2026-06-24T18:00",
    "GD-M1": "2026-06-12T21:00", "GD-M2": "2026-06-14T00:00",
    "GD-M3": "2026-06-19T00:00", "GD-M4": "2026-06-19T15:00",
    "GD-M5": "2026-06-25T22:00", "GD-M6": "2026-06-25T22:00",
    "GE-M1": "2026-06-14T13:00", "GE-M2": "2026-06-14T19:00",
    "GE-M3": "2026-06-20T16:00", "GE-M4": "2026-06-20T20:00",
    "GE-M5": "2026-06-25T16:00", "GE-M6": "2026-06-25T16:00",
    "GF-M1": "2026-06-14T16:00", "GF-M2": "2026-06-14T22:00",
    "GF-M3": "2026-06-20T13:00", "GF-M4": "2026-06-21T00:00",
    "GF-M5": "2026-06-25T19:00", "GF-M6": "2026-06-25T19:00",
    "GG-M1": "2026-06-15T15:00", "GG-M2": "2026-06-15T21:00",
    "GG-M3": "2026-06-21T15:00", "GG-M4": "2026-06-21T21:00",
    "GG-M5": "2026-06-26T23:00", "GG-M6": "2026-06-26T23:00",
    "GH-M1": "2026-06-15T12:00", "GH-M2": "2026-06-15T18:00",
    "GH-M3": "2026-06-21T12:00", "GH-M4": "2026-06-21T18:00",
    "GH-M5": "2026-06-26T20:00", "GH-M6": "2026-06-26T20:00",
    "GI-M1": "2026-06-16T15:00", "GI-M2": "2026-06-16T18:00",
    "GI-M3": "2026-06-22T17:00", "GI-M4": "2026-06-22T20:00",
    "GI-M5": "2026-06-26T15:00", "GI-M6": "2026-06-26T15:00",
    "GJ-M1": "2026-06-16T21:00", "GJ-M2": "2026-06-17T00:00",
    "GJ-M3": "2026-06-22T13:00", "GJ-M4": "2026-06-22T23:00",
    "GJ-M5": "2026-06-27T22:00", "GJ-M6": "2026-06-27T22:00",
    "GK-M1": "2026-06-17T13:00", "GK-M2": "2026-06-17T22:00",
    "GK-M3": "2026-06-23T13:00", "GK-M4": "2026-06-23T22:00",
    "GK-M5": "2026-06-27T19:30", "GK-M6": "2026-06-27T19:30",
    "GL-M1": "2026-06-17T16:00", "GL-M2": "2026-06-17T19:00",
    "GL-M3": "2026-06-23T16:00", "GL-M4": "2026-06-23T19:00",
    "GL-M5": "2026-06-27T17:00", "GL-M6": "2026-06-27T17:00",
}

# Offsets desde ET (UTC-4 en junio): peninsular = UTC+2 (ET+6), R. Unido = BST
# UTC+1 (ET+5). El idioma del dashboard decide cuál se muestra.
_ET_OFFSETS = {"es": timedelta(hours=6), "uk": timedelta(hours=5)}


def kickoff_info(code):
    """Horas de saque por idioma (peninsular para ES, británica para EN) más el
    instante absoluto para ordenar. next_<lang>=True si en esa zona ya es el día
    siguiente al de la jornada (partidos de madrugada)."""
    et = MATCH_KICKOFF_ET.get(code)
    if not et:
        return {"time_es": "", "time_uk": "", "next_es": False, "next_uk": False, "dt": ""}
    base = datetime.fromisoformat(et)
    porra_date = MATCH_DATES.get(code, "")
    # cualquier offset fijo sirve como clave de orden (mismo instante para todos)
    out = {"dt": (base + _ET_OFFSETS["es"]).isoformat()}
    for lang, off in _ET_OFFSETS.items():
        local = base + off
        out[f"time_{lang}"] = local.strftime("%H:%M")
        out[f"next_{lang}"] = local.date().isoformat() != porra_date
    return out

# Traducción EN -> ES (reconocible) + bandera
TEAMS = {
    "Mexico": ("México", "🇲🇽"), "South Africa": ("Sudáfrica", "🇿🇦"),
    "South Korea": ("Corea del Sur", "🇰🇷"), "Czech Rep.": ("Rep. Checa", "🇨🇿"),
    "Canada": ("Canadá", "🇨🇦"), "Bosnia-Herz.": ("Bosnia", "🇧🇦"),
    "Qatar": ("Catar", "🇶🇦"), "Switzerland": ("Suiza", "🇨🇭"),
    "Brazil": ("Brasil", "🇧🇷"), "Morocco": ("Marruecos", "🇲🇦"),
    "Haiti": ("Haití", "🇭🇹"), "Scotland": ("Escocia", "🏴󠁧󠁢󠁳󠁣󠁴󠁿"),
    "USA": ("EE. UU.", "🇺🇸"), "Paraguay": ("Paraguay", "🇵🇾"),
    "Australia": ("Australia", "🇦🇺"), "Turkey": ("Turquía", "🇹🇷"),
    "Germany": ("Alemania", "🇩🇪"), "Curacao": ("Curazao", "🇨🇼"),
    "Ivory Coast": ("Costa de Marfil", "🇨🇮"), "Ecuador": ("Ecuador", "🇪🇨"),
    "Netherlands": ("Países Bajos", "🇳🇱"), "Japan": ("Japón", "🇯🇵"),
    "Sweden": ("Suecia", "🇸🇪"), "Tunisia": ("Túnez", "🇹🇳"),
    "Belgium": ("Bélgica", "🇧🇪"), "Egypt": ("Egipto", "🇪🇬"),
    "Iran": ("Irán", "🇮🇷"), "New Zealand": ("Nueva Zelanda", "🇳🇿"),
    "Spain": ("España", "🇪🇸"), "Cape Verde": ("Cabo Verde", "🇨🇻"),
    "Uruguay": ("Uruguay", "🇺🇾"), "Saudi Arabia": ("Arabia Saudí", "🇸🇦"),
    "France": ("Francia", "🇫🇷"), "Senegal": ("Senegal", "🇸🇳"),
    "Iraq": ("Irak", "🇮🇶"), "Norway": ("Noruega", "🇳🇴"),
    "Argentina": ("Argentina", "🇦🇷"), "Algeria": ("Argelia", "🇩🇿"),
    "Austria": ("Austria", "🇦🇹"), "Jordan": ("Jordania", "🇯🇴"),
    "Portugal": ("Portugal", "🇵🇹"), "DR Congo": ("RD Congo", "🇨🇩"),
    "Uzbekistan": ("Uzbekistán", "🇺🇿"), "Colombia": ("Colombia", "🇨🇴"),
    "England": ("Inglaterra", "🏴󠁧󠁢󠁥󠁮󠁧󠁿"), "Croatia": ("Croacia", "🇭🇷"),
    "Ghana": ("Ghana", "🇬🇭"), "Panama": ("Panamá", "🇵🇦"),
}


ES2EN = {es: en for en, (es, _fl) in TEAMS.items()}

TRIVIA = {
    "Mexico": [
        ("Mexico City se hunde ~50cm al año porque está sobre un lago seco.", "Mexico City sinks ~50cm per year because it's built on a drained lake."),
        ("México regaló un axolote a la reina Isabel II en 1953.", "Mexico gifted an axolotl to Queen Elizabeth II in 1953."),
        ("El chicle fue inventado por el general mexicano Antonio López de Santa Anna.", "Chewing gum was invented by Mexican general Antonio López de Santa Anna."),
        ("El agave del mezcal tarda entre 7 y 30 años en madurar antes de destilarse.", "Mezcal agave takes 7 to 30 years to mature before distillation."),
    ],
    "South Africa": [
        ("Sudáfrica tiene 11 idiomas oficiales, un récord mundial.", "South Africa has 11 official languages, a world record."),
        ("Un sudafricano inventó el CT scan.", "A South African invented the CT scan."),
        ("Table Mountain tiene más especies de plantas que todo Reino Unido.", "Table Mountain has more plant species than the entire UK."),
        ("Sudáfrica es el único país con tres capitales: Pretoria, Ciudad del Cabo y Bloemfontein.", "South Africa is the only country with three capitals: Pretoria, Cape Town, and Bloemfontein."),
    ],
    "South Korea": [
        ("Corea del Sur tiene más robots per cápita que cualquier país.", "South Korea has more robots per capita than any other country."),
        ("Existe un Día del Perro Negro en Corea (3 de enero).", "There's a Black Dog Day in Korea (January 3rd)."),
        ("El servicio postal coreano entrega con drones desde 2017.", "Korea's postal service has delivered by drone since 2017."),
    ],
    "Czech Rep.": [
        ("Los checos beben más cerveza per cápita del mundo: ~140 litros/año.", "Czechs drink the most beer per capita in the world: ~140 liters/year."),
        ("La palabra 'robot' fue inventada por el escritor checo Karel Čapek en 1920.", "The word 'robot' was coined by Czech writer Karel Čapek in 1920."),
        ("Praga tiene un museo dedicado exclusivamente a Franz Kafka.", "Prague has a museum dedicated entirely to Franz Kafka."),
    ],
    "Canada": [
        ("En 1967 Canadá construyó una pista de aterrizaje para OVNIs en Alberta.", "In 1967 Canada built a UFO landing pad in Alberta."),
        ("Canadá tiene más lagos que todos los demás países juntos.", "Canada has more lakes than every other country combined."),
        ("La RCMP entrenó una vez un hipopótamo como policía montado. Fracasó.", "The RCMP once trained a hippo as a mounted police officer. It failed."),
        ("Canadá tiene la costa más larga del mundo: más de 202.000 km.", "Canada has the world's longest coastline: over 202,000 km."),
    ],
    "Bosnia-Herz.": [
        ("Sarajevo tuvo tranvías antes que Londres.", "Sarajevo had trams before London."),
        ("Bosnia alberga la última selva virgen de Europa: Perućica.", "Bosnia is home to Europe's last primeval forest: Perućica."),
        ("Un bosnio afirma haber descubierto pirámides más antiguas que las de Egipto. Nadie le cree.", "A Bosnian claims to have discovered pyramids older than Egypt's. Nobody believes him."),
        ("Sarajevo acogió los Juegos Olímpicos de Invierno de 1984.", "Sarajevo hosted the 1984 Winter Olympics."),
    ],
    "Qatar": [
        ("Qatar usó jinetes robot en carreras de camellos desde 2004.", "Qatar has used robot jockeys in camel races since 2004."),
        ("Qatar tiene cero impuesto sobre la renta personal.", "Qatar has zero personal income tax."),
        ("El aeropuerto de Doha tiene un oso de peluche gigante de 7 metros.", "Doha's airport features a giant 7-meter teddy bear."),
    ],
    "Switzerland": [
        ("Suiza tiene suficientes búnkeres nucleares para toda su población.", "Switzerland has enough nuclear bunkers for its entire population."),
        ("Es ilegal tener un solo conejillo de indias en Suiza (son sociales).", "It's illegal to own just one guinea pig in Switzerland (they're social)."),
        ("Suiza tiene un plan para invadir… Alemania. Por si acaso.", "Switzerland has a plan to invade… Germany. Just in case."),
        ("El chocolate con leche se inventó en Suiza en 1875 (Daniel Peter).", "Milk chocolate was invented in Switzerland in 1875 (Daniel Peter)."),
    ],
    "Brazil": [
        ("Brasil tuvo un emperador que se declaró 'Protector de los Animales'.", "Brazil had an emperor who declared himself 'Protector of Animals'."),
        ("La prisión de Carandiru tenía su propia liga de fútbol oficial.", "Carandiru prison had its own official football league."),
        ("Brasil es el mayor exportador mundial de piedras preciosas.", "Brazil is the world's largest exporter of gemstones."),
        ("El Cristo Redentor de Río mide 38 metros y pesa unas 635 toneladas.", "Rio's Christ the Redeemer is 38 meters tall and weighs about 635 tonnes."),
    ],
    "Morocco": [
        ("Marruecos fundó la universidad más antigua del mundo aún en funcionamiento (859 d.C.).", "Morocco founded the world's oldest continuously operating university (859 AD)."),
        ("En Marruecos hay cabras que trepan a los árboles para comer argán.", "In Morocco, goats climb trees to eat argan fruit."),
        ("El rey de Marruecos tiene un trono portátil que viaja con él.", "The King of Morocco has a portable throne that travels with him."),
        ("La película Casablanca no se rodó en Marruecos: se filmó en un estudio de Hollywood.", "The movie Casablanca wasn't filmed in Morocco — it was shot on a Hollywood soundstage."),
    ],
    "Haiti": [
        ("Haití fue el primer país negro independiente del mundo (1804).", "Haiti was the world's first independent Black republic (1804)."),
        ("Haití tiene la fortaleza más grande de América: la Citadelle Laferrière.", "Haiti has the largest fortress in the Americas: Citadelle Laferrière."),
        ("El pico más alto del Caribe está en Haití, no en República Dominicana.", "The Caribbean's highest peak is in Haiti, not the Dominican Republic."),
    ],
    "Scotland": [
        ("El animal nacional de Escocia es el unicornio.", "Scotland's national animal is the unicorn."),
        ("Escocia tiene más de 790 islas, la mayoría deshabitadas.", "Scotland has over 790 islands, most uninhabited."),
        ("Edimburgo tiene un volcán extinto en pleno centro de la ciudad.", "Edinburgh has an extinct volcano right in the city center."),
    ],
    "USA": [
        ("EE. UU. gastó 22 millones de dólares en desarrollar una pluma que escribiera en el espacio. Los rusos usaron lápiz.", "The US spent $22 million developing a pen that writes in space. The Russians used pencils."),
        ("Existe una ciudad llamada 'Hell' en Michigan. Se congela cada invierno.", "There's a town called 'Hell' in Michigan. It freezes every winter."),
        ("La bandera de EE. UU. fue diseñada por un estudiante de 17 años como proyecto escolar.", "The US flag was designed by a 17-year-old student as a school project."),
        ("Yellowstone fue el primer parque nacional del mundo (1872).", "Yellowstone was the world's first national park (1872)."),
    ],
    "Paraguay": [
        ("Paraguay fue el único país latinoamericano que envió condolencias al gobierno confederado en la Guerra Civil americana.", "Paraguay was the only Latin American country to send condolences to the Confederate government during the American Civil War."),
        ("Paraguay tuvo el primer ferrocarril de Sudamérica (1854).", "Paraguay had South America's first railway (1854)."),
        ("La Armada paraguaya es la más grande del mundo… sin acceso al mar.", "Paraguay's navy is the world's largest… with no access to the sea."),
        ("El guaraní es cooficial con el español: único país de América con una lengua indígena oficial a nivel nacional.", "Guaraní is co-official with Spanish — the only country in the Americas with an indigenous national language."),
    ],
    "Australia": [
        ("Australia perdió una guerra contra 20.000 emus en 1932.", "Australia lost a war against 20,000 emus in 1932."),
        ("Australia tiene un primer ministro que fue devorado (presuntamente) por caníbales en 1803.", "Australia had a PM who was allegedly eaten by cannibals in 1803."),
        ("Hay más canguros que personas en Australia (~50M vs 26M).", "There are more kangaroos than people in Australia (~50M vs 26M)."),
        ("La Gran Barrera de Coral mide más de 2.300 km y es visible desde el espacio.", "The Great Barrier Reef stretches over 2,300 km and is visible from space."),
    ],
    "Turkey": [
        ("Los tulipanes vienen de Turquía, no de Holanda.", "Tulips come from Turkey, not the Netherlands."),
        ("Papá Noel (San Nicolás) nació en lo que hoy es Turquía.", "Santa Claus (Saint Nicholas) was born in what is now Turkey."),
        ("Estambul tuvo un gato alcalde no oficial llamado Tombili.", "Istanbul had an unofficial cat mayor named Tombili."),
    ],
    "Germany": [
        ("Alemania tiene ~1.500 tipos de cerveza diferentes.", "Germany has ~1,500 different types of beer."),
        ("Intentaron enseñar a los perros a hablar en la Alemania nazi. No funcionó.", "They tried to teach dogs to talk in Nazi Germany. It didn't work."),
        ("Hay más clubes de fútbol que supermercados en Alemania.", "There are more football clubs than supermarkets in Germany."),
        ("Berlín tiene más puentes que Venecia: unos 1.700.", "Berlin has more bridges than Venice: about 1,700."),
    ],
    "Curacao": [
        ("El licor azul de Curazao se hace con cáscaras de naranja de la isla.", "Blue Curaçao liqueur is made from orange peels grown on the island."),
        ("Curazao fue colonia holandesa, española, francesa e inglesa… a veces todas a la vez.", "Curaçao was a Dutch, Spanish, French, and English colony… sometimes all at once."),
        ("Curazao tiene una de las sinagogas más antiguas de América (1732).", "Curaçao has one of the oldest synagogues in the Americas (1732)."),
    ],
    "Ivory Coast": [
        ("Costa de Marfil produce el 40% del cacao mundial.", "Ivory Coast produces 40% of the world's cocoa."),
        ("Tiene una basílica más grande que la de San Pedro en el Vaticano.", "It has a basilica larger than St. Peter's in the Vatican."),
        ("El nombre 'Costa de Marfil' viene de cuando era el mayor mercado de marfil del mundo.", "The name 'Ivory Coast' comes from when it was the world's largest ivory market."),
        ("Costa de Marfil es el mayor exportador mundial de nueces de cola.", "Ivory Coast is the world's largest exporter of kola nuts."),
    ],
    "Ecuador": [
        ("En la línea ecuatorial de Ecuador, el agua no hace remolino al desaguar.", "At Ecuador's equator line, water doesn't swirl when draining."),
        ("Ecuador tiene montañas donde puedes ver nieve estando en el ecuador.", "Ecuador has mountains where you can see snow while standing on the equator."),
        ("Las Islas Galápagos inspiraron la teoría de la evolución de Darwin.", "The Galápagos Islands inspired Darwin's theory of evolution."),
        ("Ecuador fue el primer país en reconocer derechos legales a la naturaleza en su Constitución (2008).", "Ecuador was the first country to grant legal rights to nature in its constitution (2008)."),
    ],
    "Netherlands": [
        ("Los Países Bajos están por debajo del nivel del mar en un 26% de su territorio.", "The Netherlands is below sea level in 26% of its territory."),
        ("Hay más bicicletas que personas en los Países Bajos.", "There are more bicycles than people in the Netherlands."),
        ("Holanda exportó tulipanes por primera vez como error: alguien se comió los bulbos pensando que eran cebollas.", "Holland first exported tulips by mistake: someone ate the bulbs thinking they were onions."),
        ("Los holandeses son los más altos del mundo: la altura media del hombre supera 1,83 m.", "The Dutch are the world's tallest people: average male height exceeds 1.83 m."),
    ],
    "Japan": [
        ("Japón tiene más mascotas que niños.", "Japan has more pets than children."),
        ("Hay una isla en Japón llena de conejos salvajes: Ōkunoshima.", "There's an island in Japan full of wild rabbits: Ōkunoshima."),
        ("Japón tiene un festival donde se lanzan habas a demonios imaginarios.", "Japan has a festival where you throw beans at imaginary demons."),
        ("Japón tiene más de 50.000 personas centenarias.", "Japan has over 50,000 centenarians."),
    ],
    "Sweden": [
        ("Suecia importa basura de otros países porque se le acabó la suya para reciclar.", "Sweden imports trash from other countries because it ran out of its own to recycle."),
        ("Hay un hotel de hielo en Suecia que se reconstruye cada invierno.", "There's an ice hotel in Sweden rebuilt every winter."),
        ("Suecia tuvo un rey que murió por comer 14 porciones de semla de una sentada.", "Sweden had a king who died from eating 14 servings of semla in one sitting."),
        ("En verano el sol no se pone en el norte de Suecia: hay sol de medianoche.", "In summer the sun doesn't set in northern Sweden — there's midnight sun."),
    ],
    "Tunisia": [
        ("Túnez fue donde se rodó Tatooine en Star Wars. Los decorados siguen ahí.", "Tunisia is where Tatooine was filmed in Star Wars. The sets are still there."),
        ("Cartago (en Túnez) tenía un puerto secreto para 220 barcos de guerra.", "Carthage (in Tunisia) had a secret harbor for 220 warships."),
        ("Túnez tiene un lago rosa natural: el Lago de Túnez.", "Tunisia has a natural pink lake: the Lake of Tunis."),
    ],
    "Belgium": [
        ("Bélgica estuvo 589 días sin gobierno (2010-2011). Nadie lo notó.", "Belgium went 589 days without a government (2010-2011). Nobody noticed."),
        ("Las patatas fritas son belgas, no francesas.", "French fries are Belgian, not French."),
        ("Bélgica tiene más cómics per cápita que cualquier otro país.", "Belgium has more comic books per capita than any other country."),
        ("Bélgica produce más de 220.000 toneladas de chocolate al año.", "Belgium produces over 220,000 tonnes of chocolate per year."),
    ],
    "Egypt": [
        ("Las pirámides ya eran antiguas cuando Cleopatra vivió. Ella está más cerca de nosotros que de su construcción.", "The pyramids were already ancient when Cleopatra lived. She's closer to us than to their construction."),
        ("Los egipcios antiguos usaban maquillaje tanto hombres como mujeres.", "Ancient Egyptians wore makeup — both men and women."),
        ("Egipto tiene la presa más grande de África: la Gran Presa del Renacimiento… bueno, esa es de Etiopía. La de Asuán también es enorme.", "Egypt has one of Africa's largest dams: the Aswan High Dam."),
        ("La palabra 'papel' viene del papiro egipcio.", "The word 'paper' comes from Egyptian papyrus."),
    ],
    "Iran": [
        ("Irán tiene más cirugías de nariz per cápita que cualquier país.", "Iran has more nose jobs per capita than any other country."),
        ("El polo fue inventado en Persia (Irán) como entrenamiento de caballería.", "Polo was invented in Persia (Iran) as cavalry training."),
        ("Irán tiene un desierto donde la temperatura alcanza los 70°C en superficie.", "Iran has a desert where surface temperatures reach 70°C."),
    ],
    "New Zealand": [
        ("Nueva Zelanda tiene más ovejas que personas (5:1).", "New Zealand has more sheep than people (5:1)."),
        ("Fue el último país grande en ser habitado por humanos (~1250 d.C.).", "It was the last major landmass settled by humans (~1250 AD)."),
        ("Nueva Zelanda tiene un pájaro que no vuela y pesa 4kg: el kiwi.", "New Zealand has a flightless bird that weighs 4kg: the kiwi."),
    ],
    "Spain": [
        ("España tiene un festival donde se lanzan tomates: La Tomatina.", "Spain has a festival where people throw tomatoes at each other: La Tomatina."),
        ("La selección española de fútbol no ganó nada importante durante 44 años… y luego ganó 3 seguidos.", "Spain's football team won nothing major for 44 years… then won 3 in a row."),
        ("España tiene más bares per cápita que cualquier país de la UE.", "Spain has more bars per capita than any EU country."),
        ("España es el segundo país más visitado del mundo (~85 millones de turistas al año).", "Spain is the world's second most visited country (~85 million tourists a year)."),
    ],
    "Cape Verde": [
        ("Cabo Verde tiene más gente viviendo fuera del país que dentro.", "Cape Verde has more people living abroad than at home."),
        ("Es el lugar de cría más importante del mundo para tortugas marinas.", "It's the world's most important breeding site for sea turtles."),
        ("Cabo Verde no tenía población humana hasta que los portugueses llegaron en 1456.", "Cape Verde had no human population until the Portuguese arrived in 1456."),
        ("El morna caboverdiano inspiró la canción 'Sodade', famosa en todo el mundo.", "Cape Verdean morna music inspired the world-famous song 'Sodade'."),
    ],
    "Uruguay": [
        ("Uruguay tiene más vacas que personas (4:1).", "Uruguay has more cows than people (4:1)."),
        ("Uruguay fue el primer país del mundo en legalizar el cannabis a nivel nacional.", "Uruguay was the first country to legalize cannabis nationally."),
        ("El himno nacional de Uruguay dura 5 minutos. Es el más largo del mundo.", "Uruguay's national anthem lasts 5 minutes. It's the world's longest."),
    ],
    "Saudi Arabia": [
        ("Arabia Saudí importa camellos de Australia.", "Saudi Arabia imports camels from Australia."),
        ("No hay ríos permanentes en Arabia Saudí.", "There are no permanent rivers in Saudi Arabia."),
        ("Arabia Saudí tiene un edificio (Jeddah Tower) diseñado para medir 1km de alto. Aún sin terminar.", "Saudi Arabia has a building (Jeddah Tower) designed to be 1km tall. Still unfinished."),
    ],
    "France": [
        ("Francia tiene 12 husos horarios (por sus territorios de ultramar). Más que cualquier país.", "France has 12 time zones (due to overseas territories). More than any country."),
        ("Francia construyó un 'Versalles falso' para engañar a los bombarderos alemanes en la IIGM.", "France built a 'fake Versailles' to fool German bombers in WWII."),
        ("El croissant no es francés. Es austríaco.", "The croissant isn't French. It's Austrian."),
        ("Francia produce más de 1.600 tipos de queso diferentes.", "France produces over 1,600 different types of cheese."),
    ],
    "Senegal": [
        ("Senegal tiene un lago rosa natural: el Lago Retba.", "Senegal has a natural pink lake: Lake Retba."),
        ("El Rally Dakar originalmente terminaba en Dakar (antes de mudarse a Sudamérica).", "The Dakar Rally originally ended in Dakar (before moving to South America)."),
        ("Senegal tiene la estatua más alta de África: el Monumento al Renacimiento Africano (49m).", "Senegal has Africa's tallest statue: the African Renaissance Monument (49m)."),
        ("La isla de Gorée, frente a Dakar, fue declarada Patrimonio de la Humanidad por la UNESCO.", "Gorée Island, off Dakar, was declared a UNESCO World Heritage Site."),
    ],
    "Iraq": [
        ("Irak es donde se inventó la escritura (~3400 a.C.).", "Iraq is where writing was invented (~3400 BC)."),
        ("Las ruinas de Babilonia están a 85km de Bagdad.", "The ruins of Babylon are 85km from Baghdad."),
        ("Irak tiene los jardines colgantes de Babilonia, una de las 7 maravillas antiguas. Nunca se han encontrado.", "Iraq had the Hanging Gardens of Babylon, one of the 7 ancient wonders. They've never been found."),
    ],
    "Norway": [
        ("El rey de Noruega es un pingüino. Literalmente. Se llama Nils Olav.", "The King of Norway is a penguin. Literally. His name is Nils Olav."),
        ("Noruega regaló un árbol de Navidad a Londres cada año desde 1947. Como agradecimiento por la WWII.", "Norway has gifted a Christmas tree to London every year since 1947. As thanks for WWII."),
        ("Noruega tiene un pueblo fantasma pirámide en el Ártico: Pyramiden.", "Norway has a ghost town pyramid in the Arctic: Pyramiden."),
        ("En verano el sol no se pone en el norte de Noruega: hay sol de medianoche.", "In summer the sun doesn't set in northern Norway — there's midnight sun."),
    ],
    "Argentina": [
        ("Argentina tuvo 5 presidentes en 10 días en 2001.", "Argentina had 5 presidents in 10 days in 2001."),
        ("El tango nació en los burdeles de Buenos Aires.", "Tango was born in the brothels of Buenos Aires."),
        ("Argentina tiene la avenida más ancha del mundo: la 9 de Julio (14 carriles).", "Argentina has the world's widest avenue: 9 de Julio (14 lanes)."),
        ("Argentina es el mayor consumidor mundial de carne de vacuno per cápita.", "Argentina is the world's largest per-capita consumer of beef."),
    ],
    "Algeria": [
        ("Argelia es el país más grande de África por superficie.", "Algeria is the largest country in Africa by area."),
        ("Argelia tiene más de 1.000 km de costa mediterránea.", "Algeria has over 1,000 km of Mediterranean coastline."),
        ("El desierto del Sahara cubre el 80% de Argelia.", "The Sahara Desert covers 80% of Algeria."),
        ("Argelia tiene 7 lugares declarados Patrimonio de la Humanidad por la UNESCO.", "Algeria has 7 UNESCO World Heritage Sites."),
    ],
    "Austria": [
        ("Austria tiene un pueblo llamado 'Fucking'. Tuvo que cambiar el nombre por los turistas.", "Austria had a village called 'Fucking'. It had to change the name because of tourists."),
        ("El 80% de la banda sonora de 'The Sound of Music' se rodó en Austria.", "80% of 'The Sound of Music' was filmed in Austria."),
        ("Austria inventó la bola de nieve (Schneekugel) en 1900.", "Austria invented the snow globe (Schneekugel) in 1900."),
        ("Viena ha sido elegida repetidamente la ciudad más habitable del mundo.", "Vienna has repeatedly been ranked the world's most liveable city."),
    ],
    "Jordan": [
        ("Jordania tiene el punto más bajo de la Tierra: el Mar Muerto (-430m).", "Jordan has the lowest point on Earth: the Dead Sea (-430m)."),
        ("Petra (en Jordania) fue tallada en roca rosa hace 2.000 años.", "Petra (in Jordan) was carved into pink rock 2,000 years ago."),
        ("Jordania tiene un castillo en medio del desierto que parece una nave espacial: Qasr Kharana.", "Jordan has a desert castle that looks like a spaceship: Qasr Kharana."),
    ],
    "Portugal": [
        ("Portugal tiene la alianza diplomática más antigua del mundo con Inglaterra (1373).", "Portugal has the world's oldest diplomatic alliance with England (1373)."),
        ("Portugal es el país más antiguo de Europa con las mismas fronteras desde 1139.", "Portugal is Europe's oldest country with the same borders since 1139."),
        ("Lisboa es más antigua que Roma.", "Lisbon is older than Rome."),
        ("El fado fue declarado Patrimonio Inmaterial de la Humanidad por la UNESCO.", "Fado music was declared UNESCO Intangible Cultural Heritage."),
    ],
    "DR Congo": [
        ("La RD Congo tiene el 50% de las reservas mundiales de cobalto.", "DR Congo has 50% of the world's cobalt reserves."),
        ("El río Congo es el más profundo del mundo (220m).", "The Congo River is the world's deepest (220m)."),
        ("La RD Congo tiene un volcán que produce lava azul: Nyiragongo.", "DR Congo has a volcano that produces blue lava: Nyiragongo."),
        ("El río Congo es el único gran río del mundo que cruza el ecuador dos veces.", "The Congo River is the only major river in the world that crosses the equator twice."),
    ],
    "Uzbekistan": [
        ("Uzbekistán tiene una de las ciudades más antiguas de la Ruta de la Seda: Samarcanda.", "Uzbekistan has one of the oldest Silk Road cities: Samarkand."),
        ("El Mar de Aral (en Uzbekistán) era el 4º lago más grande del mundo. Ahora es un desierto.", "The Aral Sea (in Uzbekistan) was the world's 4th largest lake. Now it's a desert."),
        ("Uzbekistán tiene el metro más bonito de Asia Central. Cada estación es una obra de arte.", "Uzbekistan has Central Asia's most beautiful metro. Each station is a work of art."),
    ],
    "Colombia": [
        ("Colombia tiene más especies de aves que cualquier otro país (~1.900).", "Colombia has more bird species than any other country (~1,900)."),
        ("El río Caño Cristales en Colombia tiene 5 colores y le llaman 'el río que se escapó del paraíso'.", "Colombia's Caño Cristales river has 5 colors and is called 'the river that escaped paradise'."),
        ("Colombia es el único país de Sudamérica con costa en el Pacífico Y el Caribe.", "Colombia is the only South American country with both Pacific and Caribbean coasts."),
        ("Bogotá está a 2.640 m de altitud: una de las capitales más altas del mundo.", "Bogotá sits at 2,640 m altitude — one of the world's highest capitals."),
    ],
    "England": [
        ("Inglaterra tuvo una reina que reinó solo 9 días: Lady Jane Grey.", "England had a queen who reigned just 9 days: Lady Jane Grey."),
        ("El Big Ben no es el nombre de la torre. Es el nombre de la campana.", "Big Ben isn't the tower's name. It's the bell's name."),
        ("Inglaterra inventó el fútbol moderno… y luego no ganó un Mundial durante 56 años.", "England invented modern football… then didn't win a World Cup for 56 years."),
        ("Los ingleses beben unos 165 millones de tazas de té al día.", "The English drink about 165 million cups of tea per day."),
    ],
    "Croatia": [
        ("La corbata fue inventada por mercenarios croatas en el siglo XVII.", "The necktie was invented by Croatian mercenaries in the 17th century."),
        ("Croacia tiene el museo de relaciones rotas del mundo (Zagreb).", "Croatia has the world's Museum of Broken Relationships (Zagreb)."),
        ("Dálmata (el perro) viene de Dalmacia, Croacia.", "The Dalmatian dog comes from Dalmatia, Croatia."),
        ("Croacia tiene más de 1.200 islas, pero solo unas 50 están habitadas.", "Croatia has over 1,200 islands, but only about 50 are inhabited."),
    ],
    "Ghana": [
        ("Ghana tiene el ataúd más customizado del mundo: los hacen con forma de pez, avión, Coca-Cola…", "Ghana has the world's most custom coffins: shaped like fish, planes, Coca-Cola bottles…"),
        ("Ghana fue llamada 'Costa de Oro' por los colonizadores portugueses.", "Ghana was called the 'Gold Coast' by Portuguese colonizers."),
        ("El lago Volta en Ghana es el mayor lago artificial del mundo por superficie.", "Lake Volta in Ghana is the world's largest artificial lake by surface area."),
        ("Ghana fue el primer país subsahariano en independizarse de Europa colonial (1957).", "Ghana was the first sub-Saharan country to gain independence from colonial Europe (1957)."),
    ],
    "Panama": [
        ("Panamá es el único lugar del mundo donde puedes ver el amanecer en el Pacífico y el atardecer en el Atlántico.", "Panama is the only place where you can see sunrise on the Pacific and sunset on the Atlantic."),
        ("El Canal de Panamá mueve el 6% del comercio mundial.", "The Panama Canal moves 6% of world trade."),
        ("Panamá tiene más de 1.500 islas. La mayoría sin nombre.", "Panama has over 1,500 islands. Most unnamed."),
    ],
}

# Índice del fact de TRIVIA por ronda eliminatoria (0-based). R32 usa el 4.º dato.
KO_TRIVIA_INDEX = {"r32": 3}


def match_trivia(team_en, fact_idx):
    facts = TRIVIA.get(team_en, [])
    if fact_idx < len(facts):
        es, en = facts[fact_idx]
    else:
        es, en = "", ""
    return {"es": es, "en": en}


def team_es(name):
    if name is None:
        return None
    name = _team_key(name)
    return TEAMS.get(name, (name, "🏳️"))[0]


def team_flag(name):
    if name is None:
        return "🏳️"
    return TEAMS.get(_team_key(name), (name, "🏳️"))[1]


def _team_key(name):
    name = str(name).strip()
    if name in ES2EN:
        return ES2EN[name]
    return FIFA_TEAM_ALIASES.get(name, name)


def knockout_schedule_info(code):
    item = KNOCKOUT_MATCH_SCHEDULE.get(code)
    if not item:
        return {
            "fixture_home": "",
            "fixture_away": "",
            "fixture_home_flag": "",
            "fixture_away_flag": "",
            "date": "",
            "time_es": "",
            "time_uk": "",
            "next_es": False,
            "next_uk": False,
            "dt": "",
            "venue": "",
            "city": "",
        }

    base = datetime.fromisoformat(item["kickoff_et"])
    home = _fixture_team(item["home"])
    away = _fixture_team(item["away"])
    local_es = base + _ET_OFFSETS["es"]
    # Partidos de madrugada (antes de las 6h) siguen contando como la jornada
    # anterior, igual que matchdayDateStr() en el JS del dashboard.
    matchday_date = local_es.date()
    if local_es.hour < 6:
        matchday_date -= timedelta(days=1)
    out = {
        "fixture_home": home["name"],
        "fixture_away": away["name"],
        "fixture_home_en": home["name_en"],
        "fixture_away_en": away["name_en"],
        "fixture_home_flag": home["flag"],
        "fixture_away_flag": away["flag"],
        "date": matchday_date.isoformat(),
        "dt": local_es.isoformat(),
        "venue": item["venue"],
        "city": item["city"],
    }
    for lang, off in _ET_OFFSETS.items():
        local = base + off
        out[f"time_{lang}"] = local.strftime("%H:%M")
        out[f"next_{lang}"] = local.date().isoformat() != out["date"]
    return out


def _fixture_team(name):
    if not name:
        return {"name": "", "name_en": "", "flag": ""}
    if name.startswith(("W", "RU")):
        return {"name": name, "name_en": name, "flag": ""}
    canonical = FIFA_TEAM_ALIASES.get(name, name)
    return {"name": team_es(canonical), "name_en": canonical, "flag": team_flag(canonical)}


# --------------------------------------------------------------------------
# Parsing
# --------------------------------------------------------------------------

def parse_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    raw = wb["Raw data"]

    # Localiza columnas de participantes con nombre en NAME_ROW
    participants = []  # [{name, home_col, away_col, idx}]
    for slot in range(N_PARTICIPANTS_MAX):
        home_col = RAW_FIRST_HOME_COL + 2 * slot
        name = raw.cell(NAME_ROW, home_col).value
        if name is not None and str(name).strip():
            participants.append({
                "name": str(name).strip(),
                "home_col": home_col,
                "away_col": home_col + 1,
                "idx": len(participants),
            })
    names = [p["name"] for p in participants]
    n = len(participants)

    # Partidos de fase de grupos
    matches = []
    for r in GROUP_MATCH_ROWS:
        concept = raw.cell(r, 2).value
        if not concept or "vs" not in str(concept):
            continue
        concept = str(concept).strip()
        code, _, teams = concept.partition(":")
        code = code.strip()
        home_en, _, away_en = teams.strip().partition(" vs ")
        home_en, away_en = home_en.strip(), away_en.strip()
        group = code.split("-")[0][1:]  # "GA-M1" -> "A", "GG-M3" -> "G"
        picks = []
        for p in participants:
            hg = raw.cell(r, p["home_col"]).value
            ag = raw.cell(r, p["away_col"]).value
            picks.append((_num(hg), _num(ag)))
        matches.append({
            "code": code, "group": group,
            "home_en": home_en, "away_en": away_en,
            "home": team_es(home_en), "away": team_es(away_en),
            "home_flag": team_flag(home_en), "away_flag": team_flag(away_en),
            "picks": picks,
            "date": MATCH_DATES.get(code, ""),
            **kickoff_info(code),
        })

    # Clasificados por grupo (1/2/3)
    qualifiers = defaultdict(dict)  # group -> {pos -> [team per participant]}
    for r in QUALIFIER_ROWS:
        label = raw.cell(r, 2).value
        if not label:
            continue
        label = str(label).strip().lower()
        pos = None
        if label.startswith("1st"):
            pos = 1
        elif label.startswith("2nd"):
            pos = 2
        elif label.startswith("3rd"):
            pos = 3
        if pos is None:
            continue
        group = label.split("group")[-1].strip().upper()
        picks = [raw.cell(r, p["home_col"]).value for p in participants]
        qualifiers[group][pos] = picks

    # 8 mejores terceros
    thirds = []  # list per row -> [team per participant]
    for r in THIRDS_ROWS:
        label = raw.cell(r, 2).value
        if not label:
            continue
        picks = [raw.cell(r, p["home_col"]).value for p in participants]
        thirds.append(picks)

    knockouts = parse_knockouts(raw, participants)
    knockout_results = parse_knockout_results(wb)

    # ¿Hay resultados reales?
    results = parse_results(wb)
    standings_results, thirds_results = parse_standings_results(wb)

    return {
        "participants": participants,
        "names": names,
        "n": n,
        "matches": matches,
        "qualifiers": qualifiers,
        "thirds": thirds,
        "results": results,
        "standings_results": standings_results,
        "thirds_results": thirds_results,
        "knockouts": knockouts,
        "knockout_results": knockout_results,
    }


def parse_results(wb):
    """Lee resultados reales de fase de grupos si existen. Devuelve {code:(h,a)}."""
    rr = wb["Real results"]
    out = {}
    for r in range(1, rr.max_row + 1):
        concept = rr.cell(r, 2).value
        if not concept or "vs" not in str(concept):
            continue
        code = str(concept).split(":")[0].strip()
        h = _num(rr.cell(r, 3).value)
        a = _num(rr.cell(r, 4).value)
        if h is not None and a is not None:
            out[code] = (h, a)
    return out


def parse_standings_results(wb):
    """Lee los clasificados reales por grupo (1.º/2.º/3.º) y los 8 mejores
    terceros de la pestaña 'Real results', si están rellenos.

    Devuelve (standings, thirds):
      standings -> {grupo: {1: equipo, 2: equipo, 3: equipo}}
      thirds    -> [equipo, ...]  (los terceros clasificados)
    """
    rr = wb["Real results"]
    standings = {}
    for r in QUALIFIER_ROWS:
        label = rr.cell(r, 2).value
        if not label:
            continue
        label = str(label).strip().lower()
        if label.startswith("1st"):
            pos = 1
        elif label.startswith("2nd"):
            pos = 2
        elif label.startswith("3rd"):
            pos = 3
        else:
            continue
        group = label.split("group")[-1].strip().upper()
        team = _text(rr.cell(r, 3).value)
        if team:
            standings.setdefault(group, {})[pos] = team
    thirds = []
    for r in THIRDS_ROWS:
        team = _text(rr.cell(r, 3).value)
        if team:
            thirds.append(team)
    return standings, thirds


def parse_knockouts(raw, participants):
    rounds = []
    for rnd in KNOCKOUT_ROUNDS:
        matches = []
        for i in range(rnd["matches"]):
            score_row = rnd["first_row"] + i * 3
            penalty_row = score_row + 1
            winner_row = score_row + 2
            code = f"{rnd['code']}-M{i + 1}"
            schedule = knockout_schedule_info(code)
            score_picks = [
                (_num(raw.cell(score_row, p["home_col"]).value),
                 _num(raw.cell(score_row, p["away_col"]).value))
                for p in participants
            ]
            tiebreaker_picks = [
                _text(raw.cell(penalty_row, p["home_col"]).value)
                for p in participants
            ]
            raw_winner_picks = [
                _text(raw.cell(winner_row, p["home_col"]).value)
                for p in participants
            ]
            matches.append({
                "code": code,
                "score_row": score_row,
                "penalty_row": penalty_row,
                "winner_row": winner_row,
                "score_picks": score_picks,
                "penalty_picks": tiebreaker_picks,
                "winner_picks": [
                    _knockout_winner_pick(schedule, score, tiebreaker, raw_winner)
                    for score, tiebreaker, raw_winner in zip(score_picks, tiebreaker_picks, raw_winner_picks)
                ],
                **schedule,
            })
        rounds.append({**rnd, "matches": matches})

    final_matches = []
    for match in FINAL_MATCHES:
        final_matches.append({
            **match,
            **knockout_schedule_info(match["code"]),
            "score_picks": [
                (_num(raw.cell(match["score_row"], p["home_col"]).value),
                 _num(raw.cell(match["score_row"], p["away_col"]).value))
                for p in participants
            ],
            "penalty_picks": [
                _text(raw.cell(match["penalty_row"], p["home_col"]).value)
                for p in participants
            ],
        })

    outright = {}
    for key, meta in SPECIAL_OUTRIGHT_ROWS.items():
        outright[key] = {
            **meta,
            "picks": [_text(raw.cell(meta["row"], p["home_col"]).value) for p in participants],
        }

    awards = {}
    for key, meta in AWARD_ROWS.items():
        awards[key] = {
            **meta,
            "picks": [_text(raw.cell(meta["row"], p["home_col"]).value) for p in participants],
        }

    return {"rounds": rounds, "final_matches": final_matches, "outright": outright, "awards": awards}


def _knockout_winner_pick(schedule, score, tiebreaker, raw_winner):
    home = schedule.get("fixture_home_en") or schedule.get("fixture_home")
    away = schedule.get("fixture_away_en") or schedule.get("fixture_away")
    if not home or not away or home.startswith(("W", "RU")) or away.startswith(("W", "RU")):
        return raw_winner

    h, a = score
    if h is None or a is None:
        return raw_winner
    if h > a:
        return home
    if a > h:
        return away
    if tiebreaker:
        side = _cmp_text(tiebreaker)
        if side == "home":
            return home
        if side == "away":
            return away
        return tiebreaker
    return raw_winner


def parse_knockout_results(wb):
    rr = wb["Real results"]
    out = {"matches": {}, "outright": {}, "awards": {}}

    for rnd in KNOCKOUT_ROUNDS:
        for i in range(rnd["matches"]):
            score_row = rnd["first_row"] + i * 3
            code = f"{rnd['code']}-M{i + 1}"
            entry = _knockout_result_entry(rr, score_row, score_row + 1, score_row + 2)
            if entry:
                out["matches"][code] = entry

    for match in FINAL_MATCHES:
        entry = _knockout_result_entry(rr, match["score_row"], match["penalty_row"])
        if entry:
            out["matches"][match["code"]] = entry

    for key, meta in SPECIAL_OUTRIGHT_ROWS.items():
        value = _text(rr.cell(meta["row"], 3).value)
        if value:
            out["outright"][key] = value

    for key, meta in AWARD_ROWS.items():
        value = _text(rr.cell(meta["row"], 3).value)
        if value:
            out["awards"][key] = value

    return out


def _knockout_result_entry(sheet, score_row, penalty_row=None, winner_row=None):
    entry = {}
    h = _num(sheet.cell(score_row, 3).value)
    a = _num(sheet.cell(score_row, 4).value)
    if h is not None and a is not None:
        entry["score"] = (h, a)
    if penalty_row is not None:
        penalties = _text(sheet.cell(penalty_row, 3).value)
        if penalties:
            entry["penalties"] = penalties
    if winner_row is not None:
        winner = _text(sheet.cell(winner_row, 3).value)
        if winner:
            entry["winner"] = winner
    return entry


def _text(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _num(v):
    if v is None or v == "":
        return None
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return None


# --------------------------------------------------------------------------
# Utilidades de analítica
# --------------------------------------------------------------------------

def outcome(hg, ag):
    if hg is None or ag is None:
        return None
    if hg > ag:
        return "1"
    if hg < ag:
        return "2"
    return "X"


def minmax_scale(values):
    lo, hi = min(values), max(values)
    if hi - lo < 1e-9:
        return [50.0 for _ in values]
    return [100.0 * (v - lo) / (hi - lo) for v in values]


def zscore(values):
    mu = statistics.mean(values)
    sd = statistics.pstdev(values) or 1.0
    return [(v - mu) / sd for v in values]


# --------------------------------------------------------------------------
# Motor de analíticas
# --------------------------------------------------------------------------

def compute(data):
    names = data["names"]
    n = data["n"]
    matches = data["matches"]
    nm = len(matches)

    # ---- Índice de Rebeldía (signo minoría + distancia marcador) ----
    raw_sign = [0.0] * n
    raw_score = [0.0] * n
    match_meta = []  # por partido: agreement, mediana, etc.

    for m in matches:
        picks = m["picks"]
        valid = [(i, h, a) for i, (h, a) in enumerate(picks) if h is not None and a is not None]
        if not valid:
            match_meta.append(None)
            continue
        outcomes = [outcome(h, a) for _, h, a in valid]
        oc = Counter(outcomes)
        total = len(valid)
        homes = [h for _, h, a in valid]
        aways = [a for _, h, a in valid]
        med_h = statistics.median(homes)
        med_a = statistics.median(aways)
        # distancia de marcador por persona
        sdist = {}
        for i, h, a in valid:
            sdist[i] = abs(h - med_h) + abs(a - med_a)
        max_sdist = max(sdist.values()) or 1.0
        for i, h, a in valid:
            o = outcome(h, a)
            share = oc[o] / total
            raw_sign[i] += (1.0 - share)
            raw_score[i] += sdist[i] / max_sdist
        # scoreline modal
        scl = Counter((h, a) for _, h, a in valid)
        modal_scl, modal_cnt = scl.most_common(1)[0]
        top_outcome, top_cnt = oc.most_common(1)[0]
        match_meta.append({
            "code": m["code"], "group": m["group"],
            "home": m["home"], "away": m["away"],
            "home_en": m["home_en"], "away_en": m["away_en"],
            "home_flag": m["home_flag"], "away_flag": m["away_flag"],
            "outcome_dist": dict(oc),
            "outcome_agreement": top_cnt / total,
            "top_outcome": top_outcome,
            "modal_scoreline": f"{modal_scl[0]}-{modal_scl[1]}",
            "modal_scoreline_share": modal_cnt / total,
            "median": f"{int(med_h)}-{int(med_a)}",
            "n": total,
        })

    raw_sign = [s / nm for s in raw_sign]
    raw_score = [s / nm for s in raw_score]
    z = [0.5 * a + 0.5 * b for a, b in zip(zscore(raw_sign), zscore(raw_score))]
    rebel_index = [round(v, 1) for v in minmax_scale(z)]

    rebeldia = sorted(
        [{"name": names[i], "index": rebel_index[i],
          "sign": round(raw_sign[i], 3), "score": round(raw_score[i], 3)}
         for i in range(n)],
        key=lambda x: -x["index"],
    )

    # ---- Similitud: % marcadores idénticos + distancia media de goles ----
    exact = [[0] * n for _ in range(n)]
    gdist = [[0.0] * n for _ in range(n)]
    comparable = 0
    for m in matches:
        picks = m["picks"]
        if all(h is not None for h, a in picks):
            comparable += 1
        for i in range(n):
            hi, ai = picks[i]
            if hi is None:
                continue
            for j in range(i + 1, n):
                hj, aj = picks[j]
                if hj is None:
                    continue
                if hi == hj and ai == aj:
                    exact[i][j] += 1
                    exact[j][i] += 1
                d = abs(hi - hj) + abs(ai - aj)
                gdist[i][j] += d
                gdist[j][i] += d

    sim_pct = [[0.0] * n for _ in range(n)]
    for i in range(n):
        for j in range(n):
            if i != j and comparable:
                sim_pct[i][j] = round(100.0 * exact[i][j] / comparable, 1)

    # almas gemelas (máx % idéntico) y polos opuestos (máx distancia)
    best_pair = (None, None, -1)
    worst_pair = (None, None, -1)
    for i in range(n):
        for j in range(i + 1, n):
            if sim_pct[i][j] > best_pair[2]:
                best_pair = (i, j, sim_pct[i][j])
            avg_d = gdist[i][j] / (comparable or 1)
            if avg_d > worst_pair[2]:
                worst_pair = (i, j, round(avg_d, 2))

    # gemelo de cada uno
    twins = []
    for i in range(n):
        best_j, best_v = None, -1
        for j in range(n):
            if i != j and sim_pct[i][j] > best_v:
                best_v, best_j = sim_pct[i][j], j
        twins.append({"name": names[i], "twin": names[best_j], "pct": best_v})

    # ---- Estilo de apuesta ----
    style = []
    all_scorelines = Counter()
    for i in range(n):
        goals, draws, blowouts, biggest = [], 0, 0, (None, -1)
        for m in matches:
            h, a = m["picks"][i]
            if h is None:
                continue
            goals.append(h + a)
            if h == a:
                draws += 1
            if abs(h - a) >= 3:
                blowouts += 1
            if h + a > biggest[1]:
                biggest = ((m, h, a), h + a)
            all_scorelines[(h, a)] += 1
        cnt = len(goals) or 1
        big = None
        if biggest[0]:
            mm, h, a = biggest[0]
            big = {"home": mm["home"], "away": mm["away"],
                   "home_en": mm["home_en"], "away_en": mm["away_en"],
                   "score": f"{h}-{a}",
                   "flags": f"{mm['home_flag']}{mm['away_flag']}", "goals": h + a}
        style.append({
            "name": names[i],
            "avg_goals": round(sum(goals) / cnt, 2),
            "pct_draws": round(100 * draws / cnt, 1),
            "pct_blowouts": round(100 * blowouts / cnt, 1),
            "biggest": big,
        })

    common_scorelines = [
        {"score": f"{h}-{a}", "count": c,
         "pct": round(100 * c / (nm * n), 1)}
        for (h, a), c in all_scorelines.most_common(8)
    ]

    # ---- Lobo solitario: marcadores únicos (solo 1 persona) ----
    unique_count = [0] * n
    unique_examples = []  # picks únicos llamativos
    for m in matches:
        scl = Counter()
        for h, a in m["picks"]:
            if h is not None:
                scl[(h, a)] += 1
        for i, (h, a) in enumerate(m["picks"]):
            if h is not None and scl[(h, a)] == 1:
                unique_count[i] += 1
                diff = abs(h - a) + (h + a)
                unique_examples.append({
                    "name": names[i], "home": m["home"], "away": m["away"],
                    "home_en": m["home_en"], "away_en": m["away_en"],
                    "flags": f"{m['home_flag']}{m['away_flag']}",
                    "score": f"{h}-{a}", "spice": diff,
                })
    lobo = sorted(
        [{"name": names[i], "count": unique_count[i]} for i in range(n)],
        key=lambda x: -x["count"],
    )
    unique_examples.sort(key=lambda x: -x["spice"])

    # ---- Favoritos del torneo ----
    group_consensus = []
    respect = Counter()        # prestigio ponderado: 1º=3, 2º=2, 3º=1
    pos_weight = {1: 3, 2: 2, 3: 1}
    for group in sorted(data["qualifiers"].keys()):
        pos = data["qualifiers"][group]
        firsts = [team_es(t) for t in pos.get(1, []) if t]
        c = Counter(firsts)
        for p, slot_picks in pos.items():
            w = pos_weight.get(p, 1)
            for t in slot_picks:
                if t:
                    respect[team_es(t)] += w
        if c:
            top, cnt = c.most_common(1)[0]
            group_consensus.append({
                "group": group, "favorite": top,
                "flag": _flag_es(top),
                "agreement": round(100 * cnt / sum(c.values()), 1),
                "dist": dict(c.most_common()),
            })
    most_respected = [{"team": t, "flag": _flag_es(t), "count": c}
                      for t, c in respect.most_common(12)]
    least_respected = [{"team": t, "flag": _flag_es(t), "count": c}
                       for t, c in sorted(respect.items(), key=lambda x: x[1])[:8]]
    polarizing = sorted(group_consensus, key=lambda x: x["agreement"])[:5]

    # terceros más elegidos
    thirds_counter = Counter()
    for row in data["thirds"]:
        for t in row:
            if t:
                thirds_counter[team_es(t)] += 1
    top_thirds = [{"team": t, "flag": _flag_es(t), "count": c}
                  for t, c in thirds_counter.most_common(10)]

    # ---- Partidos divisivos / unánimes ----
    valid_meta = [m for m in match_meta if m]
    divisive = sorted(valid_meta, key=lambda m: (m["outcome_agreement"], m["modal_scoreline_share"]))[:6]
    unanimous = sorted(valid_meta, key=lambda m: (-m["modal_scoreline_share"], -m["outcome_agreement"]))[:6]

    # ---- Fichas por persona ----
    rank_of = {r["name"]: i + 1 for i, r in enumerate(rebeldia)}
    idx_of = {r["name"]: r["index"] for r in rebeldia}
    style_of = {s["name"]: s for s in style}
    twin_of = {t["name"]: t for t in twins}
    lobo_of = {l["name"]: l["count"] for l in lobo}
    cards = []
    for i, name in enumerate(names):
        st = style_of[name]
        cards.append({
            "name": name,
            "rebel_rank": rank_of[name],
            "rebel_index": idx_of[name],
            "avg_goals": st["avg_goals"],
            "pct_draws": st["pct_draws"],
            "biggest": st["biggest"],
            "twin": twin_of[name]["twin"],
            "twin_pct": twin_of[name]["pct"],
            "lobo": lobo_of[name],
            "label": style_label(st, idx_of[name]),
        })

    # ---- Premios ----
    goleador = max(style, key=lambda s: s["avg_goals"])
    cerrojo = min(style, key=lambda s: s["avg_goals"])
    empate = max(style, key=lambda s: s["pct_draws"])
    awards = {
        "rebelde": rebeldia[0],
        "borrego": rebeldia[-1],
        "gemelas": {"a": names[best_pair[0]], "b": names[best_pair[1]], "pct": best_pair[2]},
        "opuestos": {"a": names[worst_pair[0]], "b": names[worst_pair[1]], "dist": worst_pair[2]},
        "goleador": {"name": goleador["name"], "avg": goleador["avg_goals"]},
        "cerrojo": {"name": cerrojo["name"], "avg": cerrojo["avg_goals"]},
        "empate": {"name": empate["name"], "pct": empate["pct_draws"]},
        "lobo": lobo[0],
    }

    # ---- Stats del hero ----
    total_goals = sum(
        h + a for m in matches for (h, a) in m["picks"] if h is not None
    )
    hero = {
        "participants": n,
        "matches": nm,
        "groups": len(data["qualifiers"]),
        "total_goals": total_goals,
        "avg_goals_match": round(total_goals / (nm * n), 2) if nm and n else 0,
        "top_scoreline": common_scorelines[0] if common_scorelines else None,
        "has_results": bool(data["results"]),
    }

    # ---- Aciertos en directo (solo si hay resultados) ----
    live = None
    if data["results"] or any(data["knockout_results"].values()):
        live = compute_live(data, matches)

    today = compute_today(data, matches)
    if live:
        for m in today["matches"]:
            m["stake"] = stake_for_group_match(m, data, matches, live["table"])
    recent_results = compute_recent_results(data, matches)
    knockout = compute_knockout(data, live["table"] if live else None, matches)

    return {
        "es2en": ES2EN,
        "hero": hero,
        "rebeldia": rebeldia,
        "matrix": {"names": names, "sim": sim_pct},
        "twins": twins,
        "best_pair": {"a": names[best_pair[0]], "b": names[best_pair[1]], "pct": best_pair[2]},
        "worst_pair": {"a": names[worst_pair[0]], "b": names[worst_pair[1]], "dist": worst_pair[2]},
        "style": style,
        "common_scorelines": common_scorelines,
        "lobo": lobo,
        "unique_examples": unique_examples[:12],
        "group_consensus": group_consensus,
        "most_respected": most_respected,
        "least_respected": least_respected,
        "polarizing": polarizing,
        "top_thirds": top_thirds,
        "divisive": divisive,
        "unanimous": unanimous,
        "cards": cards,
        "awards": awards,
        "live": live,
        "today": today,
        "recent_results": recent_results,
        "knockout": knockout,
    }


def _flag_es(es_name):
    for en, (es, fl) in TEAMS.items():
        if es == es_name:
            return fl
    return "🏳️"


def style_label(st, rebel_index):
    """Devuelve una clave; el texto/emoji se traduce en el cliente (JS)."""
    if st["avg_goals"] >= 3.0:
        return "goleador"
    if st["avg_goals"] <= 2.0:
        return "cerrojo"
    if st["pct_draws"] >= 35:
        return "empate"
    if rebel_index >= 70:
        return "rebelde"
    if rebel_index <= 25:
        return "borrego"
    return "equilibrado"


def ko_w_number(code):
    m = re.match(r"^(R32|R16|QF|SF)-M(\d+)$", code or "")
    if not m:
        return None
    return KO_WBASE[m.group(1)] + int(m.group(2))


def build_ko_winner_map(knockout_results, knockouts_raw):
    """Map W-numbers to real winners as results land."""
    winners = {}
    kr_matches = knockout_results.get("matches", {})
    for rnd in knockouts_raw["rounds"]:
        for m in rnd["matches"]:
            w = ko_w_number(m["code"])
            result = kr_matches.get(m["code"])
            if w and result and result.get("winner"):
                key = _team_key(result["winner"])
                winners[w] = {
                    "name": team_es(key),
                    "name_en": key,
                    "flag": team_flag(key),
                }
    return winners


def resolve_knockout_fixture(pub_match, winner_by_w):
    """Expose resolved team names without replacing W## feeder placeholders."""
    for side in ("home", "away"):
        val = pub_match.get(f"fixture_{side}", "")
        wm = re.match(r"^W(\d+)$", val)
        if not wm:
            continue
        team = winner_by_w.get(int(wm.group(1)))
        if team:
            pub_match[f"resolved_{side}"] = team["name"]
            pub_match[f"resolved_{side}_flag"] = team["flag"]


def compute_knockout(data, live_table=None, matches=None):
    names = data["names"]
    n = data["n"]
    raw = data["knockouts"]
    matches = matches or data.get("matches") or []
    winner_by_w = build_ko_winner_map(data["knockout_results"], raw)
    scoring = compute_knockout_scoring(data)

    total_rows = 0
    filled_rows = 0
    rounds = []
    for rnd in raw["rounds"]:
        match_rows = []
        for m in rnd["matches"]:
            score = _score_consensus(m["score_picks"], n)
            winner = _text_consensus(m["winner_picks"], n)
            result = data["knockout_results"]["matches"].get(m["code"])
            filled_rows += (
                _filled_score_picks(m["score_picks"])
                + _filled_text_picks(m["winner_picks"])
            )
            total_rows += n * 2
            pub = {
                "code": m["code"],
                "score": score,
                "winner": winner,
                "result": _knockout_public_result(result),
                **_ko_public_pick_fields(m, names, n),
                **_knockout_public_schedule(m),
            }
            resolve_knockout_fixture(pub, winner_by_w)
            _attach_matchup_ok(pub, m, data)
            trivia_idx = KO_TRIVIA_INDEX.get(rnd["key"])
            if trivia_idx is not None:
                pub["home_trivia"] = match_trivia(pub.get("fixture_home_en"), trivia_idx)
                pub["away_trivia"] = match_trivia(pub.get("fixture_away_en"), trivia_idx)
            if not result or "score" not in result:
                pub["stake"] = stake_for_ko_match(m, rnd, names, data, matches, live_table)
            match_rows.append(pub)
        rounds.append({
            "key": rnd["key"],
            "label_es": rnd["label_es"],
            "label_en": rnd["label_en"],
            "advance_points": rnd["advance_points"],
            "matches": match_rows,
        })

    final_matches = []
    for m in raw["final_matches"]:
        score = _score_consensus(m["score_picks"], n)
        result = data["knockout_results"]["matches"].get(m["code"])
        filled_rows += _filled_score_picks(m["score_picks"])
        total_rows += n
        pub = {
            "key": m["key"],
            "code": m["code"],
            "label_es": m["label_es"],
            "label_en": m["label_en"],
            "score": score,
            "result": _knockout_public_result(result),
            **_ko_public_pick_fields(m, names, n),
            **_knockout_public_schedule(m),
        }
        resolve_knockout_fixture(pub, winner_by_w)
        _attach_matchup_ok(pub, m, data)
        if not result or "score" not in result:
            pub["stake"] = stake_for_ko_match(m, None, names, data, matches, live_table)
        final_matches.append(pub)

    outright = {}
    for key, meta in raw["outright"].items():
        consensus = _text_consensus(meta["picks"], n)
        filled_rows += _filled_text_picks(meta["picks"])
        total_rows += n
        outright[key] = {
            "label_es": meta["label_es"],
            "label_en": meta["label_en"],
            "points": meta["points"],
            **consensus,
        }

    awards = {}
    for key, meta in raw["awards"].items():
        consensus = _text_consensus(meta["picks"], n)
        filled_rows += _filled_text_picks(meta["picks"])
        total_rows += n
        awards[key] = {
            "label_es": meta["label_es"],
            "label_en": meta["label_en"],
            "points": meta["points"],
            **consensus,
        }

    kr = data["knockout_results"]
    results_started = bool(kr["matches"] or kr["outright"] or kr["awards"])
    metrics = compute_knockout_metrics(data) if filled_rows > 0 else None
    progression = compute_ko_progression(data)
    return {
        "ready": filled_rows > 0,
        "filled": filled_rows,
        "total": total_rows,
        "pct": round(100 * filled_rows / total_rows, 1) if total_rows else 0,
        "results_started": results_started,
        "rounds": rounds,
        "final_matches": final_matches,
        "outright": outright,
        "awards": awards,
        "scoring": scoring,
        "metrics": metrics,
        "progression": progression,
    }


def _ko_w_map_from_results(data):
    """W-number -> ganador real (clave canónica) según resultados cargados."""
    w_map = {}
    for rnd in data["knockouts"]["rounds"]:
        for match in rnd["matches"]:
            w = ko_w_number(match["code"])
            result = data["knockout_results"]["matches"].get(match["code"])
            if w and result and result.get("winner"):
                w_map[w] = _cmp_team(result["winner"])
    return w_map


def _ko_w_map_for_person(person_idx, data):
    """Cuadro predicho de una persona: W-number -> ganador que puso (canónico)."""
    w_map = {}
    for rnd in data["knockouts"]["rounds"]:
        for match in rnd["matches"]:
            w = ko_w_number(match["code"])
            pick = match["winner_picks"][person_idx] if match.get("winner_picks") else None
            if w and pick:
                w_map[w] = _cmp_team(pick)
    return w_map


def _ko_matchup_counts(match, data):
    """Máscara por persona: ¿cuenta el marcador (signo/pleno) de este cruce?

    Solo cuenta si acertaste el CRUCE ENTERO: los dos equipos que juegan el
    partido son los que tú predijiste. Si acertaste al ganador pero el rival era
    otro (fallaste un feeder), el marcador no cuenta — pusiste los goles de otro
    partido; solo te llevas el punto del pase (que se otorga aparte).

    R32 tiene equipos reales fijos -> el cruce coincide para todos.
    Cruce sin resolver (feeder por jugar) -> no se condiciona.
    """
    n = data["n"]
    actual_pair = _ko_fixture_pair(match, _ko_w_map_from_results(data))
    if actual_pair is None:
        return [True] * n
    return [
        _ko_fixture_pair(match, _ko_w_map_for_person(i, data)) == actual_pair
        for i in range(n)
    ]


def _attach_matchup_ok(pub, match, data):
    """Marca en cada pick público si el cruce coincide con el que predijo la
    persona (lo usa la web para saber si su marcador puntúa o no)."""
    mask = _ko_matchup_counts(match, data)
    for i, pick in enumerate(pub.get("picks", [])):
        pick["matchup_ok"] = mask[i]


def compute_knockout_scoring(data):
    results = data["knockout_results"]
    if not (results["matches"] or results["outright"] or results["awards"]):
        return None

    names = data["names"]
    n = data["n"]
    totals = [0] * n
    exact = [0] * n
    outcome_hits = [0] * n
    advance_hits = [0] * n

    def add_score_points(match, result, counts):
        if "score" not in result:
            return
        rh, ra = result["score"]
        for i, (h, a) in enumerate(match["score_picks"]):
            if h is None or a is None:
                continue
            if not counts[i]:
                continue
            if outcome(h, a) == outcome(rh, ra):
                totals[i] += 3
                outcome_hits[i] += 1
                if h == rh and a == ra:
                    totals[i] += 2
                    exact[i] += 1

    def add_text_points(picks, actual, points, bucket=None):
        if not actual:
            return
        actual_key = _cmp_team(actual)
        for i, pick in enumerate(picks):
            if pick and _cmp_team(pick) == actual_key:
                totals[i] += points
                if bucket is not None:
                    bucket[i] += 1

    for rnd in data["knockouts"]["rounds"]:
        for match in rnd["matches"]:
            result = results["matches"].get(match["code"])
            if not result:
                continue
            add_score_points(match, result, _ko_matchup_counts(match, data))
            add_text_points(match["winner_picks"], result.get("winner"), rnd["advance_points"], advance_hits)

    for match in data["knockouts"]["final_matches"]:
        result = results["matches"].get(match["code"])
        if not result:
            continue
        add_score_points(match, result, _ko_matchup_counts(match, data))

    for key, meta in data["knockouts"]["outright"].items():
        actual = results["outright"].get(key)
        add_text_points(meta["picks"], actual, meta["points"], advance_hits)

    for key, meta in data["knockouts"]["awards"].items():
        actual = results["awards"].get(key)
        add_text_points(meta["picks"], actual, meta["points"], advance_hits)

    table = sorted(
        [
            {
                "name": names[i],
                "pts": totals[i],
                "exact": exact[i],
                "outcomes": outcome_hits[i],
                "advance": advance_hits[i],
            }
            for i in range(n)
        ],
        key=lambda x: (-x["pts"], x["name"].lower()),
    )
    for i, row in enumerate(table, 1):
        row["rank"] = i
    return {"table": table, "played": sum(1 for v in results["matches"].values() if "score" in v)}


KO_SURVIVAL_ROUNDS = [
    {"key": "r32", "label_es": "Dieciseisavos", "label_en": "Round of 32"},
    {"key": "r16", "label_es": "Octavos", "label_en": "Round of 16"},
    {"key": "qf", "label_es": "Cuartos", "label_en": "Quarter-finals"},
    {"key": "sf", "label_es": "Semis", "label_en": "Semi-finals"},
    {"key": "final", "label_es": "Final", "label_en": "Final"},
    {"key": "champion", "label_es": "Campeón", "label_en": "Champion"},
]
KO_BRACKET_ROUNDS = KO_SURVIVAL_ROUNDS[:-1]


def _ko_public_pick_fields(match, names, n):
    """Picks individuales y analítica agregada para la vista Hoy."""
    picks = []
    for i, (h, a) in enumerate(match["score_picks"]):
        winner = match["winner_picks"][i] if match.get("winner_picks") else None
        picks.append({
            "name": names[i],
            "home": h,
            "away": a,
            "winner": team_es(winner) if winner else None,
            "winner_flag": team_flag(winner) if winner else "",
        })
    o1 = sum(1 for h, a in match["score_picks"] if h is not None and h > a)
    ox = sum(1 for h, a in match["score_picks"] if h is not None and h == a)
    o2 = sum(1 for h, a in match["score_picks"] if h is not None and h < a)
    counter = Counter(
        f"{h}-{a}" for h, a in match["score_picks"] if h is not None and a is not None
    )
    modal = counter.most_common(1)
    modal_sl = modal[0][0] if modal else "–"
    modal_pct = round(modal[0][1] / n, 3) if modal and n else 0
    return {
        "picks": picks,
        "outcome_dist": {"1": o1, "X": ox, "2": o2},
        "modal_scoreline": modal_sl,
        "modal_scoreline_share": modal_pct,
    }


def _match_sort_key(match):
    return (match.get("date", ""), match.get("dt", ""), match.get("code", ""))


def _iter_scheduled_matches(data, matches):
    for m in matches:
        yield {**m, "kind": "group"}
    winner_by_w = build_ko_winner_map(data["knockout_results"], data["knockouts"])
    for rnd in data["knockouts"]["rounds"]:
        for m in rnd["matches"]:
            mm = {**m, "kind": "ko", "advance_points": rnd["advance_points"]}
            resolve_knockout_fixture(mm, winner_by_w)
            yield mm
    for m in data["knockouts"]["final_matches"]:
        mm = {**m, "kind": "ko", "advance_points": 0}
        resolve_knockout_fixture(mm, winner_by_w)
        yield mm


def _has_match_result(match, data):
    code = match["code"]
    if match.get("kind") == "group":
        return code in data["results"]
    result = data["knockout_results"]["matches"].get(code)
    return bool(result and "score" in result)


def _pending_earlier_same_day(data, matches, match):
    match_date = match.get("date")
    if not match_date:
        return []
    tkey = _match_sort_key(match)
    pending = []
    for m in _iter_scheduled_matches(data, matches):
        if m.get("date") != match_date:
            continue
        if _match_sort_key(m) >= tkey:
            continue
        if _has_match_result(m, data):
            continue
        pending.append({
            "code": m["code"],
            "home": m.get("resolved_home") or m.get("fixture_home") or m.get("home", ""),
            "away": m.get("resolved_away") or m.get("fixture_away") or m.get("away", ""),
            "home_flag": m.get("resolved_home_flag") or m.get("fixture_home_flag") or m.get("home_flag", ""),
            "away_flag": m.get("resolved_away_flag") or m.get("fixture_away_flag") or m.get("away_flag", ""),
            "time_es": m.get("time_es", ""),
        })
    return pending


def _earlier_unplayed_same_day(data, matches, match):
    return bool(_pending_earlier_same_day(data, matches, match))


def _filtered_data_before(data, matches, target):
    tkey = _match_sort_key(target)
    results = {
        m["code"]: data["results"][m["code"]]
        for m in matches
        if m["code"] in data["results"] and _match_sort_key(m) < tkey
    }
    ko_matches = {}
    for m in _iter_scheduled_matches(data, matches):
        if m.get("kind") == "group":
            continue
        code = m["code"]
        result = data["knockout_results"]["matches"].get(code)
        if result and "score" in result and _match_sort_key(m) < tkey:
            ko_matches[code] = result
    kr = data["knockout_results"]
    return {
        **data,
        "results": results,
        "knockout_results": {
            "matches": ko_matches,
            "outright": kr.get("outright", {}),
            "awards": kr.get("awards", {}),
        },
    }


def live_table_before_match(data, matches, target):
    """Ranking tras todos los resultados anteriores a este partido (fecha/hora)."""
    partial = _filtered_data_before(data, matches, target)
    if (
        not partial["results"]
        and not partial["knockout_results"]["matches"]
        and not partial["knockout_results"]["outright"]
        and not partial["knockout_results"]["awards"]
    ):
        return None
    live = compute_live(partial, matches)
    return live["table"] if live else None


def _deferred_stake(match, data, matches):
    return {
        "deferred": True,
        "pending_after": _pending_earlier_same_day(data, matches, match),
    }


def stake_for_ko_match(match, rnd, names, data, matches, live_table):
    if _earlier_unplayed_same_day(data, matches, match):
        return _deferred_stake(match, data, matches)
    table = live_table_before_match(data, matches, match) or live_table
    winner_by_w = build_ko_winner_map(data["knockout_results"], data["knockouts"])
    pub = _knockout_public_schedule(match)
    resolve_knockout_fixture(pub, winner_by_w)
    home = pub.get("resolved_home") or pub.get("fixture_home")
    away = pub.get("resolved_away") or pub.get("fixture_away")
    return compute_ko_match_stake(
        match, rnd, names, table, _ko_matchup_counts(match, data), home, away
    )


def stake_for_group_match(match, data, matches, live_table):
    if match["code"] in data["results"]:
        return None
    if _earlier_unplayed_same_day(data, matches, match):
        return _deferred_stake(match, data, matches)
    table = live_table_before_match(data, matches, match) or live_table
    return compute_match_stake(match, data["results"], table)


def _stake_live_context(live_table):
    sorted_table = sorted(live_table, key=lambda r: (-r["pts"], r["name"].lower()))
    rank_rows = {}
    for rank, row in enumerate(sorted_table, 1):
        rank_rows[row["name"]] = {**row, "rank": row.get("rank", rank)}
    return sorted_table, rank_rows


def _ranks_after_scenario(sorted_table, deltas_by_name):
    """Ranking tras un resultado: todos suman a la vez."""
    new_rows = [
        {"name": row["name"], "pts": row["pts"] + deltas_by_name.get(row["name"], 0)}
        for row in sorted_table
    ]
    new_rows.sort(key=lambda row: (-row["pts"], row["name"].lower()))
    return {row["name"]: rank for rank, row in enumerate(new_rows, 1)}


def _format_group_scenario(scenario):
    rh, ra = scenario
    score = f"{rh}-{ra}"
    return {"score": score, "score_es": score, "score_en": score}


def _format_ko_scenario(scenario, _match=None):
    rh, ra, res_winner = scenario
    result = {"score": f"{rh}-{ra}", "score_es": f"{rh}-{ra}", "score_en": f"{rh}-{ra}"}
    if rh == ra and res_winner:
        result["winner"] = team_es(res_winner)
        result["winner_en"] = res_winner
        result["winner_flag"] = team_flag(res_winner)
    return result


def _compute_scenario_stake(
    names, live_table, scenarios, points_for_name, format_scenario=None, extra=None,
):
    """Swing realista: simula cada apuesta única como resultado posible."""
    if not scenarios or not live_table:
        return None
    sorted_table, rank_rows = _stake_live_context(live_table)

    pickers = [name for name in names if points_for_name(name, scenarios[0]) is not None]
    if not pickers:
        return None

    personal_max = {
        name: max(points_for_name(name, scenario) for scenario in scenarios)
        for name in pickers
    }

    records = {name: [] for name in pickers}
    for scenario in scenarios:
        deltas = {}
        for name in names:
            pts = points_for_name(name, scenario)
            deltas[name] = pts if pts is not None else 0
        new_ranks = _ranks_after_scenario(sorted_table, deltas)
        for name in pickers:
            old_rank = rank_rows[name]["rank"]
            swing = old_rank - new_ranks[name]
            records[name].append({
                "swing": swing,
                "scenario": scenario,
                "at_max": deltas[name] == personal_max[name],
            })

    people = []
    total_pts = 0
    for name in pickers:
        recs = records[name]
        at_max_recs = [rec for rec in recs if rec["at_max"]]
        best_rec = max(at_max_recs, key=lambda rec: rec["swing"], default=None)
        swing_up = best_rec["swing"] if best_rec and best_rec["swing"] > 0 else 0
        best_result = (
            format_scenario(best_rec["scenario"])
            if format_scenario and best_rec and best_rec["swing"] > 0
            else None
        )

        worst_rec = min(recs, key=lambda rec: rec["swing"])
        swing_down = -worst_rec["swing"] if worst_rec["swing"] < 0 else 0
        worst_result = (
            format_scenario(worst_rec["scenario"])
            if format_scenario and worst_rec["swing"] < 0
            else None
        )

        people.append({
            "name": name,
            "max_pts": personal_max[name],
            "swing_up": swing_up,
            "swing_down": swing_down,
            "swing": swing_up,
            "best_result": best_result,
            "worst_result": worst_result,
        })
        total_pts += personal_max[name]

    result = {
        "max_swing": max((p["swing_up"] for p in people), default=0),
        "min_swing": max((p["swing_down"] for p in people), default=0),
        "max_points": total_pts,
        "picks": len(people),
        "people": people,
    }
    if people:
        top_up = max(people, key=lambda p: p["swing_up"])
        top_down = max(people, key=lambda p: p["swing_down"])
        if top_up["swing_up"] > 0:
            result["max_swing_result"] = top_up.get("best_result")
            result["max_swing_who"] = top_up["name"]
        if top_down["swing_down"] > 0:
            result["min_swing_result"] = top_down.get("worst_result")
            result["min_swing_who"] = top_down["name"]
    if extra:
        result.update(extra)
    return result


def _group_match_scenarios(match):
    scenarios = set()
    for pick in match["picks"]:
        if pick["home"] is not None:
            scenarios.add((pick["home"], pick["away"]))
    return list(scenarios)


def _group_points_for_pick(h, a, rh, ra):
    if h is None or a is None:
        return None
    if h == rh and a == ra:
        return 4
    if outcome(h, a) == outcome(rh, ra):
        return 2
    return 0


def _ko_match_scenarios(match, home=None, away=None):
    """Resultados plausibles del cruce REAL: cada marcador único de las apuestas,
    con el ganador limitado a los dos equipos que de verdad juegan (empate ->
    penaltis, cualquiera de los dos). Usa los equipos resueltos, no el hueco W##,
    para no inventar escenarios imposibles (p.ej. '1-1 · España' en Brasil-Noruega)."""
    home = home or match.get("fixture_home") or match.get("fixture_home_en", "")
    away = away or match.get("fixture_away") or match.get("fixture_away_en", "")
    has_winner = bool(match.get("winner_picks"))
    scenarios = {}
    for h, a in match["score_picks"]:
        if h is None or a is None:
            continue
        if h > a:
            options = [home]
        elif a > h:
            options = [away]
        elif has_winner:
            options = [home, away]  # empate -> penaltis a cualquiera de los dos
        else:
            continue
        for res_winner in options:
            if res_winner:
                scenarios[(h, a, _cmp_text(res_winner))] = (h, a, res_winner)
    return list(scenarios.values())


def _ko_points_for_pick(h, a, winner_pick, rh, ra, res_winner, advance_pts, marcador_counts=True):
    if h is None or a is None:
        return None
    pts = 0
    if marcador_counts and outcome(h, a) == outcome(rh, ra):
        pts += 3
        if h == rh and a == ra:
            pts += 2
    winner_ok = bool(
        res_winner and winner_pick and _cmp_team(winner_pick) == _cmp_team(res_winner)
    )
    if advance_pts and winner_ok:
        pts += advance_pts
    return pts


def compute_ko_match_stake(match, rnd, names, live_table, matchup=None, home=None, away=None):
    """Cuánto puede mover el ranking general un cruce KO (escenarios realistas).

    `matchup[i]` indica si el cruce coincide con el que predijo la persona i; si
    no, su marcador no puede puntuar (solo el pase). `home`/`away` son los equipos
    reales resueltos, para generar escenarios (y el pase) con los equipos correctos.
    """
    advance_pts = rnd["advance_points"] if rnd else 0
    has_winner = bool(match.get("winner_picks"))
    max_one = (3 + 2 + advance_pts) if has_winner else 5
    if not live_table:
        live_table = [{"name": name, "pts": 0} for name in names]
    name_index = {name: i for i, name in enumerate(names)}
    scenarios = _ko_match_scenarios(match, home, away)
    if not scenarios:
        return None

    def points_for_name(name, scenario):
        i = name_index[name]
        h, a = match["score_picks"][i]
        winner = match["winner_picks"][i] if has_winner else None
        if h is None and not winner:
            return None
        rh, ra, res_winner = scenario
        marcador_counts = matchup[i] if matchup is not None else True
        return _ko_points_for_pick(h, a, winner, rh, ra, res_winner, advance_pts, marcador_counts)

    extra = {"max_one": max_one, "advance_points": advance_pts}
    stake = _compute_scenario_stake(
        names,
        live_table,
        scenarios,
        points_for_name,
        lambda scenario: _format_ko_scenario(scenario, match),
        extra,
    )
    if not stake:
        return None
    for person in stake["people"]:
        i = name_index[person["name"]]
        h, a = match["score_picks"][i]
        winner = match["winner_picks"][i] if has_winner else None
        person["score"] = f"{h}-{a}" if h is not None else None
        person["winner"] = team_es(winner) if winner else None
        person["winner_flag"] = team_flag(winner) if winner else ""
    return stake


def _ko_match_side_keys(match, pub, winner_by_w):
    """Claves canónicas de local y visitante en un cruce resuelto."""
    keys = []
    for side in ("home", "away"):
        val = pub.get(f"fixture_{side}") or match.get(f"fixture_{side}", "")
        if val and not str(val).startswith("W"):
            keys.append(_cmp_team(val))
            continue
        resolved = pub.get(f"resolved_{side}")
        if resolved:
            keys.append(_cmp_team(resolved))
            continue
        wm = re.match(r"^W(\d+)$", str(val))
        if wm:
            team = winner_by_w.get(int(wm.group(1)))
            if team:
                keys.append(_cmp_team(team.get("name_en") or team.get("name")))
    return keys


def _ko_champion_fell_round(person_idx, data):
    """Ronda en la que el campeón predicho queda fuera del torneo."""
    raw = data["knockouts"]
    champ_pick = raw["outright"]["champion"]["picks"][person_idx]
    fell_alive = len(KO_SURVIVAL_ROUNDS)
    if not champ_pick:
        return fell_alive

    champ_key = _cmp_team(champ_pick)
    if champ_key in _ko_alive_teams(data):
        return fell_alive

    winner_by_w = build_ko_winner_map(data["knockout_results"], raw)
    results = data["knockout_results"]["matches"]

    for round_idx, rnd_meta in enumerate(KO_SURVIVAL_ROUNDS):
        if rnd_meta["key"] == "champion":
            continue
        key = rnd_meta["key"]
        if key == "final":
            matches = [m for m in raw["final_matches"] if m["key"] == "final"]
        else:
            rnd = next((r for r in raw["rounds"] if r["key"] == key), None)
            if not rnd:
                continue
            matches = rnd["matches"]
        for match in matches:
            result = results.get(match["code"])
            if not result or not result.get("winner"):
                continue
            pub = _knockout_public_schedule(match)
            resolve_knockout_fixture(pub, winner_by_w)
            if champ_key not in _ko_match_side_keys(match, pub, winner_by_w):
                continue
            if _cmp_team(result["winner"]) != champ_key:
                return round_idx

    return 0


def _ko_round_matches(raw, key):
    if key == "final":
        return [m for m in raw["final_matches"] if m["key"] == "final"]
    rnd = next((r for r in raw["rounds"] if r["key"] == key), None)
    return rnd["matches"] if rnd else []


def _ko_fixture_pair(match, w_map):
    """Par de selecciones (ordenado) que juegan el cruce, o None si no se resuelve."""
    sides = []
    for side in ("home", "away"):
        val = match.get(f"fixture_{side}", "")
        wm = re.match(r"^W(\d+)$", str(val))
        if wm:
            team = w_map.get(int(wm.group(1)))
            if not team:
                return None
            sides.append(team)
        elif val:
            sides.append(_cmp_team(val))
        else:
            return None
    return tuple(sorted(sides))


def _ko_person_bracket_rounds(person_idx, data):
    """Precisión del cuadro por fase.

    - Cruce ya jugado: acierto si el ganador que pusiste es el real; si no, fallo
      (rojo), da igual si tu rama diverge — no puntuaste.
    - Cruce futuro: si el equipo que pusiste como ganador ya está eliminado, la
      rama está muerta (rayado). Si sigue vivo, aún pendiente.
    """
    raw = data["knockouts"]
    results = data["knockout_results"]["matches"]
    alive = _ko_alive_teams(data)
    # La Final no tiene pick de ganador por partido: su "ganador" es el campeón.
    champ_meta = data["knockouts"]["outright"].get("champion")
    champ_result = data["knockout_results"]["outright"].get("champion")
    rounds = []

    for rnd_meta in KO_BRACKET_ROUNDS:
        matches = _ko_round_matches(raw, rnd_meta["key"])
        total = len(matches)
        hits = misses = drift = played = 0
        for match in matches:
            if match["code"] == "FINAL":
                pick = champ_meta["picks"][person_idx] if champ_meta else None
                actual_winner = champ_result
            else:
                result = results.get(match["code"])
                actual_winner = result["winner"] if (result and result.get("winner")) else None
                pick = match["winner_picks"][person_idx] if match.get("winner_picks") else None

            if actual_winner:
                played += 1
                if not pick:
                    continue
                if _cmp_team(pick) == _cmp_team(actual_winner):
                    hits += 1
                else:
                    misses += 1
            elif pick and _cmp_team(pick) not in alive:
                # Cruce por jugar cuyo ganador que pusiste ya está fuera: rama muerta.
                drift += 1

        rounds.append({
            "hits": hits,
            "misses": misses,
            "drift": drift,
            "played": played,
            "total": total,
        })
    return rounds


def _iter_knockout_played_matches(data):
    """Cruces KO jugados en orden cronológico."""
    results = data["knockout_results"]["matches"]
    winner_by_w = build_ko_winner_map(data["knockout_results"], data["knockouts"])
    played = []

    for rnd in data["knockouts"]["rounds"]:
        for match in rnd["matches"]:
            result = results.get(match["code"])
            if not result or "score" not in result:
                continue
            pub = _knockout_public_schedule(match)
            resolve_knockout_fixture(pub, winner_by_w)
            played.append((rnd, match, result, pub))

    for match in data["knockouts"]["final_matches"]:
        result = results.get(match["code"])
        if not result or "score" not in result:
            continue
        pub = _knockout_public_schedule(match)
        resolve_knockout_fixture(pub, winner_by_w)
        played.append((None, match, result, pub))

    played.sort(key=lambda item: (item[3]["date"], item[3].get("dt", ""), item[1]["code"]))
    return played


def compute_ko_progression(data):
    """Ranking KO partido a partido (subidón/batacazo por jornada)."""
    played = _iter_knockout_played_matches(data)
    if not played:
        return None

    names = data["names"]
    n = data["n"]
    pts = [0] * n
    exact = [0] * n
    outcomes = [0] * n
    advance = [0] * n
    previous_ranks = {}
    progression = []

    for idx, (rnd, match, result, pub) in enumerate(played, 1):
        before_pts = pts[:]
        before_exact = exact[:]
        before_out = outcomes[:]
        before_adv = advance[:]
        rh, ra = result["score"]
        ro = outcome(rh, ra)
        adv_pts = rnd["advance_points"] if rnd else 0

        counts = _ko_matchup_counts(match, data)
        for i, (h, a) in enumerate(match["score_picks"]):
            if h is None or a is None:
                continue
            if not counts[i]:
                continue
            if outcome(h, a) == ro:
                pts[i] += 3
                outcomes[i] += 1
                if h == rh and a == ra:
                    pts[i] += 2
                    exact[i] += 1

        if rnd and result.get("winner"):
            actual = _cmp_team(result["winner"])
            for i, pick in enumerate(match["winner_picks"]):
                if pick and _cmp_team(pick) == actual:
                    pts[i] += adv_pts
                    advance[i] += 1

        rows = []
        for i in range(n):
            rows.append({
                "name": names[i],
                "pts": pts[i],
                "exact": exact[i],
                "outcomes": outcomes[i],
                "advance": advance[i],
                "_order": i,
            })
        rows.sort(key=lambda row: (-row["pts"], row["_order"]))
        for rank, row in enumerate(rows, 1):
            i = row["_order"]
            row["rank"] = rank
            old_rank = previous_ranks.get(row["name"], rank)
            row["delta"] = old_rank - rank
            row["round_pts"] = pts[i] - before_pts[i]
            row["round_exact"] = exact[i] - before_exact[i]
            row["round_outcomes"] = outcomes[i] - before_out[i]
            row["round_advance"] = advance[i] - before_adv[i]
            del row["_order"]
        previous_ranks = {row["name"]: row["rank"] for row in rows}

        phase_es = rnd["label_es"] if rnd else match["label_es"]
        phase_en = rnd["label_en"] if rnd else match["label_en"]
        progression.append({
            "idx": idx,
            "code": match["code"],
            "date": pub["date"],
            "dt": pub.get("dt", ""),
            "home": pub.get("resolved_home") or pub["fixture_home"],
            "away": pub.get("resolved_away") or pub["fixture_away"],
            "home_flag": pub.get("resolved_home_flag") or pub["fixture_home_flag"],
            "away_flag": pub.get("resolved_away_flag") or pub["fixture_away_flag"],
            "phase_es": phase_es,
            "phase_en": phase_en,
            "result": {"home": rh, "away": ra, "outcome": ro},
            "table": rows,
        })

    final_table = progression[-1]["table"]
    return {
        "played": len(played),
        "steps": len(progression),
        "table": final_table,
        "progression": progression,
    }


def _team_prestige(data):
    """Prestigio por consenso de clasificación (1.º=3, 2.º=2, 3.º=1)."""
    pos_weight = {1: 3, 2: 2, 3: 1}
    prestige = Counter()
    for pos in data.get("qualifiers", {}).values():
        for slot, picks in pos.items():
            w = pos_weight.get(slot, 1)
            for pick in picks:
                if pick:
                    prestige[_cmp_text(team_es(pick))] += w
    return prestige


def _ko_resolved_side_key(match, side, w_map):
    """Clave canónica del equipo en un lado del cruce, resolviendo feeders W##."""
    val = str(match.get(f"fixture_{side}", ""))
    wm = re.match(r"^W(\d+)$", val)
    if wm:
        return w_map.get(int(wm.group(1)))
    return _cmp_team(val) if val else None


def _ko_alive_teams(data):
    """Equipos que aún pueden ganar el torneo (en R32 y sin derrota KO)."""
    alive = set()
    for rnd in data["knockouts"]["rounds"]:
        if rnd["key"] != "r32":
            continue
        for match in rnd["matches"]:
            for side in ("home", "away"):
                name = match.get(f"fixture_{side}", "")
                if name and not str(name).startswith("W"):
                    alive.add(_cmp_team(name))

    results = data["knockout_results"]["matches"]
    # Los cruces de octavos en adelante llevan feeders W## en fixture_home/away;
    # hay que resolverlos al ganador real para poder descartar al perdedor.
    w_map = _ko_w_map_from_results(data)
    for rnd in data["knockouts"]["rounds"]:
        for match in rnd["matches"]:
            result = results.get(match["code"])
            if not result or not result.get("winner"):
                continue
            winner_key = _cmp_team(result["winner"])
            for side in ("home", "away"):
                side_key = _ko_resolved_side_key(match, side, w_map)
                if side_key and side_key != winner_key:
                    alive.discard(side_key)
    return alive


def _ko_team_pair(match):
    home = match.get("fixture_home", "")
    away = match.get("fixture_away", "")
    if not home or not away or str(home).startswith("W") or str(away).startswith("W"):
        return None, None
    return team_es(home), team_es(away)


def _ko_underdog_key(home_es, away_es, prestige):
    home_key = _cmp_text(home_es)
    away_key = _cmp_text(away_es)
    if prestige.get(home_key, 0) <= prestige.get(away_key, 0):
        return home_key
    return away_key


def _ko_pick_similarity(i, j, raw, outright, awards):
    same = 0
    total = 0
    for rnd in raw["rounds"]:
        for match in rnd["matches"]:
            wi = match["winner_picks"][i]
            wj = match["winner_picks"][j]
            if wi and wj:
                total += 1
                if _cmp_text(wi) == _cmp_text(wj):
                    same += 1
    sim = round(100 * same / total, 1) if total else 0.0
    bonus = 0
    for key, pts in (("champion", 15), ("runner_up", 10)):
        meta = outright.get(key)
        if not meta:
            continue
        pi = meta["picks"][i]
        pj = meta["picks"][j]
        if pi and pj and _cmp_text(pi) == _cmp_text(pj):
            bonus += pts
    ts_i = awards["top_scorer"]["picks"][i] if awards.get("top_scorer") else None
    ts_j = awards["top_scorer"]["picks"][j] if awards.get("top_scorer") else None
    if ts_i and ts_j and _cmp_text(ts_i) == _cmp_text(ts_j):
        bonus += 10
    return min(99.0, round(sim * 0.7 + bonus, 1))


def _ko_depth_row(team_name, raw, n):
    team_key = _cmp_text(team_es(team_name))
    by_round = {rnd["key"]: 0 for rnd in raw["rounds"]}
    for rnd in raw["rounds"]:
        for match in rnd["matches"]:
            for pick in match["winner_picks"]:
                if pick and _cmp_text(pick) == team_key:
                    by_round[rnd["key"]] += 1

    champ_picks = raw["outright"]["champion"]["picks"]
    champ_count = sum(1 for p in champ_picks if p and _cmp_text(p) == team_key)
    pct = lambda c: round(100 * c / n, 1) if n else 0
    return {
        "team": team_es(team_name),
        "flag": team_flag(team_name),
        "r16": pct(by_round.get("r32", 0)),
        "qf": pct(by_round.get("r16", 0)),
        "sf": pct(by_round.get("qf", 0)),
        "fin": pct(by_round.get("sf", 0)),
        "champ": pct(champ_count),
    }


def _ko_consensus_w_map(raw, n):
    """Cuadro del pueblo: ganador modal por cruce → mapa W## → team key."""
    w_map = {}
    for rnd in raw["rounds"]:
        for match in rnd["matches"]:
            w = ko_w_number(match["code"])
            consensus = _text_consensus(match["winner_picks"], n).get("value")
            if w and consensus:
                w_map[w] = _cmp_team(consensus)
    return w_map


def _ko_pick_share(picks, pick, n):
    """Probabilidad implícita del consenso para un pick (0–1)."""
    filled = [p for p in picks if p]
    if not pick or not filled:
        return 0.0
    key = _cmp_text(team_es(pick))
    counter = Counter(_cmp_text(team_es(p)) for p in filled)
    return counter.get(key, 0) / len(filled)


def _ko_scoreline_share(picks, h, a):
    filled = [(x, y) for x, y in picks if x is not None and y is not None]
    if h is None or a is None or not filled:
        return 0.0
    counter = Counter(f"{x}-{y}" for x, y in filled)
    return counter.get(f"{h}-{a}", 0) / len(filled)


def _ko_person_risk_reward(i, raw, n, outright, awards_meta, score_row, has_scoring):
    """Riesgo (contrarianismo vs consenso) y pts esperados por alineación implícita."""
    raw_risk = 0.0
    risk_n = 0.0
    exp = 0.0

    for rnd in raw["rounds"]:
        for match in rnd["matches"]:
            pick = match["winner_picks"][i] if match.get("winner_picks") else None
            if pick:
                share = _ko_pick_share(match["winner_picks"], pick, n)
                raw_risk += 1.0 - share
                risk_n += 1.0
                exp += rnd["advance_points"] * share
            h, a = match["score_picks"][i]
            if h is not None:
                share = _ko_scoreline_share(match["score_picks"], h, a)
                raw_risk += (1.0 - share) * 0.5
                risk_n += 0.5
                exp += 5.0 * share

    for meta in outright.values():
        pick = meta["picks"][i]
        if pick:
            share = _ko_pick_share(meta["picks"], pick, n)
            raw_risk += 1.0 - share
            risk_n += 1.0
            exp += meta["points"] * share

    for meta in awards_meta.values():
        pick = meta["picks"][i]
        if pick:
            share = _ko_pick_share(meta["picks"], pick, n)
            raw_risk += 1.0 - share
            risk_n += 1.0
            exp += meta["points"] * share

    avg_risk = raw_risk / risk_n if risk_n else 0.0
    if has_scoring:
        exp = score_row.get("pts", 0)
    return avg_risk, round(exp, 1)


def _ko_person_vs_pueblo(i, raw, n):
    diffs = total = 0
    for rnd in raw["rounds"]:
        for match in rnd["matches"]:
            consensus = _text_consensus(match["winner_picks"], n).get("value")
            pick = match["winner_picks"][i] if match.get("winner_picks") else None
            if not consensus or not pick:
                continue
            total += 1
            if _cmp_team(pick) != _cmp_team(consensus):
                diffs += 1
    return diffs, total


def _ko_reventador_counts(data, prestige):
    counts = [0] * data["n"]
    results = data["knockout_results"]["matches"]
    for rnd in data["knockouts"]["rounds"]:
        for match in rnd["matches"]:
            result = results.get(match["code"])
            if not result or not result.get("winner"):
                continue
            home_es, away_es = _ko_team_pair(match)
            if not home_es:
                continue
            underdog = _ko_underdog_key(home_es, away_es, prestige)
            if _cmp_team(result["winner"]) != underdog:
                continue
            for i, pick in enumerate(match["winner_picks"]):
                if pick and _cmp_text(pick) == underdog:
                    counts[i] += 1
    return counts


def _ko_champion_path(champ_team, raw, n):
    """Camino del campeón consensual a través del cuadro del pueblo."""
    if not champ_team:
        return []
    champ_key = _cmp_team(champ_team)
    w_map = _ko_consensus_w_map(raw, n)
    path = []

    for rnd in raw["rounds"]:
        for match in rnd["matches"]:
            pair = _ko_fixture_pair(match, w_map)
            if not pair or champ_key not in pair:
                continue
            opp_key = pair[0] if pair[1] == champ_key else pair[1]
            opp_name = opp_flag = None
            for side in ("home", "away"):
                val = match.get(f"fixture_{side}", "")
                if val and not str(val).startswith("W") and _cmp_team(val) == opp_key:
                    opp_name = team_es(val)
                    opp_flag = team_flag(val)
                    break
            if not opp_name:
                for side in ("home", "away"):
                    val = match.get(f"fixture_{side}", "")
                    wm = re.match(r"^W(\d+)$", str(val))
                    if wm:
                        resolved = w_map.get(int(wm.group(1)))
                        if resolved == opp_key:
                            opp_name = team_es(resolved)
                            opp_flag = team_flag(resolved)
                            break
            w = ko_w_number(match["code"])
            winner = w_map.get(w)
            path.append({
                "code": match["code"],
                "round_es": rnd["label_es"],
                "round_en": rnd["label_en"],
                "opponent": opp_name or opp_key,
                "opponent_flag": opp_flag or "🏳️",
                "agreement": round(_ko_pick_share(match["winner_picks"], champ_team, n) * 100, 1),
            })
            if winner != champ_key:
                break
    return path


def _compute_ko_honors(names, people, reventador, scoring_table):
    risk_rank = sorted(people, key=lambda p: (-p["variance"], p["name"].lower()))
    manual = risk_rank[-1] if risk_rank else None
    agorero = risk_rank[0] if risk_rank else None
    vs_rank = sorted(people, key=lambda p: (-p["boldPct"], p["name"].lower()))
    reventador_rank = sorted(
        [{"name": names[i], "count": c} for i, c in enumerate(reventador) if c],
        key=lambda x: (-x["count"], x["name"].lower()),
    )
    rev = reventador_rank[0] if reventador_rank else {"name": "–", "count": 0}
    if scoring_table:
        prophet = {"name": scoring_table[0]["name"], "pts": scoring_table[0]["pts"], "approx": False}
    else:
        exp_rank = sorted(people, key=lambda p: (-p["exp"], p["name"].lower()))
        top = exp_rank[0] if exp_rank else None
        prophet = {"name": top["name"], "pts": top["exp"], "approx": True} if top else {"name": "–", "pts": 0, "approx": True}
    return {
        "manual": {"name": manual["name"], "risk": manual["variance"]} if manual else {"name": "–", "risk": 0},
        "agorero": {"name": agorero["name"], "risk": agorero["variance"]} if agorero else {"name": "–", "risk": 0},
        "reventador": rev,
        "profeta": prophet,
        "vsPueblo": {
            "name": vs_rank[0]["name"],
            "diff": vs_rank[0]["vsPueblo"],
            "pct": vs_rank[0]["boldPct"],
        } if vs_rank else {"name": "–", "diff": 0, "pct": 0},
    }


def compute_knockout_metrics(data):
    """Métricas del relato KO derivadas de las porras reales."""
    names = data["names"]
    n = data["n"]
    raw = data["knockouts"]
    outright = raw["outright"]
    awards = raw["awards"]
    prestige = _team_prestige(data)
    alive = _ko_alive_teams(data)
    scoring_by_name = {}
    scoring_table = None
    has_scoring = False
    if data.get("knockout_results"):
        scoring = compute_knockout_scoring(data)
        if scoring:
            has_scoring = True
            scoring_table = scoring["table"]
            scoring_by_name = {row["name"]: row for row in scoring_table}

    champ_consensus = _text_consensus(outright["champion"]["picks"], n)
    champ_rank = [
        {"team": row["value"], "flag": row["flag"], "count": row["count"]}
        for row in champ_consensus.get("dist", [])
    ]

    ts_counter = Counter(p for p in awards.get("top_scorer", {}).get("picks", []) if p)
    ts_rank = [{"name": t, "count": c} for t, c in ts_counter.most_common(8)]

    reventador = _ko_reventador_counts(data, prestige)
    champ_team = champ_consensus.get("value")
    champion_path = _ko_champion_path(champ_team, raw, n)
    raw_risks = []
    exp_vals = []
    vs_pueblo = []
    vs_pueblo_total = []
    for i in range(n):
        score_row = scoring_by_name.get(names[i], {})
        risk, exp = _ko_person_risk_reward(
            i, raw, n, outright, awards, score_row, has_scoring,
        )
        raw_risks.append(risk)
        exp_vals.append(exp)
        diffs, total = _ko_person_vs_pueblo(i, raw, n)
        vs_pueblo.append(diffs)
        vs_pueblo_total.append(total)
    risk_index = [round(v, 1) for v in minmax_scale(raw_risks)]

    twin_pairs = []
    for i in range(n):
        for j in range(i + 1, n):
            sim = _ko_pick_similarity(i, j, raw, outright, awards)
            if sim >= 45:
                twin_pairs.append({"a": names[i], "b": names[j], "sim": sim})
    twin_pairs.sort(key=lambda x: (-x["sim"], x["a"], x["b"]))

    grave = []
    for i, pick in enumerate(outright["champion"]["picks"]):
        if not pick:
            continue
        pick_es = team_es(pick)
        if _cmp_team(pick) not in alive:
            grave.append({"name": names[i], "champ": pick_es, "flag": team_flag(pick)})

    round_labels = [
        {"es": meta["label_es"], "en": meta["label_en"]}
        for meta in KO_SURVIVAL_ROUNDS
    ]
    bracket_round_labels = [
        {"es": meta["label_es"], "en": meta["label_en"]}
        for meta in KO_BRACKET_ROUNDS
    ]

    depth_teams = []
    seen = set()
    for rnd in raw["rounds"]:
        if rnd["key"] != "r32":
            continue
        for match in rnd["matches"]:
            for side in ("home", "away"):
                name = match.get(f"fixture_{side}", "")
                if not name or str(name).startswith("W"):
                    continue
                key = _cmp_text(team_es(name))
                if key not in seen:
                    seen.add(key)
                    depth_teams.append(name)
    depth_teams.sort(key=lambda t: (-prestige.get(_cmp_text(team_es(t)), 0), team_es(t)))
    depth = [_ko_depth_row(t, raw, n) for t in depth_teams[:12]]

    people = []
    for i, name in enumerate(names):
        champ_pick = outright["champion"]["picks"][i]
        runner_pick = outright.get("runner_up", {}).get("picks", [None] * n)[i]
        ts_pick = awards.get("top_scorer", {}).get("picks", [None] * n)[i]
        fell = _ko_champion_fell_round(i, data)
        bracket = _ko_person_bracket_rounds(i, data)
        champ_arr = [team_es(champ_pick), team_flag(champ_pick)] if champ_pick else ["–", "🏳️"]
        runner_arr = [team_es(runner_pick), team_flag(runner_pick)] if runner_pick else ["–", "🏳️"]
        total = vs_pueblo_total[i]
        diffs = vs_pueblo[i]
        people.append({
            "name": name,
            "champ": champ_arr,
            "runner": runner_arr,
            "ts": ts_pick or "–",
            "boldPct": round(100 * diffs / total) if total else 0,
            "fell": fell,
            "bracket": bracket,
            "exp": exp_vals[i],
            "expApprox": not has_scoring,
            "variance": risk_index[i],
            "vsPueblo": diffs,
            "vsPuebloTotal": total,
            "reventador": reventador[i],
        })

    bold_rank = sorted(people, key=lambda p: (-p["boldPct"], p["name"].lower()))
    vs_pueblo_rank = sorted(people, key=lambda p: (-p["boldPct"], p["name"].lower()))
    honors = _compute_ko_honors(names, people, reventador, scoring_table)

    return {
        "people": people,
        "rounds": round_labels,
        "bracketRounds": bracket_round_labels,
        "champRank": champ_rank,
        "tsRank": ts_rank,
        "depth": depth,
        "twins": twin_pairs,
        "grave": grave,
        "boldRank": bold_rank,
        "vsPuebloRank": vs_pueblo_rank,
        "championPath": champion_path,
        "championTeam": champ_team,
        "championFlag": champ_consensus.get("flag", "🏳️"),
        "honors": honors,
        "expApprox": not has_scoring,
        "pool": [[row["team"], row["flag"], row["champ"]] for row in depth[:12]],
    }


def _knockout_public_schedule(match):
    keys = (
        "fixture_home",
        "fixture_away",
        "fixture_home_en",
        "fixture_away_en",
        "fixture_home_flag",
        "fixture_away_flag",
        "date",
        "time_es",
        "time_uk",
        "next_es",
        "next_uk",
        "dt",
        "venue",
        "city",
    )
    return {k: match.get(k, "") for k in keys}


def _knockout_public_result(result):
    if not result:
        return None
    out = {}
    if "score" in result:
        h, a = result["score"]
        out["score"] = {"home": h, "away": a}
    if result.get("winner"):
        out["winner"] = team_es(result["winner"])
        out["winner_flag"] = team_flag(result["winner"])
    return out or None


def _filled_score_picks(picks):
    return sum(1 for h, a in picks if h is not None and a is not None)


def _filled_text_picks(picks):
    return sum(1 for p in picks if p)


def _score_consensus(picks, n):
    counter = Counter(f"{h}-{a}" for h, a in picks if h is not None and a is not None)
    if not counter:
        return {"value": None, "count": 0, "agreement": 0, "dist": []}
    value, count = counter.most_common(1)[0]
    return {
        "value": value,
        "count": count,
        "agreement": round(100 * count / n, 1) if n else 0,
        "dist": [{"value": k, "count": v} for k, v in counter.most_common(5)],
    }


def _text_consensus(picks, n):
    counter = Counter(team_es(p) for p in picks if p)
    if not counter:
        return {"value": None, "flag": "🏳️", "count": 0, "agreement": 0, "dist": []}
    value, count = counter.most_common(1)[0]
    return {
        "value": value,
        "flag": _flag_es(value),
        "count": count,
        "agreement": round(100 * count / n, 1) if n else 0,
        "dist": [
            {"value": k, "flag": _flag_es(k), "count": v}
            for k, v in counter.most_common(5)
        ],
    }


def _cmp_text(v):
    return str(v).strip().casefold()


def _cmp_team(name):
    """Clave canónica para comparar selecciones (EN/ES, acentos, alias FIFA)."""
    if name is None:
        return None
    return _cmp_text(_team_key(name))


# Baremo de clasificados (idéntico al de la hoja Calculation del Excel):
# 1.º acertado = 3 (1 por clasificar + 2 por la posición exacta);
# 2.º y 3.º = 1 si el equipo queda entre los tres del grupo; mejor tercero = 1.
STANDINGS_EXACT_FIRST_BONUS = 2
STANDINGS_QUALIFY_POINTS = 1
THIRD_POINTS = 1


def compute_standings_points(data):
    """Puntos por clasificados de grupo (1.º/2.º/3.º) y por los 8 mejores
    terceros, por participante. Replica exactamente las fórmulas del Excel.

    Devuelve (standings_pts, thirds_pts, breakdown) donde breakdown lleva el
    desglose por persona para poder mostrarlo en la web.
    """
    n = data["n"]
    standings_results = data.get("standings_results") or {}
    thirds_results = data.get("thirds_results") or []
    qualifiers = data.get("qualifiers") or {}
    thirds_picks = data.get("thirds") or []

    standings_pts = [0] * n
    thirds_pts = [0] * n

    for group, actual in standings_results.items():
        first = _cmp_text(actual[1]) if actual.get(1) else None
        qualified = {_cmp_text(actual[pos]) for pos in actual if actual.get(pos)}
        picks = qualifiers.get(group, {})
        for pos in (1, 2, 3):
            slot = picks.get(pos)
            if not slot:
                continue
            for i, pick in enumerate(slot):
                if not pick:
                    continue
                key = _cmp_text(pick)
                if key in qualified:
                    standings_pts[i] += STANDINGS_QUALIFY_POINTS
                    if pos == 1 and first and key == first:
                        standings_pts[i] += STANDINGS_EXACT_FIRST_BONUS

    if thirds_results:
        actual_thirds = {_cmp_text(t) for t in thirds_results}
        for row in thirds_picks:
            for i, pick in enumerate(row):
                if pick and _cmp_text(pick) in actual_thirds:
                    thirds_pts[i] += THIRD_POINTS

    ready = bool(standings_results) or bool(thirds_results)
    return standings_pts, thirds_pts, ready


def compute_live(data, matches):
    """Ranking simplificado de aciertos cuando hay resultados (signo=2, pleno=4)."""
    results = data["results"]
    names = data["names"]
    n = data["n"]
    pts = [0] * n
    exact = [0] * n
    sign = [0] * n
    name_index = {name: i for i, name in enumerate(names)}

    played_matches = sorted(
        [m for m in matches if m["code"] in results],
        key=lambda m: (m["date"], m.get("dt", ""), m["code"]),
    )
    standings_pts, thirds_pts, standings_ready = compute_standings_points(data)

    def snapshot_rows(before_pts=None, round_exact=None, round_sign=None, show_standings=False):
        nonlocal previous_ranks
        rows = []
        for i in range(n):
            match_only = before_pts[i] if (show_standings and before_pts) else pts[i]
            rows.append({
                "name": names[i],
                "pts": pts[i],
                "group_pts": match_only,
                "standings_pts": standings_pts[i] if show_standings else 0,
                "thirds_pts": thirds_pts[i] if show_standings else 0,
                "exact": exact[i],
                "sign": sign[i],
                "_order": i,
            })
        rows.sort(key=lambda x: (-x["pts"], x["_order"]))
        for rank, row in enumerate(rows, 1):
            row["rank"] = rank
            i = row["_order"]
            if before_pts is not None:
                old_rank = previous_ranks.get(row["name"], rank)
                row["delta"] = old_rank - rank
                row["round_pts"] = pts[i] - before_pts[i]
                row["round_exact"] = round_exact[i] if round_exact else 0
                row["round_sign"] = round_sign[i] if round_sign else 0
            del row["_order"]
        if before_pts is not None:
            previous_ranks = {row["name"]: row["rank"] for row in rows}
        return rows

    progression = []
    previous_ranks = {}
    for idx, m in enumerate(played_matches, 1):
        before_pts = pts[:]
        before_exact = exact[:]
        before_sign = sign[:]
        rh, ra = results[m["code"]]
        ro = outcome(rh, ra)
        for i, (h, a) in enumerate(m["picks"]):
            if h is None:
                continue
            if h == rh and a == ra:
                pts[i] += 4
                exact[i] += 1
            elif outcome(h, a) == ro:
                pts[i] += 2
                sign[i] += 1
        rows = snapshot_rows(before_pts, exact, sign)
        for row in rows:
            i = name_index[row["name"]]
            row["round_exact"] = exact[i] - before_exact[i]
            row["round_sign"] = sign[i] - before_sign[i]
        progression.append({
            "idx": idx,
            "code": m["code"],
            "group": m["group"],
            "date": m["date"],
            "home": m["home"],
            "away": m["away"],
            "home_en": m["home_en"],
            "away_en": m["away_en"],
            "home_flag": m["home_flag"],
            "away_flag": m["away_flag"],
            "result": {"home": rh, "away": ra, "outcome": ro},
            "table": rows,
        })

    group_table = progression[-1]["table"] if progression else snapshot_rows()
    group_pts_by_name = {row["name"]: row["pts"] for row in group_table}
    last_date = played_matches[-1]["date"] if played_matches else ""

    # Paso virtual: cierre de fase de grupos (clasificados + mejores terceros).
    if standings_ready and any(standings_pts[i] + thirds_pts[i] for i in range(n)):
        before_pts = pts[:]
        for i in range(n):
            pts[i] += standings_pts[i] + thirds_pts[i]
        rows = snapshot_rows(before_pts, show_standings=True)
        for row in rows:
            i = name_index[row["name"]]
            row["round_standings_pts"] = standings_pts[i]
            row["round_thirds_pts"] = thirds_pts[i]
        progression.append({
            "idx": len(progression) + 1,
            "code": "STANDINGS",
            "virtual": True,
            "kind": "standings",
            "label_es": "Clasificados de grupo",
            "label_en": "Group standings",
            "date": last_date,
            "group": "",
            "table": rows,
        })

    ko = compute_knockout_scoring(data)
    ko_rows_by_name = {row["name"]: row for row in ko["table"]} if ko else {}
    ko_pts_by_name = {name: row["pts"] for name, row in ko_rows_by_name.items()}
    ko_progression = compute_ko_progression(data)

    def append_ko_step(meta, step_rows_by_name):
        before_pts = pts[:]
        for i, name in enumerate(names):
            pts[i] += step_rows_by_name.get(name, {}).get("round_pts", 0)
        rows = snapshot_rows(before_pts, show_standings=True)
        for row in rows:
            i = name_index[row["name"]]
            ko_row = step_rows_by_name.get(row["name"], {})
            row["group_pts"] = group_pts_by_name.get(row["name"], 0)
            row["standings_pts"] = standings_pts[i]
            row["thirds_pts"] = thirds_pts[i]
            row["ko_pts"] = ko_row.get("pts", 0)
            row["ko_exact"] = ko_row.get("exact", 0)
            row["ko_outcomes"] = ko_row.get("outcomes", 0)
            row["ko_advance"] = ko_row.get("advance", 0)
            row["round_ko_pts"] = ko_row.get("round_pts", 0)
            row["round_exact"] = ko_row.get("round_exact", 0)
            row["round_sign"] = ko_row.get("round_outcomes", 0)
            row["round_advance"] = ko_row.get("round_advance", 0)
        step = {
            "idx": len(progression) + 1,
            "code": meta["code"],
            "virtual": True,
            "kind": "ko",
            "label_es": "Eliminatorias",
            "label_en": "Knockouts",
            "date": meta.get("date", ""),
            "dt": meta.get("dt", ""),
            "group": "",
            "phase_es": meta.get("phase_es", "Eliminatorias"),
            "phase_en": meta.get("phase_en", "Knockouts"),
            "table": rows,
        }
        for key in ("home", "away", "home_flag", "away_flag", "result"):
            if key in meta:
                step[key] = meta[key]
        progression.append(step)

    if ko_progression and ko_progression.get("progression"):
        for ko_step in ko_progression["progression"]:
            append_ko_step(
                ko_step,
                {row["name"]: row for row in ko_step["table"]},
            )
        latest_ko_rows_by_name = {
            row["name"]: row for row in ko_progression["progression"][-1]["table"]
        }
        bonus_rows_by_name = {}
        for name, total_row in ko_rows_by_name.items():
            current = latest_ko_rows_by_name.get(name, {})
            bonus_rows_by_name[name] = {
                **total_row,
                "round_pts": total_row.get("pts", 0) - current.get("pts", 0),
                "round_exact": total_row.get("exact", 0) - current.get("exact", 0),
                "round_outcomes": total_row.get("outcomes", 0) - current.get("outcomes", 0),
                "round_advance": total_row.get("advance", 0) - current.get("advance", 0),
            }
        if any(row["round_pts"] for row in bonus_rows_by_name.values()):
            append_ko_step(
                {
                    "code": "KO-BONUS",
                    "date": progression[-1].get("date", last_date),
                    "phase_es": "Premios KO",
                    "phase_en": "KO awards",
                },
                bonus_rows_by_name,
            )
    elif ko_pts_by_name:
        append_ko_step(
            {
                "code": "KO",
                "date": last_date,
                "phase_es": "Eliminatorias",
                "phase_en": "Knockouts",
            },
            {
                name: {
                    **row,
                    "round_pts": row.get("pts", 0),
                    "round_exact": row.get("exact", 0),
                    "round_outcomes": row.get("outcomes", 0),
                    "round_advance": row.get("advance", 0),
                }
                for name, row in ko_rows_by_name.items()
            },
        )

    final_delta_by_name = {
        row["name"]: row.get("delta", 0)
        for row in (progression[-1]["table"] if progression else [])
    }

    grand_rows = []
    for i in range(n):
        nm = names[i]
        kp = ko_pts_by_name.get(nm, 0)
        ko_row = ko_rows_by_name.get(nm, {})
        grand_rows.append({
            "name": nm,
            "pts": pts[i],
            "group_pts": group_pts_by_name.get(nm, 0),
            "standings_pts": standings_pts[i],
            "thirds_pts": thirds_pts[i],
            "ko_pts": kp,
            "ko_exact": ko_row.get("exact", 0),
            "ko_outcomes": ko_row.get("outcomes", 0),
            "ko_advance": ko_row.get("advance", 0),
            "exact": exact[i],
            "sign": sign[i],
            "_order": i,
        })
    grand_rows.sort(key=lambda x: (-x["pts"], x["_order"]))
    for rank, row in enumerate(grand_rows, 1):
        row["rank"] = rank
        row["delta"] = final_delta_by_name.get(row["name"], 0)
        del row["_order"]

    return {
        "played": len(played_matches),
        "steps": len(progression),
        "table": grand_rows,
        "group_table": group_table,
        "progression": progression,
        "standings_ready": standings_ready,
        "has_bonus": standings_ready or bool(ko_pts_by_name),
    }


def _latest_knockout_result_match(data):
    results = data["knockout_results"]["matches"]
    winner_by_w = build_ko_winner_map(data["knockout_results"], data["knockouts"])
    played = []

    def add_match(m, phase_es, phase_en):
        result = results.get(m["code"])
        if not result or "score" not in result:
            return
        rh, ra = result["score"]
        pub = _knockout_public_schedule(m)
        resolve_knockout_fixture(pub, winner_by_w)
        played.append({
            "code": m["code"],
            "date": pub["date"],
            "dt": pub.get("dt", ""),
            "home": pub.get("resolved_home") or pub["fixture_home"],
            "away": pub.get("resolved_away") or pub["fixture_away"],
            "home_en": pub["fixture_home_en"],
            "away_en": pub["fixture_away_en"],
            "home_flag": pub.get("resolved_home_flag") or pub["fixture_home_flag"],
            "away_flag": pub.get("resolved_away_flag") or pub["fixture_away_flag"],
            "phase_es": phase_es,
            "phase_en": phase_en,
            "result": {"home": rh, "away": ra, "outcome": outcome(rh, ra)},
        })

    for rnd in data["knockouts"]["rounds"]:
        for m in rnd["matches"]:
            add_match(m, rnd["label_es"], rnd["label_en"])
    for m in data["knockouts"]["final_matches"]:
        add_match(m, m["label_es"], m["label_en"])
    played.sort(key=lambda m: (m["date"], m.get("dt", ""), m["code"]), reverse=True)
    return played[0] if played else None


def compute_recent_results(data, matches, limit=6):
    """Últimos partidos con resultado real y quién acertó/falló."""
    results = data["results"]
    names = data["names"]
    played = []

    def score_groups(picks, rh, ra, counts=None):
        """Reparte a cada persona en pleno/signo/palmada. En cruces con regla de
        rama, quien acertó el signo pero falló quién pasa va a `voided`
        (el "cementerio del cruce"): 0 puntos, pero distinto de una palmada."""
        ro = outcome(rh, ra)
        exact, sign, miss, voided = [], [], [], []
        for idx, (name, (h, a)) in enumerate(zip(names, picks)):
            counted = counts[idx] if counts is not None else True
            entry = {"name": name, "pick": f"{h}-{a}" if h is not None else "–"}
            if h is None or a is None:
                miss.append(entry)
            elif h == rh and a == ra:
                (exact if counted else voided).append(entry)
            elif outcome(h, a) == ro:
                (sign if counted else voided).append(entry)
            else:
                miss.append(entry)
        return exact, sign, miss, voided

    for m in matches:
        if m["code"] not in results:
            continue
        rh, ra = results[m["code"]]
        ro = outcome(rh, ra)
        exact, sign, miss, _ = score_groups(m["picks"], rh, ra)
        played.append({
            "code": m["code"], "group": m["group"], "date": m["date"],
            "dt": m.get("dt", ""),
            "home_en": m["home_en"], "away_en": m["away_en"],
            "home": m["home"], "away": m["away"],
            "home_flag": m["home_flag"], "away_flag": m["away_flag"],
            "result": {"home": rh, "away": ra, "outcome": ro},
            "exact": exact,
            "sign": sign,
            "miss": miss,
        })

    winner_by_w = build_ko_winner_map(data["knockout_results"], data["knockouts"])
    for rnd in data["knockouts"]["rounds"]:
        for m in rnd["matches"]:
            result = data["knockout_results"]["matches"].get(m["code"])
            if not result or "score" not in result:
                continue
            rh, ra = result["score"]
            ro = outcome(rh, ra)
            counts = _ko_matchup_counts(m, data)
            exact, sign, miss, voided = score_groups(m["score_picks"], rh, ra, counts)
            advance = []
            winner = result.get("winner")
            if winner:
                winner_key = _cmp_team(winner)
                advance = [
                    {"name": name, "pick": team_es(pick)}
                    for name, pick in zip(names, m["winner_picks"])
                    if pick and _cmp_team(pick) == winner_key
                ]
            pub = _knockout_public_schedule(m)
            resolve_knockout_fixture(pub, winner_by_w)
            played.append({
                "code": m["code"],
                "group": "",
                "date": pub["date"],
                "dt": pub.get("dt", ""),
                "home_en": pub["fixture_home_en"],
                "away_en": pub["fixture_away_en"],
                "home": pub.get("resolved_home") or pub["fixture_home"],
                "away": pub.get("resolved_away") or pub["fixture_away"],
                "home_flag": pub.get("resolved_home_flag") or pub["fixture_home_flag"],
                "away_flag": pub.get("resolved_away_flag") or pub["fixture_away_flag"],
                "result": {
                    "home": rh,
                    "away": ra,
                    "outcome": ro,
                    "winner": team_es(winner) if winner else None,
                    "winner_flag": team_flag(winner) if winner else "",
                },
                "exact": exact,
                "sign": sign,
                "miss": miss,
                "voided": voided,
                "advance": advance,
                "is_knockout": True,
                "scoreline_gated": rnd["key"] in ("r16", "qf", "sf"),
                "phase_es": rnd["label_es"],
                "phase_en": rnd["label_en"],
                "advance_points": rnd["advance_points"],
            })

    for m in data["knockouts"]["final_matches"]:
        result = data["knockout_results"]["matches"].get(m["code"])
        if not result or "score" not in result:
            continue
        rh, ra = result["score"]
        ro = outcome(rh, ra)
        counts = _ko_matchup_counts(m, data)
        exact, sign, miss, voided = score_groups(m["score_picks"], rh, ra, counts)
        pub = _knockout_public_schedule(m)
        resolve_knockout_fixture(pub, winner_by_w)
        played.append({
            "code": m["code"],
            "group": "",
            "date": pub["date"],
            "dt": pub.get("dt", ""),
            "home_en": pub["fixture_home_en"],
            "away_en": pub["fixture_away_en"],
            "home": pub.get("resolved_home") or pub["fixture_home"],
            "away": pub.get("resolved_away") or pub["fixture_away"],
            "home_flag": pub.get("resolved_home_flag") or pub["fixture_home_flag"],
            "away_flag": pub.get("resolved_away_flag") or pub["fixture_away_flag"],
            "result": {"home": rh, "away": ra, "outcome": ro},
            "exact": exact,
            "sign": sign,
            "miss": miss,
            "voided": voided,
            "advance": [],
            "is_knockout": True,
            "scoreline_gated": True,
            "phase_es": m["label_es"],
            "phase_en": m["label_en"],
            "advance_points": 0,
        })

    played.sort(key=lambda m: (m["date"], m.get("dt", ""), m["code"]), reverse=True)
    return {"matches": played[:limit], "total": len(played)}


def compute_today(data, matches):
    names = data["names"]
    all_matches = []

    # El primer partido de cada país ya mostró un fact con la lógica antigua
    # (m_num-1)%3. Partimos de ese índice para que los siguientes partidos no
    # repitan el ya visto.
    trivia_first_idx = {}
    for m in sorted(matches, key=lambda x: (x["date"], x.get("dt", ""), x["code"])):
        for team in (m["home_en"], m["away_en"]):
            if team not in trivia_first_idx:
                m_num = int(m["code"].split("-M")[1])
                trivia_first_idx[team] = (m_num - 1) % 3
    trivia_used = defaultdict(int)

    for m in matches:
        picks = []
        for i, (h, a) in enumerate(m["picks"]):
            picks.append({"name": names[i], "home": h, "away": a})
        o1 = sum(1 for h, a in m["picks"] if h is not None and h > a)
        ox = sum(1 for h, a in m["picks"] if h is not None and h == a)
        o2 = sum(1 for h, a in m["picks"] if h is not None and h < a)
        cnt = Counter(f"{h}-{a}" for h, a in m["picks"] if h is not None)
        modal = cnt.most_common(1)
        modal_sl = modal[0][0] if modal else "–"
        modal_pct = round(modal[0][1] / len(m["picks"]) * 100, 1) if modal else 0
        unique_picks = [(n, f"{h}-{a}") for n, (h, a) in zip(names, m["picks"])
                        if h is not None and cnt.get(f"{h}-{a}", 0) == 1]
        most_unique = None
        if unique_picks:
            best = max(unique_picks, key=lambda x: sum(int(g) for g in x[1].split("-")) + abs(int(x[1].split("-")[0]) - int(x[1].split("-")[1])))
            most_unique = {"name": best[0], "score": best[1]}
        t_idx_home = (trivia_first_idx.get(m["home_en"], 0) + trivia_used[m["home_en"]]) % 3
        t_idx_away = (trivia_first_idx.get(m["away_en"], 0) + trivia_used[m["away_en"]]) % 3
        trivia_used[m["home_en"]] += 1
        trivia_used[m["away_en"]] += 1
        home_trivia = TRIVIA.get(m["home_en"], [("", "")] * 3)[t_idx_home]
        away_trivia = TRIVIA.get(m["away_en"], [("", "")] * 3)[t_idx_away]
        all_matches.append({
            "code": m["code"], "group": m["group"], "date": m["date"],
            "time_es": m.get("time_es", ""), "time_uk": m.get("time_uk", ""),
            "next_es": m.get("next_es", False), "next_uk": m.get("next_uk", False),
            "dt": m.get("dt", ""),
            "home_en": m["home_en"], "away_en": m["away_en"],
            "home": m["home"], "away": m["away"],
            "home_flag": m["home_flag"], "away_flag": m["away_flag"],
            "picks": picks,
            "outcome_dist": {"1": o1, "X": ox, "2": o2},
            "modal_scoreline": modal_sl,
            "modal_scoreline_share": modal_pct / 100,
            "most_unique_pick": most_unique,
            "home_trivia": {"es": home_trivia[0], "en": home_trivia[1]},
            "away_trivia": {"es": away_trivia[0], "en": away_trivia[1]},
        })
    all_matches.sort(key=lambda x: (x["date"], x.get("dt", ""), x["code"]))
    return {"matches": all_matches}


def compute_match_stake(match, results, live_table):
    """Cuánto puede mover el ranking un partido de grupos (escenarios realistas)."""
    if match["code"] in results or not live_table:
        return None
    picks_by_name = {p["name"]: p for p in match["picks"]}
    scenarios = _group_match_scenarios(match)
    if not scenarios:
        return None
    names = [row["name"] for row in live_table]

    def points_for_name(name, scenario):
        pick = picks_by_name.get(name)
        if not pick or pick["home"] is None:
            return None
        rh, ra = scenario
        return _group_points_for_pick(pick["home"], pick["away"], rh, ra)

    return _compute_scenario_stake(
        names,
        live_table,
        scenarios,
        points_for_name,
        _format_group_scenario,
    )


# --------------------------------------------------------------------------
# main (parte de verificación; el render se añade después)
# --------------------------------------------------------------------------

CSS = r"""
*{box-sizing:border-box;margin:0;padding:0}
:root{
  --bg:#042503;--bg2:#021802;--teal:#0a3a08;--surface:#0a3a08;--surface2:#0d4a0a;
  --mint:#6ce869;--mint2:#2fbf2c;--mint-soft:#d1fbd0;--line:rgba(108,232,105,.15);
  --text:#eef8ec;--muted:#9dbf9b;--gold:#f28536;--red:#ff8a7a;--maxw:1060px;
  --fd:'Shippori Mincho',Georgia,'Times New Roman',serif;
  --fu:'DM Sans',system-ui,-apple-system,Segoe UI,Roboto,sans-serif;
  --fm:'DM Mono',ui-monospace,'SF Mono',Menlo,monospace;
}
html{scroll-behavior:smooth}
body{background:radial-gradient(1200px 800px at 75% -10%,#0d4a0a 0%,var(--bg) 45%,var(--bg2) 100%);
  color:var(--text);font-family:var(--fu);
  line-height:1.5;-webkit-font-smoothing:antialiased;min-height:100vh;overflow-x:hidden}
h1,h2,h3,.disp{font-family:var(--fd);font-weight:600;letter-spacing:-.005em;line-height:1.1}
a{color:inherit}
.mint{color:var(--mint)}.muted{color:var(--muted)}
/* APP WRAP — top-tab structure (see top bar styles below) */
.wrap{margin:0;padding:0 0 80px}
/* HERO */
header.hero{min-height:100vh;display:flex;flex-direction:column;justify-content:center;position:relative;padding-top:40px}
.hero .eyebrow{font-family:var(--fm);color:var(--mint);font-weight:500;letter-spacing:.22em;text-transform:uppercase;font-size:.78rem;margin-bottom:18px}
.hero h1{font-size:clamp(2.6rem,8vw,5.6rem);background:linear-gradient(120deg,#fff 10%,var(--mint) 90%);
  -webkit-background-clip:text;background-clip:text;-webkit-text-fill-color:transparent;margin-bottom:10px}
.hero .lead{font-size:clamp(1rem,2vw,1.3rem);color:var(--muted);max-width:620px;margin-bottom:42px}
.chips{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;max-width:920px}
.chip{background:var(--surface);border:1px solid var(--line);border-radius:16px;padding:18px 22px;min-width:0}
.chip .big{white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.chip .big{font-family:var(--fd);font-size:2.1rem;font-weight:600;color:var(--mint)}
.chip .lab{font-size:.78rem;color:var(--muted);text-transform:uppercase;letter-spacing:.08em;margin-top:4px}
.scrollcue{position:absolute;bottom:26px;left:0;color:var(--muted);font-size:.8rem;display:flex;align-items:center;gap:8px;animation:bob 1.8s ease-in-out infinite}
@keyframes bob{0%,100%{transform:translateY(0)}50%{transform:translateY(6px)}}
/* HERO CONNECTION NET — brand motif: dots joined by lines ("todo está conectado") */
header.hero::before{content:"";position:absolute;inset:0;z-index:0;pointer-events:none;
  background-image:linear-gradient(var(--line) 1px,transparent 1px),linear-gradient(90deg,var(--line) 1px,transparent 1px);
  background-size:66px 66px;opacity:.55;
  -webkit-mask-image:radial-gradient(125% 95% at 72% 4%,#000 34%,transparent 80%);
  mask-image:radial-gradient(125% 95% at 72% 4%,#000 34%,transparent 80%)}
header.hero>*{position:relative;z-index:1}
/* Interactive brand constellation — dots + lines that react to the cursor (sits behind all content) */
#bgnet{position:fixed;inset:0;z-index:-1;pointer-events:none;opacity:.62}
/* SECTIONS */
section.sec{padding:74px 0;border-top:1px solid var(--line)}
.sec-head{margin-bottom:34px}
.kicker{font-family:var(--fm);color:var(--mint);font-weight:500;font-size:.82rem;letter-spacing:.12em;text-transform:uppercase;margin-bottom:8px}
.sec h2{font-size:clamp(1.8rem,4vw,2.9rem)}
.sec .sub{color:var(--muted);max-width:640px;margin-top:10px;font-size:1.02rem}
.grid{display:grid;gap:16px}
.g2{grid-template-columns:repeat(2,1fr)}.g3{grid-template-columns:repeat(3,1fr)}.g4{grid-template-columns:repeat(4,1fr)}
.card{background:var(--surface);border:1px solid var(--line);border-radius:18px;padding:22px}
.card.glow{box-shadow:0 0 0 1px rgba(108,232,105,.2),0 20px 60px rgba(108,232,105,.08)}
.card h3{font-size:1.05rem;margin-bottom:4px}
.card .k{font-size:.74rem;color:var(--muted);text-transform:uppercase;letter-spacing:.1em}
.big-num{font-family:var(--fd);font-size:2.6rem;font-weight:600;color:var(--mint);line-height:1}
/* BARS */
.bar-row{display:grid;grid-template-columns:30px minmax(84px,170px) 1fr 54px;align-items:center;gap:14px;padding:7px 0}
.bar-rank{font-family:var(--fu);color:var(--muted);font-size:.95rem;text-align:right;font-weight:700}
.bar-name{font-weight:600;font-size:.95rem;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.bar-co{display:flex;flex-direction:column;gap:3px;min-width:0}
.bar-track{height:13px;background:rgba(255,255,255,.06);border-radius:8px;overflow:hidden}
.bar-fill{height:100%;width:0;border-radius:8px;background:linear-gradient(90deg,var(--mint2),var(--mint));transition:width 1s cubic-bezier(.2,.8,.2,1)}
.bar-fill.cool{background:linear-gradient(90deg,#2b6f7f,#4a9aa8)}
.bar-fill.gold{background:linear-gradient(90deg,#c45f14,var(--gold))}
.bar-val{font-family:var(--fu);font-weight:700;text-align:right;font-size:.95rem}
/* live ranking (?view=live → #aciertos): match-by-match race only */
.rank-proto-card{background:rgba(0,26,31,.34);border:1px solid var(--line);border-radius:18px;padding:18px;margin-top:18px;overflow:visible}
.rank-proto-top{display:flex;justify-content:space-between;gap:16px;align-items:flex-start;margin-bottom:14px}
.rank-proto-top h3{font-size:1.08rem;margin-bottom:4px}
.rr-breakdown{display:block;font-size:.7rem;font-weight:500;color:var(--muted);letter-spacing:.01em;margin-top:1px}
.rank-delta{font-family:var(--fu);font-size:.72rem;color:var(--muted);white-space:nowrap}
.rank-delta.up{color:var(--mint)}.rank-delta.down{color:var(--red)}
.race-layout{display:grid;grid-template-columns:minmax(220px,280px) 1fr;gap:16px;align-items:start;overflow:visible}
.race-control{background:linear-gradient(160deg,rgba(108,232,105,.08),rgba(242,133,54,.055));border:1px solid var(--line);border-radius:14px;padding:14px}
.race-control input{width:100%;accent-color:var(--mint);margin:12px 0}
.race-actions{display:grid;grid-template-columns:36px 36px 1fr 36px;gap:8px;align-items:center;margin:12px 0}
.race-actions button{height:36px;border:1px solid var(--line);border-radius:50%;background:#021802;color:var(--mint);
  font:800 .98rem var(--fu);cursor:pointer}
.race-actions .race-play{background:var(--mint);color:#021802;border-color:transparent}
.race-step{font:800 .8rem var(--fu);color:var(--gold);text-align:center;letter-spacing:.06em}
.race-match{font-weight:700;margin-top:8px;line-height:1.25}
.race-now{display:grid;grid-template-columns:repeat(3,1fr);gap:7px;margin-top:12px}
.race-now span{background:rgba(0,0,0,.18);border:1px solid rgba(108,232,105,.12);border-radius:10px;padding:8px;font-size:.76rem;color:var(--muted)}
.race-now b{display:block;color:var(--text);font-family:var(--fu);font-size:1rem}
.race-legend{display:flex;flex-wrap:wrap;gap:8px;margin-top:12px;font-size:.76rem;color:var(--muted)}
.race-legend span{display:inline-flex;align-items:center;gap:6px}.race-dot{width:9px;height:9px;border-radius:50%;display:inline-block}
.race-board{display:grid;gap:8px;position:relative;z-index:0;overflow:visible}
.race-board-wrap{display:flex;flex-direction:column;gap:8px;min-width:0}
.race-row{display:grid;grid-template-columns:34px minmax(82px,170px) 68px minmax(42px,1fr) 42px 42px;gap:9px;align-items:center;background:rgba(255,255,255,.04);
  border:1px solid rgba(108,232,105,.08);border-left:4px solid var(--runner);border-radius:10px;padding:7px 10px;min-width:0;
  position:relative;will-change:transform,box-shadow;transition:border-color .25s ease,background .25s ease}
.race-row .bar-track{height:10px}.race-fill{height:100%;border-radius:8px;background:linear-gradient(90deg,color-mix(in srgb,var(--runner) 45%,#021802),var(--runner));
  transition:width .58s cubic-bezier(.2,.8,.2,1)}
.race-row.leader{border-color:rgba(242,133,54,.45);border-left-color:var(--runner);background:rgba(242,133,54,.06)}
.race-row .rank-delta{justify-self:center;background:rgba(0,0,0,.18);border-radius:999px;padding:3px 7px;font-size:.74rem}
.race-row.is-moving{z-index:5}
.race-row.moved-up{z-index:3;animation:rankGlowUp .76s ease both}.race-row.moved-down{z-index:2;animation:rankGlowDown .76s ease both}
.race-row.moved-up .rank-delta{box-shadow:0 0 0 1px rgba(108,232,105,.25),0 0 18px rgba(108,232,105,.16)}
.race-row.moved-down .rank-delta{box-shadow:0 0 0 1px rgba(255,122,122,.25),0 0 18px rgba(255,122,122,.13)}
@keyframes rankGlowUp{0%,100%{box-shadow:0 0 0 rgba(108,232,105,0)}45%{box-shadow:0 0 0 1px rgba(108,232,105,.28),0 0 24px rgba(108,232,105,.16)}}
@keyframes rankGlowDown{0%,100%{box-shadow:0 0 0 rgba(255,122,122,0)}45%{box-shadow:0 0 0 1px rgba(255,122,122,.24),0 0 24px rgba(255,122,122,.13)}}
.race-round{font-family:var(--fu);font-weight:800;color:var(--runner);text-align:right;font-size:.88rem}
.race-row .race-tip{position:absolute;left:50%;bottom:calc(100% + 8px);transform:translateX(-50%) translateY(5px);
  background:#021403;border:1px solid rgba(108,232,105,.28);border-radius:11px;padding:10px 12px;display:flex;flex-direction:column;gap:8px;
  min-width:200px;box-shadow:0 14px 34px rgba(0,0,0,.5);opacity:0;pointer-events:none;
  transition:opacity .16s ease,transform .16s ease;z-index:20}
.race-row .race-tip::after{content:'';position:absolute;left:50%;top:100%;transform:translateX(-50%);
  border:6px solid transparent;border-top-color:#021403}
.race-tip .rt-block{display:flex;flex-direction:column;gap:5px}
.race-tip .rt-block+.rt-block{padding-top:8px;border-top:1px solid rgba(108,232,105,.14)}
.race-tip .rt-head{font:800 .62rem var(--fu);letter-spacing:.09em;text-transform:uppercase;color:var(--muted)}
.race-tip .rt-items{display:flex;flex-wrap:wrap;gap:6px 10px}
.race-tip .rt-item{display:flex;flex-direction:column;align-items:flex-start;gap:1px;font-size:.62rem;
  color:var(--muted);text-transform:uppercase;letter-spacing:.06em;white-space:nowrap}
.race-tip .rt-item b{font-family:var(--fu);font-size:1rem;line-height:1.1;color:var(--text)}
.race-tip .rt-item .rt-pts{font-size:.92rem;margin-top:1px}
.race-tip .rt-sum{font:700 .72rem var(--fu);color:var(--gold);letter-spacing:.02em;text-transform:none}
.race-tip .rt-pleno b,.race-tip .rt-pleno .rt-pts{color:var(--gold)}
.race-tip .rt-sign b,.race-tip .rt-sign .rt-pts{color:var(--mint)}
.race-tip .rt-bonus b,.race-tip .rt-bonus .rt-pts{color:#75e0ff}
.race-tip .rt-ko b,.race-tip .rt-ko .rt-pts{color:#c9a8ff}
.race-row:hover{z-index:30;cursor:default}
.race-row:hover .race-tip{opacity:1;transform:translateX(-50%) translateY(0)}
/* PODIUM */
.podium{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;align-items:end;margin-bottom:30px}
.pod{background:var(--surface);border:1px solid var(--line);border-radius:16px;padding:20px;text-align:center;position:relative}
.pod .medal{font-size:1.6rem}.pod .nm{font-weight:700;font-size:1.1rem;margin:6px 0 2px}
.pod .sc{font-family:var(--fu);font-size:2rem;font-weight:700;color:var(--mint)}
.pod.p1{transform:translateY(-12px);box-shadow:0 0 0 1px rgba(108,232,105,.25),0 18px 50px rgba(108,232,105,.1)}
/* HEADLINE pair cards */
.duo{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:26px}
.duo .card{display:flex;flex-direction:column;gap:6px}
.duo .names{font-family:var(--fu);font-size:1.5rem;font-weight:700}
.tag{display:inline-block;font-size:.72rem;font-weight:700;text-transform:uppercase;letter-spacing:.08em;
  padding:4px 9px;border-radius:999px;background:rgba(108,232,105,.12);color:var(--mint);width:fit-content}
.tag.warm{background:rgba(242,133,54,.14);color:var(--gold)}
.tag.cool{background:rgba(120,180,200,.14);color:#8aa889}
/* MATRIX */
.matrix-wrap{overflow-x:auto;padding-bottom:8px}
.matrix{display:inline-grid;gap:2px;margin-top:8px}
.mcell{width:19px;height:19px;border-radius:3px;background:var(--surface2);cursor:default}
.mlabel{font-size:.62rem;color:var(--muted);display:flex;align-items:center}
.mlabel.row{justify-content:flex-end;padding-right:6px;white-space:nowrap}
.mlabel.col{writing-mode:vertical-rl;transform:rotate(180deg);justify-content:flex-end;padding-bottom:6px;height:64px}
.mdiag{background:repeating-linear-gradient(45deg,#0a3a08,#0a3a08 3px,#0d4a0a 3px,#0d4a0a 6px)}
#mtip{position:fixed;pointer-events:none;z-index:90;background:#021802;border:1px solid var(--line);
  border-radius:10px;padding:8px 12px;font-size:.82rem;opacity:0;transition:opacity .12s;box-shadow:0 8px 30px rgba(0,0,0,.5)}
.legend{display:flex;align-items:center;gap:10px;margin-top:14px;font-size:.78rem;color:var(--muted)}
.legend .scale{height:10px;width:140px;border-radius:6px;background:linear-gradient(90deg,var(--surface2),var(--mint))}
/* match distribution */
.match-row{padding:14px 0;border-bottom:1px solid var(--line)}
.match-row:last-child{border:0}
.match-top{display:flex;justify-content:space-between;gap:10px;font-weight:600;margin-bottom:8px;font-size:.96rem}
.dist{display:flex;height:24px;border-radius:7px;overflow:hidden;font-size:.72rem;font-weight:700}
.dist span{display:flex;align-items:center;justify-content:center;color:#012;min-width:0;transition:width 1s cubic-bezier(.2,.8,.2,1)}
.dist .s1{background:var(--mint)}.dist .sx{background:#8aa889}.dist .s2{background:var(--gold)}
.modal-pill{color:var(--muted);font-size:.84rem;margin-top:6px}
/* chips teams */
.teamchip{display:inline-flex;align-items:center;gap:7px;background:var(--surface);border:1px solid var(--line);
  border-radius:999px;padding:7px 13px;font-size:.9rem;font-weight:600;margin:4px 4px 0 0}
.teamchip .n{color:var(--muted);font-family:var(--fu);font-weight:700}
/* group cards */
.gcard{background:var(--surface);border:1px solid var(--line);border-radius:14px;padding:16px}
.gcard .gl{font-family:var(--fu);color:var(--muted);font-size:.8rem;letter-spacing:.1em}
.gcard .fav{font-size:1.15rem;font-weight:700;margin:4px 0 10px;display:flex;align-items:center;gap:8px}
/* knockouts */
.ko-summary{display:grid;grid-template-columns:1.3fr 1fr 1fr;gap:16px;margin-bottom:22px}
.ko-hero{background:linear-gradient(160deg,rgba(242,133,54,.14),rgba(108,232,105,.08));border:1px solid rgba(242,133,54,.3)}
.ko-hero .fav{font-family:var(--fu);font-size:2rem;font-weight:800;color:var(--gold);margin:8px 0 4px}
.ko-round{margin-top:22px}
.ko-round h3{font-size:1.15rem;margin-bottom:12px}
.ko-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(210px,1fr));gap:10px}
.ko-match{background:rgba(0,0,0,.14);border:1px solid rgba(108,232,105,.12);border-radius:13px;padding:12px}
.ko-code{font-family:var(--fu);font-size:.76rem;color:var(--muted);letter-spacing:.08em;margin-bottom:7px}
.ko-main{font-weight:800;font-size:1rem;min-height:1.45em}
.ko-mini{color:var(--muted);font-size:.78rem;margin-top:5px}
.ko-pills{display:flex;flex-wrap:wrap;gap:7px;margin-top:12px}
.ko-pill{display:inline-flex;align-items:center;gap:6px;border:1px solid var(--line);border-radius:999px;padding:5px 9px;background:rgba(255,255,255,.045);font-size:.8rem}
.ko-score{font-family:var(--fu);font-weight:800;color:var(--mint)}
/* fichas */
.search{width:100%;max-width:340px;background:var(--surface);border:1px solid var(--line);color:var(--text);
  border-radius:12px;padding:11px 15px;font-size:.95rem;margin-bottom:20px;font-family:inherit}
.search:focus{outline:none;border-color:var(--mint)}
.ficha{background:var(--surface);border:1px solid var(--line);border-radius:16px;padding:18px;transition:.2s}
.ficha:hover{border-color:rgba(108,232,105,.4);transform:translateY(-3px)}
.ficha .fh{display:flex;justify-content:space-between;align-items:start;margin-bottom:10px}
.ficha .fn{font-size:1.2rem;font-weight:700}
.ficha .rk{font-family:var(--fu);font-size:.8rem;color:var(--muted)}
.ficha .lab{font-size:.8rem;margin:2px 0 12px}
.fstats{display:grid;grid-template-columns:1fr 1fr;gap:8px;font-size:.84rem}
.fstats .v{font-family:var(--fu);font-weight:700;color:var(--mint)}
.fline{font-size:.82rem;color:var(--muted);margin-top:10px;border-top:1px solid var(--line);padding-top:10px}
.fline b{color:var(--text)}
/* premios */
.award{background:linear-gradient(160deg,var(--surface2),var(--surface));border:1px solid var(--line);
  border-radius:18px;padding:22px;text-align:center}
.award .em{font-size:2.4rem}.award .ti{font-weight:700;margin:8px 0 2px}
.award .wn{font-family:var(--fu);font-size:1.5rem;font-weight:700;color:var(--mint)}
.award .dt{font-size:.82rem;color:var(--muted);margin-top:2px}
/* teaser */
.teaser{text-align:center;padding:46px 22px}
.teaser .em{font-size:3rem}
/* reveal */
.reveal{opacity:0;transform:translateY(20px);transition:opacity .65s ease,transform .65s ease}
.reveal.in{opacity:1;transform:none}
footer{border-top:1px solid var(--line);padding:36px 0 60px;color:var(--muted);font-size:.82rem}
footer .brand{color:var(--text);height:20px;display:inline-block;vertical-align:middle;margin-right:8px}
footer .brand svg{height:20px;width:auto}
@media(max-width:900px){
  .wrap{padding:0 0 70px}
  .g3,.g4{grid-template-columns:repeat(2,1fr)}.duo,.podium,.ko-summary{grid-template-columns:1fr}
  .chips{grid-template-columns:repeat(2,1fr)}
  .race-layout{grid-template-columns:1fr}
}
@media(max-width:560px){
  .g2,.g3,.g4{grid-template-columns:1fr}
  .race-row{grid-template-columns:28px minmax(58px,1fr) 64px 34px 34px;gap:7px}
  .race-row .bar-track{grid-column:2 / -1;grid-row:2}
  .race-round,.race-row .bar-val{font-size:.8rem}
  .race-row .rank-delta{font-size:.68rem;padding:3px 6px}
}
.today-date{font-family:var(--fu);font-size:1.1rem;color:var(--mint);margin-bottom:0;letter-spacing:.04em}
.hoy-nav{display:flex;align-items:center;justify-content:center;gap:14px;margin-bottom:22px}
.hoy-nav-mid{display:flex;flex-direction:column;align-items:center;gap:6px;min-width:0;text-align:center}
.hoy-arrow{flex:0 0 auto;width:38px;height:38px;border-radius:50%;border:1px solid var(--line);background:#021802;color:var(--mint);
  font:800 1.1rem var(--fu);cursor:pointer;display:flex;align-items:center;justify-content:center;transition:opacity .15s ease}
.hoy-arrow:disabled{opacity:.28;cursor:default}
.hoy-today-btn{border:1px solid var(--line);background:rgba(108,232,105,.1);color:var(--mint);border-radius:999px;
  font:700 .72rem var(--fu);letter-spacing:.04em;padding:4px 12px;cursor:pointer}
@media(max-width:560px){.hoy-nav{gap:8px}.today-date{font-size:.98rem}}
.today-match{background:var(--surface);border:1px solid var(--line);border-radius:18px;padding:24px;margin-bottom:20px}
.today-match .tm-head{display:flex;align-items:center;justify-content:space-between;margin-bottom:16px;flex-wrap:wrap;gap:8px}
.today-match .tm-teams{font-family:var(--fu);font-size:1.3rem;font-weight:700;display:flex;align-items:center;gap:8px}
.today-match .tm-group{font-size:.78rem;color:var(--muted);background:rgba(108,232,105,.08);padding:4px 10px;border-radius:8px;letter-spacing:.06em}
.today-match .tm-tags{display:flex;gap:8px;align-items:center;flex-wrap:wrap}
.today-match .tm-time{font-size:.78rem;font-weight:700;color:var(--mint);background:rgba(108,232,105,.14);padding:4px 10px;border-radius:8px;letter-spacing:.04em}
.tm-next{font-size:.62rem;font-weight:700;color:var(--bg,#021802);background:var(--mint);border-radius:5px;padding:1px 4px;margin-left:3px;vertical-align:top}
.today-match .tm-stats{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:18px}
.today-match .tm-stat{background:rgba(0,0,0,.15);border-radius:12px;padding:12px;text-align:center}
.today-match .tm-stat .val{font-family:var(--fu);font-size:1.3rem;font-weight:700;color:var(--mint)}
.today-match .tm-stat .lab{font-size:.72rem;color:var(--muted);text-transform:uppercase;letter-spacing:.06em;margin-top:2px}
.today-match .tm-stat.tm-stat-dead{opacity:.5;background:rgba(0,0,0,.28)}
.today-match .tm-stat.tm-stat-dead .val{color:var(--muted)}
.today-match .tm-stat.tm-stat-dead .lab{text-decoration:line-through;text-decoration-color:var(--muted)}
.today-match .tm-stat.tm-stat-dead .dead-x{color:var(--red);text-decoration:none;font-weight:700;margin-left:2px}
.today-picks{margin-top:14px}
.today-picks .tp-title{font-size:.82rem;color:var(--muted);text-transform:uppercase;letter-spacing:.08em;margin-bottom:10px}
.pick-groups{margin-top:4px}
.pick-groups .result-person b{max-width:none}
.pick-groups .result-box.draw .rb-count{color:var(--gold)}
.recent-meta{display:flex;align-items:center;gap:12px;flex-wrap:wrap;margin-bottom:14px;color:var(--muted);font-size:.86rem}
.score-final{font-family:var(--fu);font-weight:700;font-size:1.7rem;color:var(--gold);white-space:nowrap}
.result-groups{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:12px;margin-top:18px}
.result-box{background:rgba(0,0,0,.15);border-radius:12px;padding:13px;min-width:0}
.result-box .rb-title{font-size:.72rem;color:var(--muted);text-transform:uppercase;letter-spacing:.06em;margin-bottom:8px}
.result-box .rb-count{font-family:var(--fu);font-weight:700;color:var(--mint)}
.result-box.miss .rb-count{color:var(--red)}
.result-box.voided,.result-box.dead{background:rgba(150,120,190,.1);border:1px solid rgba(160,130,200,.28)}
.result-box.voided .rb-count,.result-box.dead .rb-count{color:#c9a9ff}
.result-box .rb-sub{font-size:.68rem;color:var(--muted);margin:-4px 0 8px;font-style:italic}
.result-names{display:flex;flex-wrap:wrap;gap:6px}
.result-person{display:inline-flex;gap:5px;align-items:center;background:rgba(255,255,255,.06);border:1px solid var(--line);border-radius:999px;padding:4px 8px;font-size:.78rem;min-width:0}
.result-person b{font-weight:700;max-width:92px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.result-person span{font-family:var(--fu);color:var(--muted);white-space:nowrap}
.no-today{text-align:center;padding:46px 22px}
.no-today .em{font-size:3rem}
.no-today-img{width:100%;max-width:340px;border-radius:14px;margin:0 auto 6px;display:block;box-shadow:0 8px 24px rgba(0,0,0,.3)}
.no-today .next{margin-top:22px;text-align:left}
.no-today .next-match{display:flex;align-items:center;gap:10px;padding:10px 0;border-bottom:1px solid var(--line);font-size:.92rem}
.no-today .next-date{color:var(--muted);font-size:.78rem;min-width:80px}
@media(max-width:720px){.result-groups{grid-template-columns:1fr}}
.trivia-block{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-top:16px;padding-top:16px;border-top:1px solid var(--line)}
.trivia-item{background:rgba(0,0,0,.15);border-radius:12px;padding:12px 14px;font-size:.84rem;line-height:1.45}
.trivia-item .trivia-flag{font-size:.72rem;color:var(--muted);text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px}
.trivia-item .trivia-text{color:var(--text)}
@media(max-width:560px){.trivia-block{grid-template-columns:1fr}}
/* ============================================================
   APP SHELL — top tabs (En directo / Eliminatorias / Fase de grupos)
   + knockout bracket visualization.
   ============================================================ */
.proto-shell{position:sticky;top:0;z-index:80;background:rgba(3,20,3,.86);backdrop-filter:blur(14px);border-bottom:1px solid var(--line)}
.proto-shell-inner{max-width:1180px;margin:0 auto;display:flex;align-items:center;gap:14px;padding:12px 22px;flex-wrap:wrap}
.proto-shell .brand{display:flex;align-items:center;color:var(--text);height:26px}
.proto-shell .brand svg{height:26px;width:auto}
.proto-tabs{display:flex;gap:6px;flex:1;flex-wrap:wrap}
.proto-tab{border:1px solid var(--line);background:rgba(255,255,255,.04);color:var(--muted);border-radius:999px;
  padding:9px 16px;font:800 .82rem var(--fu);cursor:pointer;display:flex;align-items:center;gap:8px}
.proto-tab .pt-em{font-size:.95rem}
.proto-tab.on{background:var(--mint);border-color:transparent;color:#021802}
.proto-tab .pt-badge{font-size:.66rem;background:rgba(0,0,0,.18);color:inherit;border-radius:999px;padding:1px 7px}
.proto-tab.on .pt-badge{background:rgba(0,26,31,.18)}
.proto-lang{display:flex;gap:4px}
.proto-lang button{border:1px solid var(--line);background:transparent;color:var(--muted);border-radius:8px;padding:5px 9px;font:800 .72rem var(--fu);cursor:pointer}
.proto-lang button.on{background:var(--mint);color:#021802;border-color:transparent}
.proto-body{max-width:1180px;margin:0 auto;padding:0 22px}
/* bracket visualization */
.bracket-wrap{background:linear-gradient(160deg,rgba(0,26,31,.45),rgba(6,58,68,.5));border:1px solid var(--line);
  border-radius:20px;padding:20px;overflow-x:auto;margin-top:8px}
/* --- desktop: real tournament tree with measured SVG connectors --- */
.bk-tree{position:relative;display:flex;gap:34px;min-width:max-content;align-items:stretch}
.bk-lines{position:absolute;left:0;top:0;pointer-events:none;z-index:0;overflow:visible}
.bk-lines path{fill:none;stroke:rgba(108,232,105,.3);stroke-width:1.5}
.bk-col{position:relative;z-index:1;display:flex;flex-direction:column;min-width:184px}
.bk-col-head{font:800 .72rem var(--fu);letter-spacing:.08em;text-transform:uppercase;color:var(--muted);text-align:center;margin-bottom:12px}
.bk-col-body{flex:1;display:flex;flex-direction:column;justify-content:space-around}
.bk-finalcol .bk-col-body{justify-content:center}
.bk-finalcol .bk-col-head{margin:0 0 8px}
.bk-third-block{margin-top:22px}
.bk-match{position:relative;background:rgba(0,0,0,.3);border:1px solid rgba(108,232,105,.16);border-radius:12px;padding:9px 11px;margin:7px 0}
.bk-match.played{border-color:rgba(108,232,105,.38);background:rgba(0,26,31,.42)}
.bk-match.final{border-color:rgba(242,133,54,.42);background:linear-gradient(160deg,rgba(242,133,54,.13),rgba(0,0,0,.28))}
.bk-match.third{border-color:rgba(255,255,255,.14);opacity:.9}
.bk-code{font:700 .62rem var(--fu);color:var(--muted);letter-spacing:.05em;margin-bottom:6px;display:flex;justify-content:space-between;gap:6px}
.bk-side{display:flex;align-items:center;gap:7px;font-size:.84rem;font-weight:700;padding:2px 0}
.bk-side .bk-flag{width:18px;text-align:center;flex:none}
.bk-side .bk-nm{overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.bk-side .bk-goals{margin-left:auto;font:800 .78rem var(--fu);color:var(--text);flex:none}
.bk-side .bk-pct{margin-left:auto;font:800 .72rem var(--fu);color:var(--mint);flex:none}
.bk-side.dim{color:var(--muted);font-weight:600}
.bk-side.won{color:var(--mint)}
.bk-side.out{opacity:.42;font-weight:600}
.bk-cbar{height:5px;border-radius:999px;background:rgba(255,255,255,.08);margin:7px 0 2px;overflow:hidden}
.bk-cbar span{display:block;height:100%;background:linear-gradient(90deg,var(--mint2),var(--mint))}
.bk-cbar.split{display:flex}
.bk-cbar.split span{flex:none}
.bk-cbar .away{background:var(--gold)}
.bk-tip{color:var(--muted);font-size:.68rem;margin-top:4px}
.bk-tip.bk-real{color:var(--gold);font-weight:600}
.bk-match.bk-has-tip,.ko-match.bk-has-tip{cursor:help}
.bk-match:not(.played).bk-has-tip:hover{border-color:rgba(108,232,105,.45)}
#bktip{position:fixed;pointer-events:none;z-index:95;background:#021802;border:1px solid var(--line);
  border-radius:12px;padding:10px 12px;font-size:.78rem;opacity:0;transition:opacity .12s;
  box-shadow:0 8px 30px rgba(0,0,0,.5);max-width:300px;line-height:1.45}
.bk-ht-title{font:800 .72rem var(--fu);color:var(--mint);margin-bottom:8px;letter-spacing:.04em}
.bk-ht-slot{margin-bottom:8px}.bk-ht-slot:last-child{margin-bottom:0}
.bk-ht-lbl{font:700 .68rem;color:var(--muted);margin-bottom:4px}
.bk-ht-row{display:flex;justify-content:space-between;gap:10px;padding:2px 0}
.bk-ht-row span:last-child{color:var(--mint);font-weight:700;font-size:.72rem}
/* --- mobile: round chips + full-width match list --- */
.bk-mobile{display:none}
.bk-chips{display:flex;gap:7px;overflow-x:auto;padding-bottom:12px;margin-bottom:4px;-webkit-overflow-scrolling:touch;scrollbar-width:none}
.bk-chips::-webkit-scrollbar{display:none}
.bk-chip{flex:none;border:1px solid var(--line);background:rgba(255,255,255,.04);color:var(--muted);
  border-radius:999px;padding:8px 15px;font:800 .82rem var(--fu);cursor:pointer;white-space:nowrap}
.bk-chip.on{background:var(--mint);border-color:transparent;color:#021802}
.bk-list{display:flex;flex-direction:column;gap:10px}
.bk-list[hidden]{display:none}
.bk-list .bk-match{margin:0}
@media(max-width:760px){
  .bracket-wrap{padding:14px;overflow:hidden}
  .bk-tree{display:none}
  .bk-mobile{display:block}
}
/* survival timeline */
.survive{margin-top:22px;background:rgba(0,0,0,.18);border:1px solid var(--line);border-radius:18px;padding:20px}
.survive h3{font:800 1.1rem var(--fu);margin-bottom:4px}
.survive .demo-pill{display:inline-block;font:800 .64rem var(--fu);letter-spacing:.05em;text-transform:uppercase;
  background:rgba(242,133,54,.15);color:var(--gold);border-radius:999px;padding:3px 9px;margin-bottom:14px}
.surv-grid{display:grid;grid-template-columns:130px repeat(var(--rounds,5),1fr);gap:0 6px;align-items:center;min-width:560px}
.surv-head{font:800 .66rem var(--fu);letter-spacing:.05em;text-transform:uppercase;color:var(--muted);text-align:center;padding-bottom:8px}
.surv-name{font-weight:700;font-size:.82rem;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;padding:5px 0}
.surv-cell{height:12px;border-radius:4px;margin:3px 2px;background:rgba(108,232,105,.1);position:relative;overflow:hidden}
.surv-cell.br-stack{display:block}
.surv-cell.br-stack .br-layer{position:absolute;inset:0}
.surv-cell.br-stack .br-ok{background:linear-gradient(90deg,var(--mint2),var(--mint))}
.surv-cell.br-stack .br-bad{background:#ff8e7d}
.surv-cell.br-stack .br-drift{background:repeating-linear-gradient(-45deg,rgba(255,142,125,.35),rgba(255,142,125,.35) 4px,rgba(255,142,125,.08) 4px,rgba(255,142,125,.08) 8px)}
.surv-legend{display:flex;flex-wrap:wrap;gap:10px 16px;margin-top:14px;font-size:.74rem;color:var(--muted)}
.surv-legend span{display:inline-flex;align-items:center;gap:6px}
.surv-legend i{display:inline-block;width:18px;height:10px;border-radius:3px}
.surv-legend .lg-grad{background:linear-gradient(90deg,var(--mint2) 0%,var(--mint) 70%,rgba(108,232,105,.1) 70%,#ff8e7d 100%)}
.survive .surv-scroll{overflow-x:auto}

/* ---- Eliminatorias: relato + métricas KO ---- */
.koproto-note{background:rgba(242,133,54,.12);border:1px solid rgba(242,133,54,.3);color:var(--gold);
  border-radius:12px;padding:11px 15px;font-size:.86rem;margin-bottom:18px}
.kp-dossier{background:rgba(0,0,0,.18);border:1px solid var(--line);border-radius:18px;padding:18px}
.kp-chips{display:flex;flex-wrap:wrap;gap:8px;margin-bottom:16px}
.kp-teamchip{border:1px solid var(--line);background:transparent;color:var(--muted);border-radius:999px;
  padding:6px 13px;font:800 .8rem var(--fu);cursor:pointer}
.kp-teamchip.on{background:var(--mint);color:#021802;border-color:var(--mint)}
.kp-dossier-head{font:800 1rem var(--fu);margin:0 0 12px}
.kp-scatter{width:100%;height:auto}
.kp-scatter text{fill:var(--muted);font:700 10px var(--fu)}
.kp-scatter .axis{stroke:var(--line)}
.kp-grave{display:grid;grid-template-columns:repeat(auto-fill,minmax(150px,1fr));gap:12px}
.kp-grave-card{background:rgba(255,142,125,.08);border:1px solid rgba(255,142,125,.25);border-radius:14px;
  padding:14px;text-align:center}
.kp-grave-card .x{font-size:1.6rem}
.kp-grave-card .nm{font:800 1rem var(--fu);margin-top:4px}
.kp-grave-card .ch{color:#ff8e7d;font-size:.8rem;margin-top:2px}
.kp-twin-row{display:flex;align-items:center;gap:12px;padding:8px 0;border-bottom:1px solid var(--line)}
.kp-twin-row .pct{margin-left:auto;font:800 1rem var(--fu);color:var(--mint)}
.kp-twin-bar{flex:1;height:8px;border-radius:4px;background:rgba(108,232,105,.12);overflow:hidden;min-width:60px}
.kp-twin-bar span{display:block;height:100%;background:linear-gradient(90deg,var(--mint2),var(--mint))}
.kp-stake{display:flex;align-items:center;justify-content:space-between;gap:14px;flex-wrap:wrap}
.kp-stake.tm-stake{margin-top:14px;padding-top:14px;border-top:1px solid var(--line)}
.kp-stake .swing{font:800 1.8rem var(--fu);color:var(--gold)}
.kp-act{padding:46px 0;border-bottom:1px solid var(--line)}
.kp-act-kicker{font:800 .8rem var(--fu);letter-spacing:.16em;text-transform:uppercase;color:var(--mint);margin-bottom:10px}
.kp-act h2{font-size:clamp(1.8rem,4vw,3rem);margin-bottom:14px}
.kp-act-num{font:800 clamp(3rem,9vw,6rem) var(--fu);color:var(--mint);line-height:.95;margin-bottom:12px}
.kp-act-lead{font-size:1.05rem;color:var(--muted);max-width:60ch}
.kp-duo{display:grid;grid-template-columns:1fr 1fr;gap:18px;margin-top:18px}
.kp-duo .b{background:rgba(0,0,0,.2);border:1px solid var(--line);border-radius:16px;padding:18px}
.kp-duo .b .big{font:800 1.6rem var(--fu);margin:6px 0}
.kp-act-viz{margin-top:22px;background:rgba(0,0,0,.2);border:1px solid var(--line);border-radius:16px;padding:18px}
.kp-act-grid{display:grid;grid-template-columns:1fr 1fr;gap:18px;margin-top:22px}
.kp-act-grid .kp-act-viz{margin-top:0}
.kp-prog-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-top:8px}
.kp-prog-row{display:flex;justify-content:space-between;gap:10px;padding:7px 0;border-bottom:1px solid var(--line);font-size:.88rem}
.kp-prog-row .nm{font-weight:700}
.kp-tip{position:fixed;z-index:200;pointer-events:none;background:#021802;border:1px solid var(--line);
  border-radius:10px;padding:7px 11px;font:700 .82rem var(--fu);color:var(--text);
  box-shadow:0 10px 28px rgba(0,0,0,.45);opacity:0;transition:opacity .08s;max-width:240px}
.kp-tip .muted{display:block;font-weight:600;font-size:.74rem;margin-top:2px}
.kp-scatter circle{cursor:pointer;transition:stroke-width .08s}
.kp-scatter circle:hover{stroke:#fff;stroke-width:2}
.kp-fichas-wrap .search{margin-bottom:18px}
.kp-path{display:flex;flex-wrap:wrap;gap:10px;margin-top:14px;align-items:center}
.kp-path-step{background:rgba(0,0,0,.22);border:1px solid var(--line);border-radius:12px;padding:10px 14px;min-width:120px}
.kp-path-step .lab{font:800 .72rem var(--fu);color:var(--muted);text-transform:uppercase;letter-spacing:.08em}
.kp-path-step .opp{font:800 .95rem var(--fu);margin-top:4px}
.kp-path-arrow{color:var(--mint);font:800 1.2rem var(--fu)}
.kp-honors{display:grid;grid-template-columns:repeat(auto-fill,minmax(180px,1fr));gap:14px;margin-top:18px}
.kp-honor{background:rgba(0,0,0,.2);border:1px solid var(--line);border-radius:16px;padding:16px;text-align:center}
.kp-honor .em{font-size:1.8rem}
.kp-honor .ti{font:800 .82rem var(--fu);color:var(--mint);margin:8px 0 4px;text-transform:uppercase;letter-spacing:.06em}
.kp-honor .wn{font:800 1.1rem var(--fu)}
.kp-honor .dt{font-size:.82rem;color:var(--muted);margin-top:4px}
.kp-prog-timeline{display:flex;flex-direction:column;gap:10px;margin-top:16px}
.kp-prog-step{padding:10px 14px;border:1px solid var(--line);border-radius:12px;background:rgba(0,0,0,.15);font-size:.86rem}
@media(max-width:720px){.kp-act-grid,.kp-prog-grid{grid-template-columns:1fr}}
@media(max-width:560px){.kp-duo{grid-template-columns:1fr}}

@media(max-width:560px){.proto-shell-inner{padding:10px 14px}.proto-body{padding:0 14px}}
/* user identity picker + personalization */
.proto-user{display:inline-flex;align-items:center;gap:7px;border:1px solid var(--line);background:rgba(255,255,255,.04);
  color:var(--text);border-radius:999px;padding:6px 12px;font:700 .78rem var(--fu);cursor:pointer;max-width:160px}
.proto-user .pu-dot{width:9px;height:9px;border-radius:50%;flex:none}
.proto-user .pu-label{overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
@media(max-width:560px){.proto-user{max-width:44px;padding:6px 9px}.proto-user .pu-label{display:none}}
.user-overlay{position:fixed;inset:0;z-index:200;background:rgba(0,10,12,.72);backdrop-filter:blur(6px);
  display:flex;align-items:center;justify-content:center;padding:18px;opacity:0;pointer-events:none;transition:opacity .18s}
.user-overlay.open{opacity:1;pointer-events:auto}
.user-modal{background:var(--surface);border:1px solid var(--line);border-radius:20px;padding:22px;width:min(420px,100%);
  box-shadow:0 24px 80px rgba(0,0,0,.55);max-height:min(88vh,560px);display:flex;flex-direction:column}
.user-modal h3{font-size:1.15rem;margin-bottom:4px}
.user-modal .sub{color:var(--muted);font-size:.88rem;margin-bottom:14px}
.user-modal .search{max-width:none;margin-bottom:10px}
.user-list{overflow-y:auto;display:flex;flex-direction:column;gap:4px;min-height:80px;max-height:240px;margin-bottom:14px}
.user-option{border:1px solid transparent;background:rgba(255,255,255,.04);color:var(--text);border-radius:10px;
  padding:10px 12px;font:600 .92rem var(--fu);cursor:pointer;text-align:left;display:flex;align-items:center;gap:8px}
.user-option:hover,.user-option.on{border-color:rgba(108,232,105,.35);background:rgba(108,232,105,.08)}
.user-option .uo-dot{width:8px;height:8px;border-radius:50%;flex:none}
.user-actions{display:flex;flex-wrap:wrap;gap:8px}
.user-actions button{border:1px solid var(--line);background:transparent;color:var(--muted);border-radius:10px;
  padding:8px 12px;font:700 .78rem var(--fu);cursor:pointer}
.user-actions button.primary{background:var(--mint);color:#021802;border-color:transparent}
.is-me{box-shadow:0 0 0 1px rgba(108,232,105,.45)!important;background:rgba(108,232,105,.07)!important}
.bar-row.is-me,.race-row.is-me{border-color:rgba(108,232,105,.55)!important}
.race-row.is-me .bar-name{font-weight:800}
.result-person.is-me{border-color:rgba(108,232,105,.55);background:rgba(108,232,105,.12)}
.result-person.is-me b{color:var(--mint)}
.ficha.is-me{border-color:rgba(108,232,105,.55);box-shadow:0 0 0 1px rgba(108,232,105,.25)}
.mlabel.is-me{color:var(--mint);font-weight:700}
.mcell.is-me{outline:2px solid var(--mint);outline-offset:-1px}
.surv-name.is-me{color:var(--mint);font-weight:800}
.me-summary{margin-bottom:22px}
.me-summary-inner{background:linear-gradient(160deg,rgba(108,232,105,.1),rgba(0,0,0,.2));border:1px solid rgba(108,232,105,.28);
  border-left:4px solid var(--me-color,var(--mint));border-radius:16px;padding:16px 18px}
.me-summary-head{display:flex;align-items:center;gap:10px;flex-wrap:wrap;margin-bottom:12px}
.me-dot{width:11px;height:11px;border-radius:50%;background:var(--me-color,var(--mint));flex:none}
.me-summary-stats{display:flex;flex-wrap:wrap;gap:14px 22px;margin-bottom:12px}
.me-stat .lab{display:block;font-size:.68rem;color:var(--muted);text-transform:uppercase;letter-spacing:.06em}
.me-stat .val{font:800 1.1rem var(--fu);color:var(--mint)}
.me-summary-link{border:0;background:transparent;color:var(--mint);font:700 .82rem var(--fu);cursor:pointer;padding:0}
.race-me-pin{margin-bottom:8px}
.race-me-pin-row{pointer-events:none}
.me-bet-block{margin:12px 0;background:linear-gradient(160deg,rgba(108,232,105,.08),rgba(0,0,0,.15));
  border:1px solid rgba(108,232,105,.25);border-left:3px solid var(--me-color,var(--mint));border-radius:12px;padding:12px 14px}
.me-bet-head{font:800 .88rem var(--fu);color:var(--mint);margin-bottom:8px;text-transform:uppercase;letter-spacing:.04em}
.me-bet-pick{font:800 1.25rem var(--fu);margin-bottom:6px}
.me-bet-consensus{font-size:.82rem;margin-bottom:8px;line-height:1.4}
.me-bet-consensus .badge{display:inline-block;padding:2px 8px;border-radius:6px;font-size:.72rem;font-weight:700;margin-left:6px}
.me-bet-consensus .badge.agree{background:rgba(108,232,105,.15);color:var(--mint)}
.me-bet-consensus .badge.disagree{background:rgba(255,180,80,.12);color:#ffb450}
.me-bet-stake{font-size:.82rem;line-height:1.45}
.me-bet-stake .swing{font:800 1.1rem var(--fu);color:var(--mint)}
.me-bet-void{margin-top:6px;padding:11px 13px;border-radius:10px;background:rgba(150,120,190,.16);border:1px solid rgba(170,140,210,.4);
  font:800 1.08rem var(--fu);color:#d3b8ff;display:flex;flex-direction:column;gap:3px}
.me-bet-void span{font:500 .8rem/1.4 var(--fu);color:var(--muted)}
.race-me-step{margin-top:10px;padding:10px 12px;background:rgba(108,232,105,.08);border:1px solid rgba(108,232,105,.22);
  border-radius:10px;font-size:.84rem;line-height:1.45}
.race-me-step b{color:var(--mint)}
.me-result-strip{margin:10px 0 4px;padding:8px 12px;background:rgba(108,232,105,.1);border:1px solid rgba(108,232,105,.28);
  border-radius:8px;font:700 .84rem var(--fu);color:var(--mint)}
.me-result-strip.miss{color:#ff8a8a;background:rgba(255,100,100,.08);border-color:rgba(255,100,100,.25)}
.me-result-strip.sign{color:#c8d4ff;background:rgba(150,170,255,.08);border-color:rgba(150,170,255,.22)}
.me-result-strip.voided{color:#d3b8ff;background:rgba(150,120,190,.1);border-color:rgba(160,130,200,.28)}
.kp-scatter circle.is-me{stroke:#fff;stroke-width:2.5}
"""

JS = r"""
const D = window.__PORRA__;
const N = D.hero.participants;
const ES2EN = D.es2en || {};
const logo = `__LOGO__`;
let LANG = 'es';
let wrap = null;
let HOY_DATE = null; // null = jornada real de hoy; si no, fecha YYYY-MM-DD elegida con las flechas

/* ---- USER IDENTITY (localStorage + ?user=) ---- */
const USER_KEY = 'porra2026.currentUser';
const PICKER_DISMISS_KEY = 'porra2026.pickerDismissed';
let ME = null;
let _pickerEl = null;

function participantNames(){
  return ((D.matrix && D.matrix.names) || []).slice();
}
function isKnownParticipant(name){
  return participantNames().includes(name);
}
function loadMe(){
  const params = new URLSearchParams(location.search);
  const fromUrl = params.get('user');
  if(fromUrl){
    const decoded = decodeURIComponent(fromUrl);
    if(isKnownParticipant(decoded)){
      ME = decoded;
      try { localStorage.setItem(USER_KEY, ME); } catch(e){}
      params.delete('user');
      const u = new URL(location.href);
      u.search = params.toString();
      history.replaceState(null, '', u.pathname + (u.search ? '?' + u.search : '') + u.hash);
      return;
    }
  }
  try {
    const stored = localStorage.getItem(USER_KEY);
    if(stored && isKnownParticipant(stored)) ME = stored;
    else {
      if(stored){
        localStorage.removeItem(USER_KEY);
        localStorage.removeItem(PICKER_DISMISS_KEY);
      }
      ME = null;
    }
  } catch(e){ ME = null; }
}
function saveMe(name){
  ME = name || null;
  try {
    if(ME) localStorage.setItem(USER_KEY, ME);
    else localStorage.removeItem(USER_KEY);
    localStorage.setItem(PICKER_DISMISS_KEY, '1');
  } catch(e){}
}
function dismissPicker(){
  try { localStorage.setItem(PICKER_DISMISS_KEY, '1'); } catch(e){}
}
function pickerDismissed(){
  try { return localStorage.getItem(PICKER_DISMISS_KEY) === '1'; } catch(e){ return false; }
}
function isMe(name){ return !!(ME && name === ME); }
function meClass(name){ return isMe(name) ? ' is-me' : ''; }
function meCard(){
  if(!ME) return null;
  return (D.cards || []).find(c => c.name === ME) || null;
}
function meLiveRow(){
  if(!ME || !D.live || !D.live.table) return null;
  return D.live.table.find(r => r.name === ME) || null;
}
function meStakePerson(stake){
  if(!ME || !stake || !stake.people) return null;
  return stake.people.find(p => p.name === ME) || null;
}
function mePickInMatch(m){
  if(!ME || !m || !m.picks) return null;
  return m.picks.find(p => p.name === ME) || null;
}
function meProgressionAt(idx){
  const hist = liveHistory();
  const snap = hist[idx];
  if(!snap || !ME) return null;
  return snap.table.find(r => r.name === ME) || null;
}
function meRecentOutcome(m){
  if(!ME || !m) return null;
  const lists = [
    {key:'exact', kind:'exact'},
    {key:'sign', kind:'sign'},
    {key:'voided', kind:'voided'},
    {key:'miss', kind:'miss'},
    {key:'advance', kind:'advance'},
  ];
  const advanced = (m.advance || []).some(p => p.name === ME);
  const advancePts = m.advance_points || 0;
  for(const {key, kind} of lists){
    const hit = (m[key] || []).find(p => p.name === ME);
    if(hit) return {kind, pick: hit.pick, advanced, advancePts};
  }
  return null;
}
function sortPeople(arr){
  return [...arr].sort((a,b) => {
    if(isMe(a.name)) return -1;
    if(isMe(b.name)) return 1;
    return a.name.localeCompare(b.name, LANG === 'es' ? 'es' : 'en');
  });
}
function userBarHtml(){
  const dot = ME ? `<span class="pu-dot" style="background:${personColor(ME)}"></span>` : '';
  const short = ME && ME.length > 14 ? ME.slice(0, 13) + '…' : ME;
  const label = ME ? esc(short) : L('¿Quién eres?', 'Who are you?');
  return `<button type="button" class="proto-user" data-action="open-user-picker">${dot}<span class="pu-label">${label}</span></button>`;
}
function ensureUserPicker(){
  if(_pickerEl) return _pickerEl;
  _pickerEl = el('div', 'user-overlay');
  _pickerEl.id = 'user-picker';
  _pickerEl.innerHTML = `<div class="user-modal" role="dialog" aria-modal="true">
    <h3>${L('¿Quién eres en la porra?', 'Who are you in the pool?')}</h3>
    <p class="sub">${L('Elige tu nombre para resaltar tu ranking, picks y ficha.', 'Pick your name to highlight your ranking, picks and card.')}</p>
    <input type="search" class="search user-search" autocomplete="off" placeholder="${L('🔎 Busca tu nombre…', '🔎 Search your name…')}">
    <div class="user-list"></div>
    <div class="user-actions">
      <button type="button" class="primary" data-pick-confirm disabled>${L('Confirmar', 'Confirm')}</button>
      <button type="button" data-pick-spectator>${L('Solo miro', 'Just browsing')}</button>
      <button type="button" data-pick-later>${L('Ahora no', 'Not now')}</button>
    </div>
  </div>`;
  document.body.appendChild(_pickerEl);
  const modal = _pickerEl.querySelector('.user-modal');
  const inp = _pickerEl.querySelector('.user-search');
  const list = _pickerEl.querySelector('.user-list');
  const confirmBtn = _pickerEl.querySelector('[data-pick-confirm]');
  let selected = ME;
  function renderList(){
    const q = inp.value.toLowerCase().trim();
    const names = participantNames().filter(n => !q || n.toLowerCase().includes(q));
    list.innerHTML = names.length
      ? names.map(n => `<button type="button" class="user-option${n === selected ? ' on' : ''}" data-name="${esc(n)}">
          <span class="uo-dot" style="background:${personColor(n)}"></span>${esc(n)}</button>`).join('')
      : `<span class="muted" style="padding:8px">${L('Sin coincidencias', 'No matches')}</span>`;
    confirmBtn.disabled = !selected || !names.includes(selected);
  }
  function closePicker(){
    _pickerEl.classList.remove('open');
    selected = ME;
  }
  function openPicker(){
    selected = ME;
    inp.value = '';
    renderList();
    _pickerEl.classList.add('open');
    inp.focus();
  }
  inp.addEventListener('input', () => {
    const q = inp.value.toLowerCase().trim();
    const first = participantNames().find(n => !q || n.toLowerCase().includes(q));
    selected = first || null;
    renderList();
  });
  inp.addEventListener('keydown', e => {
    if(e.key === 'Enter'){
      const first = list.querySelector('.user-option');
      if(first && first.dataset.name){ selected = first.dataset.name; confirmBtn.click(); }
    }
    if(e.key === 'Escape'){ dismissPicker(); closePicker(); }
  });
  list.addEventListener('click', e => {
    const btn = e.target.closest('.user-option');
    if(!btn) return;
    selected = btn.dataset.name;
    renderList();
  });
  confirmBtn.addEventListener('click', () => {
    if(!selected) return;
    saveMe(selected);
    closePicker();
    rebuild();
  });
  _pickerEl.querySelector('[data-pick-spectator]').addEventListener('click', () => {
    saveMe(null);
    closePicker();
    rebuild();
  });
  _pickerEl.querySelector('[data-pick-later]').addEventListener('click', () => {
    dismissPicker();
    closePicker();
  });
  _pickerEl.addEventListener('click', e => {
    if(e.target === _pickerEl){ dismissPicker(); closePicker(); }
  });
  modal.addEventListener('click', e => e.stopPropagation());
  _pickerEl.openPicker = openPicker;
  return _pickerEl;
}
function showUserPicker(){ ensureUserPicker().openPicker(); }
function maybeShowUserPicker(){
  if(!ME && !pickerDismissed()) showUserPicker();
}
function buildMeSummary(parent){
  if(!ME || !parent) return;
  const row = meLiveRow();
  const card = meCard();
  const stats = [];
  if(row){
    stats.push(`<div class="me-stat"><span class="lab">${L('Puesto', 'Rank')}</span><span class="val">#${row.rank}</span></div>`);
    stats.push(`<div class="me-stat"><span class="lab">${L('Puntos', 'Points')}</span><span class="val">${row.pts}</span></div>`);
    if(row.delta) stats.push(`<div class="me-stat"><span class="lab">${L('Último mov.', 'Last move')}</span><span class="val">${row.delta > 0 ? '+' : ''}${row.delta}</span></div>`);
  }
  if(card){
    stats.push(`<div class="me-stat"><span class="lab">${L('Gemelo', 'Twin')}</span><span class="val">${esc(card.twin)}</span></div>`);
    stats.push(`<div class="me-stat"><span class="lab">${L('Rebeldía', 'Maverick')}</span><span class="val">#${card.rebel_rank}</span></div>`);
  }
  const statsHtml = stats.length
    ? stats.join('')
    : `<span class="muted">${L('Sin datos en directo todavía', 'No live data yet')}</span>`;
  const banner = el('div', 'me-summary reveal',
    `<div class="me-summary-inner" style="--me-color:${personColor(ME)}">
      <div class="me-summary-head"><span class="me-dot"></span><b>${esc(ME)}</b><span class="muted">${L('Tu resumen', 'Your summary')}</span></div>
      <div class="me-summary-stats">${statsHtml}</div>
      <button type="button" class="me-summary-link" data-goto-ficha>${L('Ver tu ficha →', 'See your card →')}</button>
    </div>`);
  banner.querySelector('[data-goto-ficha]').addEventListener('click', () => {
    const u = new URL(location.href);
    u.searchParams.set('view', 'groups');
    u.hash = 'fichas';
    history.replaceState(null, '', u);
    rebuild();
  });
  parent.appendChild(banner);
}
function afterRenderMeHooks(view){
  if(view !== 'groups' || !ME) return;
  if(location.hash !== '#fichas') return;
  requestAnimationFrame(() => {
    const ficha = document.querySelector('#fichas .ficha.is-me');
    const inp = document.querySelector('#fichas .search');
    if(inp && !inp.value) inp.value = ME;
    if(ficha) setTimeout(() => ficha.scrollIntoView({behavior:'smooth', block:'center'}), 400);
  });
}

function L(es, en){ return LANG === 'es' ? es : en; }
function team(es){ return LANG === 'es' ? es : (ES2EN[es] || es); }
function esc(s){ return String(s).replace(/[&<>]/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;'}[c])); }
// Hora de saque según idioma: ES -> peninsular, EN -> Reino Unido (BST).
function koTime(m){ return (LANG==='es' ? m.time_es : m.time_uk) || ''; }
function koNext(m){ return LANG==='es' ? !!m.next_es : !!m.next_uk; }
function koTz(){ return L('hora peninsular','UK time'); }
function fmt(v){
  let [a,b] = String(v).split('.');
  const th = LANG==='es' ? '.' : ',', dp = LANG==='es' ? ',' : '.';
  a = a.replace(/\B(?=(\d{3})+(?!\d))/g, th);
  return b !== undefined ? a + dp + b : a;
}
function matchdayDateStr(now){
  const d = new Date(now.getTime());
  if (d.getHours() < 6) d.setDate(d.getDate() - 1);
  return d.getFullYear() + '-' + String(d.getMonth()+1).padStart(2,'0') + '-' + String(d.getDate()).padStart(2,'0');
}
function pf(x){ let s = String(x); if (LANG==='es') s = s.replace('.', ','); return s + '%'; }
function animateCount(node){
  const to = parseFloat(node.dataset.count), dec = parseInt(node.dataset.decimals||'0'), suf = node.dataset.suffix||'';
  const dur = 900, t0 = performance.now();
  (function step(t){ let p = Math.min(1,(t-t0)/dur); p = 1-Math.pow(1-p,3);
    node.textContent = fmt((to*p).toFixed(dec)) + suf; if (p<1) requestAnimationFrame(step); })(t0);
}
function el(t,c,h){ const e = document.createElement(t); if(c) e.className=c; if(h!=null) e.innerHTML=h; return e; }

const LABELS = {
  goleador:['🥅 Goleador','🥅 Goal machine'], cerrojo:['🔒 Cerrojo','🔒 Parked bus'],
  empate:['🤝 Amigo del empate','🤝 Draw lover'], rebelde:['🐺 Rebelde','🐺 Maverick'],
  borrego:['🐑 Borrego','🐑 Sheep'], equilibrado:['⚖️ Equilibrado','⚖️ Balanced'],
};
function labelOf(k){ const x = LABELS[k] || LABELS.equilibrado; return L(x[0], x[1]); }

/* ---- LIVE RANKING (?view=live → #aciertos) ----
   Single UI: match-by-match race (slider + bar board). No ?variant= switcher. */
/* Colour por persona, determinista: cada participante conserva SIEMPRE el mismo
   color en todas las secciones, independientemente de su puesto actual en el
   ranking. El tono sale de un índice alfabético estable repartido por el ángulo
   áureo; saturación y luminosidad son fijas para que todos los colores destaquen
   sobre el fondo verde oscuro (sin marrones ni mostazas apagados). */
let _personColorMap = null;
function personColorMap(){
  if(_personColorMap) return _personColorMap;
  const names = ((D.live && D.live.table) || []).map(r => r.name)
    .slice().sort((a,b) => a.localeCompare(b, 'es'));
  const map = {};
  names.forEach((name,i) => { map[name] = `hsl(${(i * 137.508 % 360).toFixed(1)} 72% 64%)`; });
  _personColorMap = map;
  return map;
}
function personColor(name){
  const m = personColorMap();
  if(m[name]) return m[name];
  let h = 0;
  for(let i=0;i<name.length;i++) h = (h * 31 + name.charCodeAt(i)) >>> 0;
  return `hsl(${h % 360} 72% 64%)`;
}
let racePlayTimer = null;
function liveHistory(){ return (D.live && D.live.progression) || []; }
function latestSnapshot(){ const h = liveHistory(); return h[h.length - 1] || null; }
function matchTitle(m){
  if(!m) return '–';
  if(m.virtual && m.kind === 'ko' && m.result && m.home){
    return `${m.home_flag} ${esc(team(m.home))} ${m.result.home}-${m.result.away} ${esc(team(m.away))} ${m.away_flag}`;
  }
  if(m.virtual) return L(m.label_es || 'Clasificados de grupo', m.label_en || 'Group standings');
  return `${m.home_flag} ${esc(team(m.home))} ${m.result.home}-${m.result.away} ${esc(team(m.away))} ${m.away_flag}`;
}
function snapshotSubtitle(snap){
  if(snap.virtual){
    const label = snap.kind === 'ko'
      ? L(snap.phase_es || 'Eliminatorias', snap.phase_en || 'Knockouts')
      : L('Cierre fase de grupos','End of group stage');
    return snap.date ? `${snap.date} · ${label}` : label;
  }
  return `${snap.date} · ${L('Grupo','Group')} ${snap.group}`;
}
function rankDelta(row){
  if(!row.delta) return `<span class="rank-delta">=</span>`;
  const cls = row.delta > 0 ? 'up' : 'down';
  return `<span class="rank-delta ${cls}">${row.delta > 0 ? '+' : ''}${row.delta}</span>`;
}
function rankDeltaLong(row){
  if(!row.delta) return `<span class="rank-delta">=</span>`;
  const cls = row.delta > 0 ? 'up' : 'down';
  const word = row.delta > 0 ? L('sube','up') : L('baja','down');
  return `<span class="rank-delta ${cls}">${row.delta > 0 ? '+' : ''}${row.delta} ${word}</span>`;
}
function rankingColorMap(){
  const map = {};
  ((D.live && D.live.table) || []).forEach(r => { map[r.name] = personColor(r.name); });
  return map;
}
function clearRaceTimer(){
  if(racePlayTimer){
    clearInterval(racePlayTimer);
    racePlayTimer = null;
  }
}
function tipMetric(count, pts, label, cls){
  if(!pts && !count) return '';
  const countLine = count != null ? `<b>${count}</b>${label}` : `<span>${label}</span>`;
  return `<span class="rt-item ${cls || ''}">${countLine}<b class="rt-pts">+${pts}</b></span>`;
}
function rankTip(r){
  const gExact = r.exact || 0;
  const gSign = r.sign || 0;
  const gPlenoPts = gExact * 4;
  const gSignPts = gSign * 2;
  const gStand = r.standings_pts || 0;
  const gThird = r.thirds_pts || 0;
  const gTotal = gPlenoPts + gSignPts + gStand + gThird;
  const groupItems = [
    tipMetric(gExact, gPlenoPts, L('plenos','exact'), 'rt-pleno'),
    tipMetric(gSign, gSignPts, L('signos','outcomes'), 'rt-sign'),
    gStand ? tipMetric(null, gStand, L('clasificados','standings'), 'rt-bonus') : '',
    gThird ? tipMetric(null, gThird, L('terceros','thirds'), 'rt-bonus') : '',
  ].filter(Boolean).join('');
  if(!groupItems && !gTotal) return '';
  const groupBlock = `<div class="rt-block">
    <div class="rt-head">${L('Fase de grupos','Group stage')}</div>
    <div class="rt-items">${groupItems}</div>
    <div class="rt-sum">= ${gTotal} pts</div>
  </div>`;

  const koPts = r.ko_pts || 0;
  if(!koPts) return `<div class="race-tip">${groupBlock}</div>`;

  const koExact = r.ko_exact || 0;
  const koOut = r.ko_outcomes || 0;
  const koSignOnly = Math.max(0, koOut - koExact);
  const koExactPts = koExact * 5;
  const koSignPts = koSignOnly * 3;
  const koAdv = r.ko_advance || 0;
  const koAdvPts = Math.max(0, koPts - koExactPts - koSignPts);
  const koItems = [
    koExact ? tipMetric(koExact, koExactPts, L('plenos','exact'), 'rt-ko') : '',
    koSignOnly ? tipMetric(koSignOnly, koSignPts, L('signos','outcomes'), 'rt-ko') : '',
    koAdv ? tipMetric(koAdv, koAdvPts, L('pases','advances'), 'rt-ko') : '',
    !koExact && !koSignOnly && !koAdv ? tipMetric(null, koPts, L('puntos','points'), 'rt-ko') : '',
  ].filter(Boolean).join('');
  const koBlock = `<div class="rt-block">
    <div class="rt-head">${L('Eliminatorias','Knockouts')}</div>
    <div class="rt-items">${koItems}</div>
    <div class="rt-sum">= ${koPts} pts</div>
  </div>`;
  return `<div class="race-tip">${groupBlock}${koBlock}</div>`;
}
function buildLiveRanking(s){
  const snap = latestSnapshot();
  if(!snap){
    s.appendChild(el('div','card teaser reveal',
      `<div class="em">📈</div><h3 style="margin:10px 0 6px">${L('Sin histórico todavía','No history yet')}</h3>
       <p class="muted">${L('El ranking necesita al menos un resultado real cargado.','The ranking needs at least one real score loaded.')}</p>`));
    return;
  }
  clearRaceTimer();
  const hist = liveHistory(), finalRows = D.live.table, mx = finalRows[0].pts || 1, colors = rankingColorMap();
  const card = el('div','rank-proto-card reveal',
    `<div class="rank-proto-top">
      <div><h3>${L('Carrera partido a partido','Match-by-match race')}</h3>
      <p class="muted">${L('Barras acumuladas partido a partido; los pasos ★ reparten puntos de clasificados o eliminatorias.','Accumulated bars match by match; ★ steps award standings or knockout points.')}</p></div>
    </div>
    <div class="race-layout">
      <div class="race-control">
        <div class="k">${L('Partido','Match')}</div>
        <input type="range" min="0" max="${hist.length-1}" value="${hist.length-1}">
        <div class="race-actions">
          <button type="button" class="race-prev" aria-label="${L('Partido anterior','Previous match')}">‹</button>
          <button type="button" class="race-play" aria-label="${L('Reproducir','Play')}">▶</button>
          <div class="race-step"></div>
          <button type="button" class="race-next" aria-label="${L('Partido siguiente','Next match')}">›</button>
        </div>
        <div class="race-match"></div>
        <div class="muted race-date"></div>
        <div class="race-now"></div>
        <div class="race-me-step" hidden></div>
        <div class="race-legend">
          <span><i class="race-dot" style="background:var(--mint)"></i>${L('sube en ranking','rank up')}</span>
          <span><i class="race-dot" style="background:var(--red)"></i>${L('baja en ranking','rank down')}</span>
          <span><i class="race-dot" style="background:var(--gold)"></i>${L('pts del partido','match pts')}</span>
        </div>
      </div>
      <div class="race-board-wrap">
        <div class="race-me-pin" hidden></div>
        <div class="race-board"></div>
      </div>
    </div>`);
  const input = card.querySelector('input');
  const board = card.querySelector('.race-board');
  const mePin = card.querySelector('.race-me-pin');
  const match = card.querySelector('.race-match');
  const date = card.querySelector('.race-date');
  const step = card.querySelector('.race-step');
  const now = card.querySelector('.race-now');
  const meStep = card.querySelector('.race-me-step');
  const play = card.querySelector('.race-play');
  const prev = card.querySelector('.race-prev');
  const next = card.querySelector('.race-next');
  function setPlaying(isPlaying){
    play.textContent = isPlaying ? 'Ⅱ' : '▶';
    play.setAttribute('aria-label', isPlaying ? L('Pausar','Pause') : L('Reproducir','Play'));
  }
  function stopPlaying(){
    clearRaceTimer();
    setPlaying(false);
  }
  function goTo(idx){
    input.value = String(Math.max(0, Math.min(hist.length - 1, idx)));
    paint();
  }
  function collectRowRects(){
    const rects = new Map();
    board.querySelectorAll('.race-row').forEach(row => {
      rects.set(row.dataset.name, row.getBoundingClientRect());
    });
    return rects;
  }
  function rowHtml(r,i){
    const me = isMe(r.name) ? ' is-me' : '';
    return `<div class="race-row${me} ${i===0?'leader':''} ${r.delta>0?'moved-up':(r.delta<0?'moved-down':'')}" data-name="${esc(r.name)}" style="--runner:${colors[r.name] || personColor(r.name)}">
      <div class="bar-rank">${r.rank}</div>
      <div class="bar-name">${esc(r.name)}</div>
      ${rankDeltaLong(r)}
      <div class="bar-track"><div class="race-fill" style="width:${(r.pts/mx*100).toFixed(1)}%"></div></div>
      <div class="bar-val">${r.pts}</div>
      <div class="race-round">+${r.round_pts || 0}</div>
      ${rankTip(r)}
    </div>`;
  }
  function racePinHtml(r){
    return `<div class="race-row is-me race-me-pin-row" data-name="${esc(r.name)}" style="--runner:${colors[r.name] || personColor(r.name)}">
      <div class="bar-rank">${r.rank}</div>
      <div class="bar-name">${L('Tú', 'You')}: ${esc(r.name)}</div>
      ${rankDeltaLong(r)}
      <div class="bar-track"><div class="race-fill" style="width:${(r.pts/mx*100).toFixed(1)}%"></div></div>
      <div class="bar-val">${r.pts}</div>
      <div class="race-round">+${r.round_pts || 0}</div>
    </div>`;
  }
  function updateMePin(snap){
    if(!mePin || !ME){ if(mePin) mePin.hidden = true; return; }
    const r = snap.table.find(x => x.name === ME);
    if(!r){ mePin.hidden = true; return; }
    mePin.hidden = false;
    mePin.innerHTML = racePinHtml(r);
  }
  function updateMeStep(snap){
    if(!meStep || !ME){ if(meStep) meStep.hidden = true; return; }
    const r = snap.table.find(x => x.name === ME);
    if(!r){ meStep.hidden = true; return; }
    meStep.hidden = false;
    const pts = r.round_pts || 0;
    const delta = r.delta || 0;
    let moveTxt = '';
    if(delta > 0) moveTxt = L(', subiste ', ', you moved up ') + delta;
    else if(delta < 0) moveTxt = L(', bajaste ', ', you moved down ') + Math.abs(delta);
    else moveTxt = L(', sin cambio de puesto', ', no rank change');
    if(pts > 0){
      meStep.innerHTML = `${L('En este partido','This match')}: <b>+${pts} pts</b>${moveTxt}`;
    } else if(delta !== 0){
      meStep.innerHTML = `${L('En este partido','This match')}: <b>${L('sin puntos','no points')}</b>${moveTxt}`;
    } else {
      meStep.innerHTML = `${L('En este partido','This match')}: ${L('sin puntos este paso','no points this step')}`;
    }
  }
  function animateRowsFrom(previousRects){
    if(!previousRects || !previousRects.size || !board.isConnected || typeof Element === 'undefined') return;
    const reduceMotion = window.matchMedia && window.matchMedia('(prefers-reduced-motion: reduce)').matches;
    if(reduceMotion) return;
    board.querySelectorAll('.race-row').forEach(row => {
      const before = previousRects.get(row.dataset.name);
      if(!before || typeof row.animate !== 'function') return;
      const after = row.getBoundingClientRect();
      const dx = before.left - after.left;
      const dy = before.top - after.top;
      if(Math.abs(dx) < .5 && Math.abs(dy) < .5) return;
      const duration = Math.min(880, Math.max(540, Math.round(430 + Math.abs(dy) * 1.05)));
      row.classList.add('is-moving');
      const animation = row.animate([
        {transform:`translate3d(${dx}px,${dy}px,0)`},
        {transform:'translate3d(0,0,0)'}
      ], {duration, easing:'cubic-bezier(.16,.9,.2,1)', fill:'both'});
      const cleanup = () => {
        row.classList.remove('is-moving');
        animation.cancel();
      };
      animation.finished.then(cleanup).catch(() => row.classList.remove('is-moving'));
    });
  }
  function paint(options){
    const animate = !options || options.animate !== false;
    const previousRects = animate ? collectRowRects() : null;
    const snap = hist[Number(input.value)];
    const leader = snap.table[0];
    const movers = snap.table.filter(r => r.delta !== 0).length;
    const pointsNow = snap.table.reduce((sum,r) => sum + (r.round_pts || 0), 0);
    step.textContent = `${snap.idx}/${hist.length}`;
    match.innerHTML = `${snap.idx}. ${matchTitle(snap)}`;
    date.textContent = snapshotSubtitle(snap);
    now.innerHTML = `
      <span><b>${esc(leader.name)}</b>${L('líder','leader')} · ${leader.pts} pts</span>
      <span><b>${movers}</b>${L('movimientos','moves')}</span>
      <span><b>+${pointsNow}</b>${L('pts repartidos','pts awarded')}</span>`;
    board.innerHTML = snap.table.map(rowHtml).join('');
    updateMePin(snap);
    updateMeStep(snap);
    requestAnimationFrame(() => animateRowsFrom(previousRects));
  }
  input.addEventListener('input', () => { stopPlaying(); paint(); });
  prev.addEventListener('click', () => { stopPlaying(); goTo(Number(input.value) - 1); });
  next.addEventListener('click', () => { stopPlaying(); goTo(Number(input.value) + 1); });
  play.addEventListener('click', () => {
    if(racePlayTimer){ stopPlaying(); return; }
    if(Number(input.value) >= hist.length - 1) goTo(0);
    setPlaying(true);
    racePlayTimer = setInterval(() => {
      const idx = Number(input.value);
      if(idx >= hist.length - 1){ stopPlaying(); return; }
      goTo(idx + 1);
    }, 900);
  });
  paint({animate:false});
  s.appendChild(card);
}

/* shared tooltip for affinity matrix */
const tip = el('div'); tip.id = 'mtip'; document.body.appendChild(tip);

function section(id, kicker, title, sub){
  const s = el('section','sec'); s.id = id;
  s.appendChild(el('div','sec-head reveal',
    `<div class="kicker">${kicker}</div><h2>${title}</h2>${sub?`<p class="sub">${sub}</p>`:''}`));
  wrap.appendChild(s); return s;
}

/* ---- HERO ---- */
function chip(val, lab, dec){
  return `<div class="chip reveal"><div class="big" data-count="${val}"${dec?` data-decimals="${dec}"`:''}>0</div><div class="lab">${lab}</div></div>`;
}
function chipText(val, lab){
  return `<div class="chip reveal"><div class="big">${val}</div><div class="lab">${lab}</div></div>`;
}
function buildHero(){
  const h = el('header','hero'); h.id = 'inicio';
  const ts = D.hero.top_scoreline, bp = D.best_pair;
  h.innerHTML = `
    <div class="eyebrow reveal">${L('Reveni · Porra Mundial 2026','Reveni · World Cup 2026 Pool')}</div>
    <h1 class="reveal">${L('La quiniela,<br>bajo el microscopio.','The pool,<br>under the microscope.')}</h1>
    <p class="lead reveal">${L(
      'Hemos analizado las '+N+' quinielas de la porra del Mundial: quién se la juega, quién va a lo seguro, quién piensa igual que quién y qué partidos nos parten en dos.',
      'We crunched all '+N+' predictions in the office World Cup pool: who gambles, who plays it safe, who thinks like who, and which matches split us in two.')}</p>
    <div class="chips">
      ${chip(N, L('Participantes','Players'))}
      ${chip(D.hero.matches, L('Partidos de grupos','Group matches'))}
      ${chip(D.hero.groups, L('Grupos','Groups'))}
      ${chip(N*D.hero.matches, L('Marcadores analizados','Predictions analysed'))}
      ${chip(D.hero.total_goals, L('Goles apostados','Goals predicted'))}
      ${chip(D.hero.avg_goals_match, L('Goles/partido de media','Avg goals per match'), 1)}
      ${chipText(ts?ts.score:'–', L('Marcador estrella','Star scoreline')+' ('+pf(ts?ts.pct:0)+')')}
      ${chipText(pf(bp.pct), L('Máx. afinidad','Top affinity')+' · '+esc(bp.a)+' & '+esc(bp.b))}
    </div>
    <div class="scrollcue">${L('desliza para empezar','scroll to start')} ↓</div>`;
  wrap.appendChild(h);
}

/* ---- HOY ---- */
function todayScheduleMatches(){
  const groupMatches = (D.today && D.today.matches) || [];
  const GATED_KO = {r16:1, qf:1, sf:1};
  const koRounds = ((D.knockout && D.knockout.rounds) || []).flatMap(r =>
    (r.matches || []).map(m => ({
      ...m,
      phase_es:r.label_es,
      phase_en:r.label_en,
      is_knockout:true,
      advance_points:r.advance_points,
      scoreline_gated: !!GATED_KO[r.key],
    })));
  const koFinals = ((D.knockout && D.knockout.final_matches) || [])
    .map(m => ({...m, phase_es:m.label_es, phase_en:m.label_en, is_knockout:true, advance_points:0, scoreline_gated:false}));
  const knockoutMatches = koRounds.concat(koFinals).map(m => ({
    ...m,
    home:m.resolved_home || m.fixture_home,
    away:m.resolved_away || m.fixture_away,
    home_flag:m.resolved_home_flag || m.fixture_home_flag,
    away_flag:m.resolved_away_flag || m.fixture_away_flag,
  }));
  return groupMatches.concat(knockoutMatches);
}

function stakeResultLbl(r){
  if(!r) return '';
  let s = r.score;
  if(r.winner){
    const w = LANG === 'es' ? r.winner : (r.winner_en || r.winner);
    s += ` · ${r.winner_flag || ''} ${esc(team(w))}`;
  }
  return s;
}

function stakeSwingHtml(st){
  const meSt = meStakePerson(st);
  const up = st.max_swing || 0, down = st.min_swing || 0;
  const meUp = meSt ? (meSt.swing_up || 0) : 0;
  const meDown = meSt ? (meSt.swing_down || 0) : 0;
  if(ME && meSt && (meUp || meDown)){
    const parts = [];
    if(meUp) parts.push(`+${meUp}`);
    if(meDown) parts.push(`−${meDown}`);
    let hints = '';
    if(meUp && meSt.best_result){
      hints += `<div class="muted" style="font-size:.75rem;margin-top:4px;line-height:1.35">+${meUp} ${L('si sale','if')} <b>${stakeResultLbl(meSt.best_result)}</b></div>`;
    }
    if(meDown && meSt.worst_result){
      hints += `<div class="muted" style="font-size:.75rem;margin-top:2px;line-height:1.35">−${meDown} ${L('si sale','if')} <b>${stakeResultLbl(meSt.worst_result)}</b></div>`;
    }
    let globalHint = '';
    if(up && st.max_swing_who && st.max_swing_who !== ME){
      globalHint += `<div class="muted" style="font-size:.72rem;margin-top:6px;line-height:1.35">${L('En la porra, quien más se juega es','In the pool, most at stake is')} <b>${esc(st.max_swing_who)}</b> (+${up})</div>`;
    }
    if(down && st.min_swing_who && st.min_swing_who !== ME && st.min_swing_who !== st.max_swing_who){
      globalHint += `<div class="muted" style="font-size:.72rem;margin-top:2px;line-height:1.35">${L('Quien más puede bajar es','Most downside for')} <b>${esc(st.min_swing_who)}</b> (−${down})</div>`;
    }
    return `<div class="swing">${parts.join(' / ')}</div>
      <div class="muted" style="font-size:.82rem">${L('tus puestos arriba / abajo','your places up / down')}</div>${hints}${globalHint}`;
  }
  if(!up && !down) return `<div class="swing">0</div>
      <div class="muted" style="font-size:.82rem">${L('sin movimiento','no movement')}</div>`;
  const parts = [];
  if(up) parts.push(`+${up}`);
  if(down) parts.push(`−${down}`);
  let hints = '';
  if(up && st.max_swing_result){
    const who = st.max_swing_who ? ` <span class="muted">(${esc(st.max_swing_who)})</span>` : '';
    hints += `<div class="muted" style="font-size:.75rem;margin-top:4px;line-height:1.35">+${up} ${L('si sale','if')} <b>${stakeResultLbl(st.max_swing_result)}</b>${who}</div>`;
  }
  if(down && st.min_swing_result){
    const who = st.min_swing_who ? ` <span class="muted">(${esc(st.min_swing_who)})</span>` : '';
    hints += `<div class="muted" style="font-size:.75rem;margin-top:2px;line-height:1.35">−${down} ${L('si sale','if')} <b>${stakeResultLbl(st.min_swing_result)}</b>${who}</div>`;
  }
  return `<div class="swing">${parts.join(' / ')}</div>
      <div class="muted" style="font-size:.82rem">${L('puestos arriba / abajo','places up / down')}</div>${hints}`;
}

function stakeDeferredHtml(st){
  const wait = (st.pending_after || [])[0];
  if(!wait) return '';
  const lbl = `${wait.home_flag} ${esc(team(wait.home))} – ${esc(team(wait.away))} ${wait.away_flag}`;
  const when = wait.time_es ? ` · ${esc(wait.time_es)}` : '';
  return `<div class="muted" style="font-size:.82rem;margin-top:6px;line-height:1.45;max-width:280px;text-align:right;margin-left:auto">${L(
    'El ranking en juego se calcula cuando entren los resultados anteriores de hoy',
    'Stakes are calculated once earlier today\'s results are in'
  )}<br><span style="color:var(--mint)">${lbl}</span>${when}</div>`;
}

function pickScoreLbl(p, knockout){
  if(p.home == null || p.away == null) return '–';
  let s = `${p.home}-${p.away}`;
  if(knockout && p.home === p.away && p.winner){
    s += ` · ${p.winner_flag || ''} ${esc(team(p.winner))}`;
  }
  return s;
}

function meTodayOutcome(m){
  if(!ME) return null;
  const pick = mePickInMatch(m);
  if(!pick || pick.home == null || !m.result) return null;
  const resultScore = m.result.score || m.result;
  const rh = resultScore.home, ra = resultScore.away;
  if(rh == null || ra == null) return null;
  const ph = pick.home, pa = pick.away;
  const pickStr = `${ph}-${pa}`;
  // En cruces con hueco, el marcador solo cuenta si acertaste el cruce entero
  // (los dos equipos). Si el rival era otro, tu signo/pleno se anula → cementerio.
  const branchFell = pick.matchup_ok === false;
  const ps = ph > pa ? '1' : ph === pa ? 'X' : '2';
  const rs = rh > ra ? '1' : rh === ra ? 'X' : '2';
  if(ph === rh && pa === ra) return {kind: branchFell ? 'voided' : 'exact', pick: pickStr};
  if(ps === rs) return {kind: branchFell ? 'voided' : 'sign', pick: pickStr};
  return {kind:'miss', pick: pickStr};
}

function meResultStripText(outcome){
  const pick = esc(outcome.pick || '');
  const pase = outcome.advanced && outcome.advancePts
    ? ` · ${L('acertaste el pase','right advance')} +${outcome.advancePts}`
    : '';
  switch(outcome.kind){
    case 'exact': return `✓ ${L('Pleno','Exact score')} — ${L('apostaste','you picked')} ${pick}${pase}`;
    case 'sign': return `~ ${L('Signo','Outcome')} — ${L('apostaste','you picked')} ${pick}${pase}`;
    case 'voided':
      return outcome.advanced && outcome.advancePts
        ? `✓ ${L('Acertaste el pase','Right advance')} +${outcome.advancePts} · 💀 ${L('tu marcador cae en el cementerio del cruce','your scoreline falls in the match graveyard')} (${pick})`
        : `💀 ${L('Cementerio del cruce','Match graveyard')} — ${L('acertaste el signo pero cayó tu rama, 0 pts','right outcome but your branch fell, 0 pts')} (${pick})`;
    case 'miss': return `✗ ${L('Fallaste','Missed')} — ${L('tenías','you had')} ${pick}${pase}`;
    case 'advance': return `✓ ${L('Acertaste el pase','Correct advance')}${outcome.advancePts ? ` +${outcome.advancePts}` : ''}`;
    default: return '';
  }
}

function meResultStripHtml(m){
  if(!ME) return '';
  const outcome = meRecentOutcome(m);
  if(!outcome) return '';
  // Si acertaste el pase, la franja es positiva aunque el marcador caiga en el cementerio.
  const scored = outcome.advanced && outcome.advancePts;
  const cls = scored ? '' : outcome.kind === 'miss' ? ' miss' : outcome.kind === 'voided' ? ' voided' : outcome.kind === 'sign' ? ' sign' : '';
  return `<div class="me-result-strip${cls}">${meResultStripText(outcome)}</div>`;
}

function meBetBlockHtml(m){
  if(!ME) return '';
  const pick = mePickInMatch(m);
  if(!pick) return '';
  const knockout = !!m.is_knockout;
  const dead = branchDeadForPick(m, pick);
  const scoreVoid = scorelineVoidForPick(m, pick);
  const pickLbl = pickScoreLbl(pick, knockout);
  const pickScore = pick.home != null ? `${pick.home}-${pick.away}` : '';
  const consensus = knockout
    ? (m.score && m.score.value ? m.score.value : null)
    : (m.modal_scoreline || null);
  const consensusPct = knockout
    ? (m.score ? Math.round(m.score.agreement || 0) : null)
    : (m.modal_scoreline_share != null ? Math.round(m.modal_scoreline_share * 100) : null);
  let consensusHtml = '';
  if(consensus && !dead && !scoreVoid){
    const agree = pickScore === consensus;
    const badge = agree
      ? `<span class="badge agree">${L('Vas con el pueblo','With the crowd')}</span>`
      : `<span class="badge disagree">${L('Discrepas','You disagree')}</span>`;
    consensusHtml = `<div class="me-bet-consensus">${L('Consenso','Consensus')}: <b>${esc(consensus)}</b>${consensusPct != null ? ` (${consensusPct}%)` : ''} ${badge}</div>`;
  }
  let stakeHtml = '';
  const st = m.stake;
  const meSt = st ? meStakePerson(st) : null;
  const meSwingHints = () => {
    if(!meSt || (!meSt.swing_up && !meSt.swing_down)) return '';
    const parts = [];
    if(meSt.swing_up) parts.push(`+${meSt.swing_up}`);
    if(meSt.swing_down) parts.push(`−${meSt.swing_down}`);
    let hints = `<div style="margin-top:6px">${L('Tu swing','Your swing')}: <span class="swing">${parts.join(' / ')}</span> ${L('puestos','places')}</div>`;
    if(meSt.swing_up && meSt.best_result){
      hints += `<div class="muted" style="font-size:.75rem;margin-top:2px">+${meSt.swing_up} ${L('si sale','if')} <b>${stakeResultLbl(meSt.best_result)}</b></div>`;
    }
    if(meSt.swing_down && meSt.worst_result){
      hints += `<div class="muted" style="font-size:.75rem;margin-top:2px">−${meSt.swing_down} ${L('si sale','if')} <b>${stakeResultLbl(meSt.worst_result)}</b></div>`;
    }
    return hints;
  };
  if(dead){
    stakeHtml = `<div class="me-bet-void">💀 <b>${L('No importa','Doesn\'t matter')}</b><span>${L('pusiste que pasa','you picked')} ${pick.winner_flag || ''} ${esc(team(pick.winner))} — ${L('no está en este cruce, 0 pts','not in this tie, 0 pts')}</span></div>`;
  } else if(scoreVoid){
    const nm = v => (v || '').toString().trim().toLowerCase();
    const other = nm(pick.winner) === nm(m.home) ? m.away : m.home;
    const adv = (meSt && meSt.max_pts) || m.advance_points || 0;
    stakeHtml = `<div class="me-bet-void">💀 <b>${L('Tu marcador no cuenta','Your scoreline is void')}</b><span>${L('acertaste que pasa','you have')} ${pick.winner_flag || ''} ${esc(team(pick.winner))}, ${L('pero el rival del cruce es','but the opponent in this tie is')} ${esc(team(other))} ${L('(tú pusiste otro) — solo el pase sigue vivo','(you picked another) — only the advance point is live')}${adv ? ` (${L('hasta','up to')} +${adv})` : ''}</span>${meSwingHints()}</div>`;
  } else if(st && st.deferred){
    stakeHtml = `<div class="me-bet-stake muted">${L('Tu swing se calcula cuando caigan los resultados anteriores de hoy', 'Your swing is calculated once earlier today\'s results are in')}</div>`;
  } else if(meSt && (meSt.swing_up || meSt.swing_down || meSt.max_pts)){
    const parts = [];
    if(meSt.swing_up) parts.push(`+${meSt.swing_up}`);
    if(meSt.swing_down) parts.push(`−${meSt.swing_down}`);
    let hints = '';
    if(meSt.swing_up && meSt.best_result){
      hints += `<div class="muted" style="font-size:.75rem;margin-top:4px">+${meSt.swing_up} ${L('si sale','if')} <b>${stakeResultLbl(meSt.best_result)}</b></div>`;
    }
    if(meSt.swing_down && meSt.worst_result){
      hints += `<div class="muted" style="font-size:.75rem;margin-top:2px">−${meSt.swing_down} ${L('si sale','if')} <b>${stakeResultLbl(meSt.worst_result)}</b></div>`;
    }
    stakeHtml = `<div class="me-bet-stake">${L('En juego','At stake')}: ${L('hasta','up to')} +${meSt.max_pts} pts${parts.length ? ` · <span class="swing">${parts.join(' / ')}</span> ${L('puestos','places')}` : ''}${hints}</div>`;
  } else if(!st && m.result){
    const outcome = meTodayOutcome(m);
    if(outcome) stakeHtml = `<div class="me-bet-stake">${meResultStripText(outcome)}</div>`;
  }
  return `<div class="me-bet-block" style="--me-color:${personColor(ME)}">
    <div class="me-bet-head">${L('Tu apuesta','Your pick')}</div>
    <div class="me-bet-pick">${esc(pickLbl)}</div>
    ${consensusHtml}
    ${stakeHtml}
  </div>`;
}

function koStakeSwingers(st){
  const all = (st.people || []).filter(p => p.swing_up > 0 || p.swing_down > 0)
    .sort((a,b) => (b.swing_up - a.swing_up) || (b.swing_down - a.swing_down) || b.max_pts - a.max_pts);
  if(!ME) return all.slice(0, 4);
  const meEntry = all.find(p => p.name === ME);
  const top = all.slice(0, 4);
  if(meEntry && !top.some(p => p.name === ME)){
    return all.slice(0, 3).concat([meEntry]);
  }
  return top;
}

function groupMatchStakeHtml(m){
  const st = m.stake;
  if(!st) return '';
  if(st.deferred){
    return `<div class="kp-stake tm-stake">
      <div><div style="font:800 1.05rem var(--fu)">${L('En juego en el ranking','At stake in the standings')}</div></div>
      <div style="text-align:right">${stakeDeferredHtml(st)}</div>
    </div>`;
  }
  return `<div class="kp-stake tm-stake">
    <div><div style="font:800 1.05rem var(--fu)">${L('En juego en el ranking','At stake in the standings')}</div>
      <div class="muted" style="font-size:.82rem">${L('Hasta','Up to')} ${st.max_points} ${L('pts repartibles','pts on the table')} · ${st.picks}/${N} ${L('con apuesta','with a pick')}</div></div>
    <div style="text-align:right">${stakeSwingHtml(st)}</div>
  </div>`;
}

function koMatchStakeHtml(m){
  const st = m.stake;
  if(!st) return '';
  if(st.deferred){
    return `<div class="kp-stake tm-stake">
      <div><div style="font:800 1.05rem var(--fu)">${L('En juego (eliminatorias)','At stake (knockouts)')}</div></div>
      <div style="text-align:right">${stakeDeferredHtml(st)}</div>
    </div>`;
  }
  const swingers = koStakeSwingers(st);
  const swingHtml = swingers.length
    ? `<div class="muted" style="font-size:.82rem;margin-top:8px">${L('Más en juego','Most at stake')}: ${swingers.map(p => {
        const mv = [];
        if(p.swing_up) mv.push(`+${p.swing_up}${p.best_result ? ` (${stakeResultLbl(p.best_result)})` : ''}`);
        if(p.swing_down) mv.push(`−${p.swing_down}${p.worst_result ? ` (${stakeResultLbl(p.worst_result)})` : ''}`);
        return `<b>${esc(p.name)}</b> ${mv.join(' · ')}`;
      }).join(' · ')}</div>`
    : '';
  return `<div class="kp-stake tm-stake">
    <div><div style="font:800 1.05rem var(--fu)">${L('En juego (eliminatorias)','At stake (knockouts)')}</div>
      <div class="muted" style="font-size:.82rem">${L('Hasta','Up to')} +${st.max_one} ${L('pts/persona','pts/person')} (+${st.advance_points} ${L('por pase','per advance')}) · ${st.picks}/${N} ${L('con apuesta','with a pick')}</div></div>
    <div style="text-align:right">${stakeSwingHtml(st)}</div>
    ${swingHtml}
  </div>`;
}

function branchDeadForPick(m, pick){
  // Un pick de cruce está "muerto" si el equipo que pusiste como ganador no es
  // ninguno de los dos del cruce: apostaste por un equipo que no llegó / ya está
  // fuera, así que tu apuesta de este partido no puede puntuar (0 pts).
  if(!m || !m.is_knockout || !pick || !pick.winner) return false;
  const home = m.home || '', away = m.away || '';
  if(!home || !away || /^W\d+$/i.test(home) || /^W\d+$/i.test(away)) return false; // cruce sin resolver
  const nm = v => (v || '').toString().trim().toLowerCase();
  const w = nm(pick.winner);
  return w !== nm(home) && w !== nm(away);
}

function scorelineVoidForPick(m, pick){
  // Cruce ya resuelto donde acertaste el equipo que pasa (tu pick NO está muerto)
  // pero fallaste el RIVAL: tú pusiste el marcador de otro cruce, así que tu
  // signo/pleno no puntúa. Solo el punto del pase sigue vivo. Es el "cementerio
  // del cruce" pero detectable ANTES de que se juegue (los feeders ya cayeron).
  if(!m || !m.is_knockout || !pick || pick.home == null) return false;
  if(pick.matchup_ok !== false) return false;
  return !branchDeadForPick(m, pick);
}

function todayPicksHtml(picks, match){
  if(!picks || !picks.length) return '';
  const knockout = !!match.is_knockout;
  const buckets = {'1': [], 'X': [], '2': [], '?': [], dead: []};
  picks.forEach(p => {
    if(branchDeadForPick(match, p)){ buckets.dead.push(p); return; }
    let key = '?';
    if(p.home != null && p.away != null){
      key = p.home > p.away ? '1' : p.home === p.away ? 'X' : '2';
    }
    buckets[key].push(p);
  });
  const sort = arr => sortPeople(arr);
  let anyVoid = false;
  const pill = p => {
    const score = p.home != null ? `${p.home}-${p.away}` : '–';
    let extra = '';
    if(knockout && p.home != null && p.away != null && p.home === p.away && p.winner){
      extra = ` · ${p.winner_flag || ''} ${esc(team(p.winner))}`;
    }
    if(scorelineVoidForPick(match, p)){
      anyVoid = true;
      return `<span class="result-person${meClass(p.name)}" title="${L('acertó el equipo que pasa pero no el rival del cruce → solo suma el pase','right side advancing but wrong opponent → only the advance point counts')}"><b>${esc(p.name)}</b><span><s style="opacity:.55">${score}</s> · <span style="color:var(--gold);font-weight:700">${L('solo pase','advance only')}</span>${extra}</span></span>`;
    }
    return `<span class="result-person${meClass(p.name)}"><b>${esc(p.name)}</b><span>${score}${extra}</span></span>`;
  };
  const people = arr => arr.length
    ? arr.map(pill).join('')
    : `<span class="muted">${L('Nadie','Nobody')}</span>`;
  const homeLbl = esc(team(match.home));
  const awayLbl = esc(team(match.away));
  const boxes = [
    {key: '1', cls: 'home-win', title: L(`Gana ${homeLbl}`, `Win ${homeLbl}`)},
    {key: 'X', cls: 'draw', title: L('Empate', 'Draw')},
    {key: '2', cls: 'away-win', title: L(`Gana ${awayLbl}`, `Win ${awayLbl}`)},
  ];
  let html = boxes.map(b => {
    const list = sort(buckets[b.key]);
    return `<div class="result-box ${b.cls}">
      <div class="rb-title">${b.title} <span class="rb-count">${list.length}</span></div>
      <div class="result-names">${people(list)}</div>
    </div>`;
  }).join('');
  if(buckets.dead.length){
    const list = sort(buckets.dead);
    const deadPill = p => {
      const score = p.home != null ? `${p.home}-${p.away}` : '–';
      return `<span class="result-person${meClass(p.name)}"><b>${esc(p.name)}</b><span>${score} · 💀 ${p.winner_flag || ''} ${esc(team(p.winner))}</span></span>`;
    };
    html += `<div class="result-box dead">
      <div class="rb-title">💀 ${L('Apuesta muerta','Dead pick')} <span class="rb-count">${list.length}</span></div>
      <div class="rb-sub">${L('su ganador no está en este cruce → 0 pts','their winner isn\'t in this tie → 0 pts')}</div>
      <div class="result-names">${list.map(deadPill).join('')}</div>
    </div>`;
  }
  if(buckets['?'].length){
    const list = sort(buckets['?']);
    html += `<div class="result-box pick-empty">
      <div class="rb-title">${L('Sin apuesta','No pick')} <span class="rb-count">${list.length}</span></div>
      <div class="result-names">${people(list)}</div>
    </div>`;
  }
  const voidNote = anyVoid ? `<div class="muted" style="font-size:.78rem;margin:2px 0 6px">${L('Marcador tachado','Struck scoreline')} = ${L('acertó el equipo que pasa pero falló el rival del cruce → su marcador no puntúa, solo el punto del pase','right team advancing but wrong opponent → scoreline scores nothing, only the advance point')}.</div>` : '';
  return `<div class="today-picks"><div class="tp-title">${L('Qué ha puesto cada uno','What everyone picked')}</div>${voidNote}<div class="result-groups pick-groups">${html}</div></div>`;
}


function hoyDateLabel(dateStr){
  const dateObj = new Date(dateStr + 'T12:00:00');
  const opts = {weekday:'long', day:'numeric', month:'long', year:'numeric'};
  const str = dateObj.toLocaleDateString(LANG==='es'?'es-ES':'en-US', opts);
  return str.charAt(0).toUpperCase() + str.slice(1);
}
function setHoyDate(d){ HOY_DATE = d; rebuild(); }
function buildHoy(){
  const s = section('hoy', L('⚽ Hoy','⚽ Today'),
    L('Los partidos de hoy','Today\'s matches'),
    L('Qué se juega hoy, qué ha puesto cada uno y cuánto puede mover el ranking.','What\'s on today, everyone\'s picks, and how much the standings can swing.'));
  const allM = todayScheduleMatches();
  const realToday = matchdayDateStr(new Date());
  const dates = [...new Set(allM.map(m => m.date).filter(Boolean))].sort();
  const viewDate = (HOY_DATE && dates.includes(HOY_DATE)) ? HOY_DATE : realToday;
  const before = dates.filter(d => d < viewDate);
  const after = dates.filter(d => d > viewDate);
  const prevDate = before.length ? before[before.length - 1] : null;
  const nextDate = after.length ? after[0] : null;
  const nav = el('div','hoy-nav reveal',
    `<button class="hoy-arrow" data-dir="prev"${prevDate?'':' disabled'} aria-label="${L('Día anterior','Previous day')}">‹</button>
     <div class="hoy-nav-mid">
       <div class="today-date">${esc(hoyDateLabel(viewDate))}</div>
       ${viewDate !== realToday ? `<button class="hoy-today-btn" data-action="today">${L('Volver a hoy','Back to today')}</button>` : ''}
     </div>
     <button class="hoy-arrow" data-dir="next"${nextDate?'':' disabled'} aria-label="${L('Día siguiente','Next day')}">›</button>`);
  nav.addEventListener('click', e => {
    if(e.target.closest('[data-action="today"]')){ setHoyDate(null); return; }
    const btn = e.target.closest('[data-dir]');
    if(!btn || btn.disabled) return;
    const target = btn.dataset.dir === 'prev' ? prevDate : nextDate;
    if(target) setHoyDate(target);
  });
  s.appendChild(nav);
  const todayMatches = allM.filter(m => m.date === viewDate)
    .sort((a,b) => (a.dt||'').localeCompare(b.dt||'') || a.code.localeCompare(b.code));
  if(!todayMatches.length){
    const upcoming = allM.filter(m => m.date > viewDate).sort((a,b) => a.date.localeCompare(b.date) || (a.dt||'').localeCompare(b.dt||'') || a.code.localeCompare(b.code)).slice(0,6);
    let nextHtml = '';
    if(upcoming.length){
      const grouped = {};
      upcoming.forEach(m => { if(!grouped[m.date]) grouped[m.date] = []; grouped[m.date].push(m); });
      nextHtml = '<div class="next">' + Object.entries(grouped).map(([dt, ms]) =>
        ms.map(m => `<div class="next-match"><span class="next-date" title="${koTime(m)?esc(koTz()):''}">${dt.slice(5)}${koTime(m)?' · '+esc(koTime(m))+(koNext(m)?'<span class="tm-next">+1</span>':''):''}</span>${m.home_flag} ${esc(team(m.home))} – ${esc(team(m.away))} ${m.away_flag}</div>`).join('')
      ).join('') + '</div>';
    }
    const emptyTitle = viewDate === realToday ? L('Hoy no hay partidos','No matches today') : L('Ese día no hay partidos','No matches that day');
    s.appendChild(el('div','card no-today reveal',
      `<img class="no-today-img" src="pablo.jpg" alt="" loading="lazy"><h3 style="margin:10px 0 6px">${emptyTitle}</h3>
       <p class="muted">${L('Próximos partidos:','Upcoming matches:')}</p>${nextHtml}`));
    return;
  }
  todayMatches.forEach(m => {
    if(m.is_knockout){
      const resultHtml = m.result && m.result.score
        ? `<div class="tm-stat"><div class="score-final">${m.result.score.home}-${m.result.score.away}</div><div class="lab">${L('Resultado final','Final result')}${m.result.winner ? ` · ${m.result.winner_flag || ''} ${esc(team(m.result.winner))}` : ''}</div></div>`
        : '';
      const winnerHtml = m.winner && m.winner.value
        ? `<div class="tm-stat"><div class="val" style="font-size:1rem">${m.winner.flag || ''} ${esc(team(m.winner.value))}</div><div class="lab">${L('Consenso ganador','Winner consensus')} · ${pf(m.winner.agreement)} · ${m.winner.count}/${N}</div></div>`
        : '';
      const scoreHtml = m.score && m.score.value
        ? `<div class="tm-stat"><div class="val">${esc(m.score.value)}</div><div class="lab">${L('Marcador más repetido','Most common scoreline')} (${pf(Math.round((m.score.agreement || 0)))})</div></div>`
        : '';
      const statsCount = 1 + (resultHtml ? 1 : 0) + (winnerHtml ? 1 : 0) + (scoreHtml ? 1 : 0);
      // En eliminatorias cada uno predijo su propio cuadro, así que el signo del
      // marcador (local/visitante del hueco) NO equivale a "gana <equipo real>".
      // Mostramos el reparto real: a qué equipo puso cada uno ganando el cruce.
      const wd = (m.winner && m.winner.dist) || [];
      // Cruce ya resuelto (equipos reales, no huecos W##/RU): los picks de
      // equipos que no están en el cruce ya no pueden ganar → se marcan en gris.
      const koResolved = m.home && m.away && !/^(W\d|RU)/.test(m.home) && !/^(W\d|RU)/.test(m.away);
      const koAlive = new Set(koResolved ? [m.home, m.away] : []);
      const outcomeHtml = wd.length ? `<div class="tm-stats" style="grid-template-columns:repeat(${Math.min(wd.length,4)},1fr)">
        ${wd.slice(0,4).map(d => { const dead = koResolved && !koAlive.has(d.value); return `<div class="tm-stat${dead?' tm-stat-dead':''}"${dead?` title="${L('Ya no está en este cruce','No longer in this tie')}"`:''}><div class="val">${N ? Math.round(d.count/N*100) : 0}%</div><div class="lab">${L('Gana','Win')} ${d.flag||''} ${esc(team(d.value))}${dead?' <span class="dead-x">✕</span>':''}</div></div>`; }).join('')}
      </div>` : '';
      const picksHtml = todayPicksHtml(m.picks, m);
      s.appendChild(el('div','today-match reveal',
        `<div class="tm-head">
          <div class="tm-teams">${m.home_flag ? m.home_flag + ' ' : ''}${esc(team(m.home))} – ${esc(team(m.away))}${m.away_flag ? ' ' + m.away_flag : ''}</div>
          <div class="tm-tags">${koTime(m)?`<span class="tm-time" title="${esc(koTz())}">⏱ ${esc(koTime(m))}${koNext(m)?` <span class="tm-next" title="${L('madrugada del día siguiente','after midnight, next day')}">+1</span>`:''}</span>`:''}<span class="tm-group">${L(m.phase_es || 'ELIMINATORIA', m.phase_en || 'KNOCKOUT')}</span></div>
        </div>
        <div class="tm-stats" style="grid-template-columns:repeat(${statsCount},1fr)">
          <div class="tm-stat"><div class="val" style="font-size:1rem">${esc(m.venue || '')}</div><div class="lab">${esc(m.city || '')}</div></div>
          ${resultHtml}
          ${winnerHtml}
          ${scoreHtml}
        </div>
        ${outcomeHtml}
        ${meBetBlockHtml(m)}
        ${koMatchStakeHtml(m)}
        ${picksHtml}
        ${m.home_trivia && m.home_trivia.es ? `<div class="trivia-block">
        <div class="trivia-item">
          <div class="trivia-flag">${L('🤯 ¿Sabías que…?','🤯 Did you know…?')} ${m.home_flag} ${esc(team(m.home))}</div>
          <div class="trivia-text">${esc(L(m.home_trivia.es, m.home_trivia.en))}</div>
        </div>
        <div class="trivia-item">
          <div class="trivia-flag">${L('🤯 ¿Sabías que…?','🤯 Did you know…?')} ${m.away_flag} ${esc(team(m.away))}</div>
          <div class="trivia-text">${esc(L(m.away_trivia.es, m.away_trivia.en))}</div>
        </div>
      </div>` : ''}`));
      return;
    }
    const o = m.outcome_dist, tot = (o['1']||0)+(o['X']||0)+(o['2']||0);
    const pct = k => tot ? Math.round((o[k]||0)/tot*100) : 0;
    const uniqueHtml = m.most_unique_pick
      ? `<span class="muted">${L('🔥 El más atrevido:','🔥 Boldest call:')}</span> <b>${esc(m.most_unique_pick.name)}</b> <span class="mint">${m.most_unique_pick.score}</span>`
      : '';
    const picksHtml = todayPicksHtml(m.picks, m);
    s.appendChild(el('div','today-match reveal',
      `<div class="tm-head">
        <div class="tm-teams">${m.home_flag} ${esc(team(m.home))} – ${esc(team(m.away))} ${m.away_flag}</div>
        <div class="tm-tags">${koTime(m)?`<span class="tm-time" title="${esc(koTz())}">⏱ ${esc(koTime(m))}${koNext(m)?` <span class="tm-next" title="${L('madrugada del día siguiente','after midnight, next day')}">+1</span>`:''}</span>`:''}<span class="tm-group">${L('GRUPO','GROUP')} ${m.group}</span></div>
      </div>
      <div class="tm-stats">
        <div class="tm-stat"><div class="val">${pct('1')}%</div><div class="lab">${L('Gana','Win')} ${esc(team(m.home))}</div></div>
        <div class="tm-stat"><div class="val">${pct('X')}%</div><div class="lab">${L('Empate','Draw')}</div></div>
        <div class="tm-stat"><div class="val">${pct('2')}%</div><div class="lab">${L('Gana','Win')} ${esc(team(m.away))}</div></div>
      </div>
      <div class="tm-stats" style="grid-template-columns:1fr 1fr">
        <div class="tm-stat"><div class="val">${m.modal_scoreline}</div><div class="lab">${L('Marcador más repetido','Most common scoreline')} (${pf(Math.round(m.modal_scoreline_share*100))})</div></div>
        <div class="tm-stat"><div class="val" style="font-size:1rem">${uniqueHtml||'–'}</div><div class="lab">${L('Pick único más salvaje','Wildest unique pick')}</div></div>
      </div>
      ${meBetBlockHtml(m)}
      ${groupMatchStakeHtml(m)}
      ${picksHtml}
      <div class="trivia-block">
        <div class="trivia-item">
          <div class="trivia-flag">${L('🤯 ¿Sabías que…?','🤯 Did you know…?')} ${m.home_flag} ${esc(team(m.home))}</div>
          <div class="trivia-text">${esc(L(m.home_trivia.es, m.home_trivia.en))}</div>
        </div>
        <div class="trivia-item">
          <div class="trivia-flag">${L('🤯 ¿Sabías que…?','🤯 Did you know…?')} ${m.away_flag} ${esc(team(m.away))}</div>
          <div class="trivia-text">${esc(L(m.away_trivia.es, m.away_trivia.en))}</div>
        </div>
      </div>`));
  });
}

/* ---- ÚLTIMOS RESULTADOS ---- */
function buildUltimos(){
  const s = section('ultimos', L('🧾 Últimos','🧾 Latest'),
    L('Últimos partidos jugados','Latest finished matches'),
    L('Resultado final y reparto de alegrías: plenos, signos acertados y los que han palmado.',
      'Final score and who got it right: exact hits, correct outcomes and misses.'));
  const recent = (D.recent_results && D.recent_results.matches) || [];
  if(!recent.length){
    s.appendChild(el('div','card teaser reveal',
      `<div class="em">🧾</div><h3 style="margin:10px 0 6px">${L('Aún no hay resultados cargados','No results loaded yet')}</h3>
       <p class="muted">${L('Cuando rellenes marcadores en <b>Real results</b>, aquí aparecerán los últimos partidos terminados.','Once scores are filled in <b>Real results</b>, the latest finished matches will appear here.')}</p>`));
    return;
  }
  s.appendChild(el('div','muted reveal', `${recent.length} ${L('últimos de','latest of')} ${D.recent_results.total} ${L('partidos jugados','finished matches')}`));
  recent.forEach(m => {
    const people = arr => arr.length
      ? arr.map(p => `<span class="result-person${meClass(p.name)}"><b>${esc(p.name)}</b><span>${esc(p.pick)}</span></span>`).join('')
      : `<span class="muted">${L('Nadie','Nobody')}</span>`;
    const dateObj = new Date(m.date + 'T12:00:00');
    const dateStr = dateObj.toLocaleDateString(LANG==='es'?'es-ES':'en-US', {day:'numeric', month:'short'});
    const phase = m.is_knockout ? L(m.phase_es || 'Eliminatorias', m.phase_en || 'Knockouts') : `${L('Grupo','Group')} ${m.group}`;
    const winner = m.result && m.result.winner ? `<span>${m.result.winner_flag || ''} ${esc(team(m.result.winner))}</span>` : '';
    const advanceBox = m.is_knockout ? `
        <div class="result-box">
          <div class="rb-title">${L('Pase','Advance')} <span class="rb-count">${(m.advance || []).length}</span></div>
          <div class="result-names">${people(m.advance || [])}</div>
        </div>` : '';
    const graveBox = (m.voided && m.voided.length) ? `
        <div class="result-box voided">
          <div class="rb-title">💀 ${L('Cementerio del cruce','Match graveyard')} <span class="rb-count">${m.voided.length}</span></div>
          <div class="rb-sub">${L('el signo era suyo, la rama no','right outcome, wrong branch')}</div>
          <div class="result-names">${people(m.voided)}</div>
        </div>` : '';
    s.appendChild(el('div','today-match reveal',
      `<div class="tm-head">
        <div class="tm-teams">${m.home_flag} ${esc(team(m.home))} – ${esc(team(m.away))} ${m.away_flag}</div>
        <span class="score-final">${m.result.home}-${m.result.away}</span>
      </div>
      <div class="recent-meta"><span>${dateStr}</span><span>${phase}</span><span>${L('Resultado final','Final score')}</span>${winner}</div>
      ${meResultStripHtml(m)}
      <div class="result-groups">
        <div class="result-box">
          <div class="rb-title">${L('Pleno','Exact score')} <span class="rb-count">${m.exact.length}</span></div>
          <div class="result-names">${people(m.exact)}</div>
        </div>
        <div class="result-box">
          <div class="rb-title">${L('Signo','Outcome')} <span class="rb-count">${m.sign.length}</span></div>
          <div class="result-names">${people(m.sign)}</div>
        </div>
        ${graveBox}
        <div class="result-box miss">
          <div class="rb-title">${L('Palmada','Missed')} <span class="rb-count">${m.miss.length}</span></div>
          <div class="result-names">${people(m.miss)}</div>
        </div>
        ${advanceBox}
      </div>`));
  });
}

/* ---- REBELDÍA ---- */
function buildRebeldia(){
  const s = section('rebeldia', L('01 · La estrella','01 · The star'),
    L('El ranking de rebeldía 🐺','The maverick ranking 🐺'),
    L('Combina dos cosas en cada partido: cuántas veces eliges el signo (1/X/2) en minoría, y cuánto se aleja tu marcador del marcador típico del grupo. 100 = el más inconformista. 0 = el más borrego.',
      'Two things per match: how often you back the (1/X/2) outcome in the minority, and how far your scoreline sits from the group typical one. 100 = the biggest maverick. 0 = the biggest sheep.'));
  const R = D.rebeldia, max = R[0].index || 100;
  const pod = el('div','podium reveal');
  const order = [1,0,2], medals = ['🥈','🥇','🥉'];
  order.forEach((oi,k) => { const r = R[oi]; if(!r) return;
    pod.appendChild(el('div','pod '+(oi===0?'p1':''),
      `<div class="medal">${medals[k]}</div><div class="nm">${esc(r.name)}</div>
       <div class="sc" data-count="${r.index}" data-decimals="1">0</div><div class="k">${L('índice','index')}</div>`)); });
  s.appendChild(pod);
  const list = el('div','reveal');
  R.forEach((r,i) => {
    const badge = i===0?'🐺 ':(i===R.length-1?'🐑 ':'');
    const cls = i===R.length-1?'gold':'';
    list.appendChild(el('div','bar-row'+meClass(r.name),
      `<div class="bar-rank">${i+1}</div><div class="bar-name">${badge}${esc(r.name)}</div>
       <div class="bar-track"><div class="bar-fill ${cls}" data-w="${(r.index/max*100).toFixed(1)}"></div></div>
       <div class="bar-val" data-count="${r.index}" data-decimals="1">0</div>`));
  });
  s.appendChild(list);
}

/* ---- AFINIDAD ---- */
function buildAfinidad(){
  const s = section('gemelas', L('02 · Afinidad','02 · Affinity'),
    L('¿Quién piensa como quién? 💞','Who thinks like who? 💞'),
    L('Comparamos las quinielas marcador a marcador. Cuanto más brillante el cuadro, más coinciden esas dos personas en el resultado exacto.',
      'We compare predictions scoreline by scoreline. The brighter the cell, the more those two people match on the exact result.'));
  const bp = D.best_pair, wp = D.worst_pair;
  const duo = el('div','duo reveal');
  duo.innerHTML = `
    <div class="card glow"><span class="tag">${L('💞 Almas gemelas','💞 Soulmates')}</span>
      <div class="names">${esc(bp.a)} & ${esc(bp.b)}</div>
      <div class="muted">${bp.pct}% ${L('de marcadores','of scorelines')} <b class="mint">${L('idénticos','identical')}</b>${bp.pct>=100?L(' — sí, los 72. 👀 ¿alguien ha copiado?',' — yes, all 72. 👀 did someone copy?'):''}</div></div>
    <div class="card"><span class="tag cool">${L('🧊 Polos opuestos','🧊 Opposites')}</span>
      <div class="names">${esc(wp.a)} & ${esc(wp.b)}</div>
      <div class="muted">${L('los que más se contradicen (distancia media '+wp.dist+' goles por partido)','the most contradictory pair (avg distance '+wp.dist+' goals per match)')}</div></div>`;
  s.appendChild(duo);
  const names = D.matrix.names, sim = D.matrix.sim;
  const meIdx = ME ? names.indexOf(ME) : -1;
  let mx = 0; for(let i=0;i<N;i++) for(let j=0;j<N;j++) if(i!==j) mx = Math.max(mx, sim[i][j]);
  const wrapm = el('div','matrix-wrap reveal');
  const grid = el('div','matrix');
  grid.style.gridTemplateColumns = `120px repeat(${N},19px)`;
  grid.appendChild(el('div','mlabel'));
  names.forEach((n,j) => grid.appendChild(el('div','mlabel col'+meClass(n), esc(n))));
  for(let i=0;i<N;i++){
    grid.appendChild(el('div','mlabel row'+meClass(names[i]), esc(names[i])));
    for(let j=0;j<N;j++){
      if(i===j){ grid.appendChild(el('div','mcell mdiag')); continue; }
      const v = sim[i][j], a = mx ? Math.max(.05, v/mx) : 0;
      const c = el('div','mcell'+(i===meIdx||j===meIdx?' is-me':''));
      c.style.background = `rgba(108,232,105,${a.toFixed(3)})`;
      c.dataset.t = `${esc(names[i])} ↔ ${esc(names[j])} · ${v}% ${L('idénticos','identical')}`;
      grid.appendChild(c);
    }
  }
  wrapm.appendChild(grid);
  wrapm.appendChild(el('div','legend',`<span>${L('menos afín','less alike')}</span><span class="scale"></span><span>${L('más afín','more alike')}</span>`));
  s.appendChild(wrapm);
  grid.addEventListener('mouseover', e => { const t = e.target.dataset.t; if(t){ tip.textContent = t; tip.style.opacity = 1; } });
  grid.addEventListener('mousemove', e => { tip.style.left = (e.clientX+14)+'px'; tip.style.top = (e.clientY+14)+'px'; });
  grid.addEventListener('mouseout', () => { tip.style.opacity = 0; });
}

/* ---- ESTILO ---- */
function buildEstilo(){
  const s = section('estilo', L('03 · Estilo','03 · Style'),
    L('Goleadores vs cerrojos 🥅','Goal-fests vs parked buses 🥅'),
    L('Media de goles que cada uno mete en sus marcadores, qué marcadores se repiten más en toda la porra y quién es el rey del empate.',
      'Average goals in everyone scorelines, the most repeated results across the pool, and who loves a draw.'));
  const st = D.style.slice().sort((a,b) => b.avg_goals-a.avg_goals);
  const gol = st[0], cer = st[st.length-1];
  const emp = D.style.slice().sort((a,b) => b.pct_draws-a.pct_draws)[0];
  const duo = el('div','grid g3 reveal');
  duo.innerHTML = `
    <div class="card glow"><span class="k">${L('🥅 El goleador','🥅 The goal machine')}</span><h3>${esc(gol.name)}</h3>
      <div class="big-num">${gol.avg_goals}</div><div class="muted">${L('goles de media por partido','avg goals per match')}</div></div>
    <div class="card"><span class="k">${L('🔒 El cerrojo','🔒 The parked bus')}</span><h3>${esc(cer.name)}</h3>
      <div class="big-num">${cer.avg_goals}</div><div class="muted">${L('el más tacaño con los goles','the stingiest with goals')}</div></div>
    <div class="card"><span class="k">${L('🤝 Amigo del empate','🤝 Draw lover')}</span><h3>${esc(emp.name)}</h3>
      <div class="big-num">${pf(emp.pct_draws)}</div><div class="muted">${L('de sus partidos, en tablas','of their matches drawn')}</div></div>`;
  s.appendChild(duo);
  const list = el('div','reveal'); list.style.marginTop = '22px';
  const mxg = st[0].avg_goals;
  list.appendChild(el('div','k', L('Media de goles por quiniela','Average goals per prediction')));
  st.forEach((r,i) => list.appendChild(el('div','bar-row'+meClass(r.name),
    `<div class="bar-rank">${i+1}</div><div class="bar-name">${esc(r.name)}</div>
     <div class="bar-track"><div class="bar-fill" data-w="${(r.avg_goals/mxg*100).toFixed(1)}"></div></div>
     <div class="bar-val" data-count="${r.avg_goals}" data-decimals="2">0</div>`)));
  s.appendChild(list);
  const cs = el('div','card reveal'); cs.style.marginTop = '22px';
  cs.innerHTML = `<span class="k">${L('Marcadores más apostados (de las '+fmt(N*D.hero.matches)+' apuestas)','Most-predicted scorelines (out of '+fmt(N*D.hero.matches)+' bets)')}</span>`;
  const mxc = D.common_scorelines[0].count;
  D.common_scorelines.forEach(c => cs.appendChild(el('div','bar-row',
    `<div class="bar-rank">${c.score}</div><div class="bar-name muted">${pf(c.pct)}</div>
     <div class="bar-track"><div class="bar-fill" data-w="${(c.count/mxc*100).toFixed(1)}"></div></div>
     <div class="bar-val" data-count="${c.count}">0</div>`)));
  s.appendChild(cs);
}

/* ---- FAVORITOS ---- */
function buildFavoritos(){
  const s = section('favoritos', L('04 · Favoritos','04 · Favourites'),
    L('Los favoritos del torneo 🏆','Tournament favourites 🏆'),
    L('El consenso de la oficina: a quién ve todo el mundo ganando su grupo, qué selecciones se respetan más y qué grupo nos tiene más divididos.',
      'The office consensus: who everyone sees topping their group, the most respected teams, and the group that divides us most.'));
  const gc = D.group_consensus.slice().sort((a,b) => b.agreement-a.agreement);
  const grid = el('div','grid g4 reveal');
  gc.forEach(g => grid.appendChild(el('div','gcard',
    `<div class="gl">${L('GRUPO','GROUP')} ${g.group}</div>
     <div class="fav">${g.flag} ${esc(team(g.favorite))}</div>
     <div class="bar-track"><div class="bar-fill" data-w="${g.agreement}"></div></div>
     <div class="muted" style="font-size:.8rem;margin-top:6px">${pf(g.agreement)} ${L('lo ve primero','see them top')}</div>`)));
  s.appendChild(grid);
  const mr = el('div','card reveal'); mr.style.marginTop = '22px';
  mr.innerHTML = `<span class="k">${L('Selecciones más respetadas (prestigio: 1.º=3, 2.º=2, 3.º=1)','Most respected teams (prestige: 1st=3, 2nd=2, 3rd=1)')}</span>`;
  const mxr = D.most_respected[0].count;
  D.most_respected.forEach((t,i) => mr.appendChild(el('div','bar-row',
    `<div class="bar-rank">${i+1}</div><div class="bar-name">${t.flag} ${esc(team(t.team))}</div>
     <div class="bar-track"><div class="bar-fill gold" data-w="${(t.count/mxr*100).toFixed(1)}"></div></div>
     <div class="bar-val" data-count="${t.count}">0</div>`)));
  s.appendChild(mr);
  const two = el('div','grid g2 reveal'); two.style.marginTop = '22px';
  const pol = D.polarizing.slice(0,4).map(p => {
    const top = Object.entries(p.dist).slice(0,3).map(([t,c]) => `${esc(team(t))} (${c})`).join(' · ');
    return `<div style="padding:8px 0;border-bottom:1px solid var(--line)"><b>${L('Grupo','Group')} ${p.group}</b> — ${pf(p.agreement)} ${L('de acuerdo','agreement')}<br><span class="muted" style="font-size:.84rem">${top}</span></div>`;
  }).join('');
  const th = D.top_thirds.slice(0,10).map(t => `<span class="teamchip">${t.flag} ${esc(team(t.team))} <span class="n">${t.count}</span></span>`).join('');
  two.innerHTML = `
    <div class="card"><span class="k">${L('🔥 Los grupos que más nos dividen','🔥 The groups that divide us most')}</span><div style="margin-top:10px">${pol}</div></div>
    <div class="card"><span class="k">${L('🎟️ Terceros más elegidos para colarse','🎟️ Most-picked third places')}</span><div style="margin-top:10px">${th}</div></div>`;
  s.appendChild(two);
}

/* ---- ELIMINATORIAS ---- */
function buildEliminatorias(){
  const k = D.knockout || {};
  const s = section('eliminatorias', L('05 · Eliminatorias','05 · Knockouts'),
    L('El cuadro que viene 🧩','The bracket ahead 🧩'),
    L('Calendario descargado de FIFA: cruces, fechas, horas y sedes de la fase eliminatoria. Encima aparecerá el consenso cuando se rellene el Excel.',
      'Schedule pulled from FIFA: ties, dates, kick-off times and venues for the knockout stage. Consensus appears above once the Excel is filled.'));

  if(!k.ready){
    s.appendChild(el('div','card teaser reveal',
      `<div class="em">🧩</div><h3 style="margin:10px 0 6px">${L('Calendario de eliminatorias cargado','Knockout schedule loaded')}</h3>
       <p class="muted">${L('Los cruces, fechas y horas salen de FIFA; el Excel solo aporta las predicciones y resultados. Cuando pegues las eliminatorias, aparecerá también el consenso.',
        'Fixtures, dates and kick-off times come from FIFA; the Excel only supplies predictions and results. Paste knockout picks to show consensus too.')}</p>`));
  } else {
    s.appendChild(koSummaryBlock(k));
  }
  buildEliminatoriasSchedule(s);
}

function koConsensusName(c){ return c && c.value ? `${c.flag || ''} ${esc(team(c.value))}` : '–'; }
function koAgreement(c){ return c && c.count ? `${pf(c.agreement)} · ${c.count}/${N}` : L('sin datos','no data'); }
function koMiniDist(c){
  if(!c || !c.dist || !c.dist.length) return '';
  return c.dist.slice(0,3).map(x => `${x.flag || ''} ${esc(team(x.value))} <span class="n">${x.count}</span>`).join('');
}
function koFixtureName(m, side){
  const flag = m[`fixture_${side}_flag`] || m[`resolved_${side}_flag`] || '';
  const name = m[`resolved_${side}`] || m[`fixture_${side}`] || '–';
  return `${flag ? flag + ' ' : ''}${esc(team(name))}`;
}
function koFixtureTime(m){
  const t = koTime(m);
  if(!t) return '';
  return `${t}${koNext(m) ? `<sup class="tm-next">+1</sup>` : ''} · ${koTz()}`;
}
function koMatchCardHtml(m, k){
  const winner = k.ready && m.winner && m.winner.value
    ? `<div class="ko-mini">${L('Consenso ganador','Winner consensus')}: <b>${koConsensusName(m.winner)}</b> · ${koAgreement(m.winner)}</div>` : '';
  const scoring = k.ready && m.score && m.score.value
    ? `<span class="ko-pill"><span class="ko-score">${m.score.value}</span> ${L('90 min','90 min')}</span>` : '';
  const result = m.result && m.result.score
    ? `<span class="ko-pill"><span class="ko-score">${m.result.score.home}-${m.result.score.away}</span> ${L('real','FT')}</span>` : '';
  return `<div class="ko-match" data-code="${m.code || ''}">
    <div class="ko-code">${m.code} · ${m.date || '–'}</div>
    <div class="ko-main">${koFixtureName(m,'home')} <span class="muted">vs</span> ${koFixtureName(m,'away')}</div>
    <div class="ko-mini">${koFixtureTime(m)}${m.venue ? ` · ${esc(m.venue)}` : ''}</div>
    ${winner}
    <div class="ko-pills">${scoring}${result}</div>
  </div>`;
}
function koSummaryBlock(k){
  const champ = k.outright.champion || {};
  const runner = k.outright.runner_up || {};
  const topScorer = k.awards.top_scorer || {};
  const box = el('div','ko-summary reveal');
  box.innerHTML = `
    <div class="card ko-hero"><span class="k">${L('🏆 Campeón más apostado','🏆 Most-picked champion')}</span>
      <div class="fav">${koConsensusName(champ)}</div><div class="muted">${koAgreement(champ)}</div>
      <div class="ko-pills">${koMiniDist(champ)}</div></div>
    <div class="card"><span class="k">${L('🥈 Subcampeón','🥈 Runner-up')}</span><h3>${koConsensusName(runner)}</h3><p class="muted">${koAgreement(runner)}</p></div>
    <div class="card"><span class="k">${L('👟 Máximo goleador','👟 Top scorer')}</span><h3>${koConsensusName(topScorer)}</h3><p class="muted">${koAgreement(topScorer)}</p></div>`;
  return box;
}
function koCoverageBlock(k){
  if(!k.ready) return null;
  const progress = el('div','card reveal');
  progress.innerHTML = `<span class="k">${L('Progreso de carga','Pick coverage')}</span>
    <div class="bar-row" style="grid-template-columns:minmax(84px,170px) 1fr 70px;padding-top:14px">
      <div class="bar-name">${L('Eliminatorias','Knockouts')}</div>
      <div class="bar-track"><div class="bar-fill gold" data-w="${k.pct}"></div></div>
      <div class="bar-val">${fmt(k.filled)}/${fmt(k.total)}</div>
    </div>`;
  return progress;
}
function koScoringBlock(k){
  if(!k.scoring || !k.scoring.table || !k.scoring.table.length) return null;
  const ranking = el('div','card reveal');
  const mx = k.scoring.table[0].pts || 1;
  ranking.innerHTML = `<span class="k">${L('Marcador eliminatorio','Knockout scoring')}</span>`;
  k.scoring.table.slice(0,12).forEach(r => ranking.appendChild(el('div','bar-row'+meClass(r.name),
    `<div class="bar-rank">${r.rank}</div><div class="bar-name">${r.rank===1?'👑 ':''}${esc(r.name)}
      <span class="muted" style="font-size:.78rem">(${r.exact} ${L('plenos','exact')} · ${r.advance} ${L('pases/premios','advances/awards')})</span></div>
     <div class="bar-track"><div class="bar-fill gold" data-w="${(r.pts/mx*100).toFixed(1)}"></div></div>
     <div class="bar-val" data-count="${r.pts}">0</div>`)));
  return ranking;
}
function buildEliminatoriasSchedule(parent){
  const k = D.knockout || {};
  const s = parent || section('ko-schedule', L('📅 Calendario completo','📅 Full schedule'),
    L('Cruces y consenso','Fixtures and consensus'),
    L('Todos los partidos de la fase eliminatoria con horario y apuestas agregadas.',
      'Every knockout tie with kick-off time and aggregated picks.'));
  (k.rounds || []).forEach(r => {
    const block = el('div','ko-round reveal');
    block.innerHTML = `<h3>${L(r.label_es, r.label_en)} <span class="muted" style="font-size:.82rem">+${r.advance_points} ${L('puntos por pase','pts per advance')}</span></h3>
      <div class="ko-grid">${r.matches.map(m => koMatchCardHtml(m, k)).join('')}</div>`;
    s.appendChild(block);
  });
  if(k.final_matches && k.final_matches.length){
    const finals = el('div','ko-round reveal');
    finals.innerHTML = `<h3>${L('Finales','Final weekend')}</h3><div class="ko-grid">${k.final_matches.map(m => koMatchCardHtml(m, k)).join('')}</div>`;
    s.appendChild(finals);
  }
  if(!parent) return;
  attachBracketHovers(s);
}

/* ---- PARTIDOS ---- */
function buildPartidos(){
  const s = section('partidos', L('06 · Partidos','06 · Matches'),
    L('Los partidos que nos parten en dos 🔪','The matches that split us 🔪'),
    L('Dónde hay guerra de pronósticos (poco acuerdo en el 1/X/2) y dónde casi todos ponemos lo mismo.',
      'Where there is a prediction war (little agreement on 1/X/2) and where we nearly all agree.'));
  function matchBlock(m){
    const o = m.outcome_dist, tot = (o['1']||0)+(o['X']||0)+(o['2']||0);
    const p = k => tot ? Math.round((o[k]||0)/tot*100) : 0;
    return `<div class="match-row">
      <div class="match-top"><span>${m.home_flag} ${esc(team(m.home))} – ${esc(team(m.away))} ${m.away_flag}</span>
        <span class="muted">${pf(Math.round(m.outcome_agreement*100))} ${L('acuerdo','agree')}</span></div>
      <div class="dist">
        <span class="s1" style="width:${p('1')}%">${p('1')?p('1')+'%':''}</span>
        <span class="sx" style="width:${p('X')}%">${p('X')?'X '+p('X')+'%':''}</span>
        <span class="s2" style="width:${p('2')}%">${p('2')?p('2')+'%':''}</span></div>
      <div class="modal-pill">${L('marcador más repetido','most common scoreline')}: <b class="mint">${m.modal_scoreline}</b> (${pf(Math.round(m.modal_scoreline_share*100))})</div>
    </div>`;
  }
  const two = el('div','grid g2 reveal');
  two.innerHTML = `
    <div class="card"><span class="k">${L('😵‍💫 Máxima guerra de pronósticos','😵‍💫 Maximum disagreement')}</span><div style="margin-top:8px">${D.divisive.map(matchBlock).join('')}</div></div>
    <div class="card"><span class="k">${L('🤖 Casi unanimidad','🤖 Near unanimity')}</span><div style="margin-top:8px">${D.unanimous.map(matchBlock).join('')}</div></div>`;
  s.appendChild(two);
}

/* ---- LOBO SOLITARIO ---- */
function buildLobo(){
  const s = section('lobo', L('07 · Atrevimiento','07 · Boldness'),
    L('Lobo solitario 🐺','Lone wolf 🐺'),
    L('Marcadores que solo apostó UNA persona en todo el grupo. El que más tiene, es el que más se aleja del rebaño.',
      'Scorelines only ONE person predicted in the whole pool. The more you have, the further from the flock you roam.'));
  const mx = D.lobo[0].count;
  const list = el('div','reveal');
  D.lobo.forEach((r,i) => list.appendChild(el('div','bar-row'+meClass(r.name),
    `<div class="bar-rank">${i+1}</div><div class="bar-name">${i===0?'🐺 ':''}${esc(r.name)}</div>
     <div class="bar-track"><div class="bar-fill" data-w="${(r.count/mx*100).toFixed(1)}"></div></div>
     <div class="bar-val" data-count="${r.count}">0</div>`)));
  s.appendChild(list);
  const ex = el('div','card reveal'); ex.style.marginTop = '22px';
  ex.innerHTML = `<span class="k">${L('Apuestas en solitario más salvajes','Wildest solo calls')}</span><div style="margin-top:10px">` +
    D.unique_examples.map(e => `<div style="padding:7px 0;border-bottom:1px solid var(--line)">
      <b>${esc(e.name)}</b> <span class="muted">${L('es el único que firmó','is the only one who called')}</span> ${e.flags} <b class="mint">${esc(team(e.home))} ${e.score} ${esc(team(e.away))}</b></div>`).join('') + '</div>';
  s.appendChild(ex);
}

/* ---- FICHAS ---- */
function buildFichas(){
  const s = section('fichas', L('08 · Uno a uno','08 · One by one'),
    L('La ficha de cada uno 🪪',"Everyone's card 🪪"),
    L('Resumen por persona: su puesto en rebeldía, su estilo, su gemelo y su apuesta más loca. Busca tu nombre.',
      'A summary per person: their maverick rank, style, twin and wildest bet. Search your name.'));
  const inp = el('input','search'); inp.placeholder = L('🔎 Busca tu nombre…','🔎 Search your name…');
  if(ME) inp.value = ME;
  s.appendChild(inp);
  const grid = el('div','grid g3 reveal');
  D.cards.slice().sort((a,b) => a.rebel_rank-b.rebel_rank).forEach(c => {
    const big = c.biggest ? `${c.biggest.flags} ${esc(team(c.biggest.home))} ${c.biggest.score} ${esc(team(c.biggest.away))}` : '–';
    const card = el('div','ficha'+meClass(c.name)); card.dataset.name = c.name.toLowerCase();
    card.innerHTML = `
      <div class="fh"><div><div class="fn">${esc(c.name)}</div><div class="lab">${labelOf(c.label)}</div></div>
        <div class="rk">#${c.rebel_rank}<br>${L('rebeldía','maverick')}</div></div>
      <div class="fstats">
        <div>${L('Índice rebeldía','Maverick index')}<br><span class="v">${c.rebel_index}</span></div>
        <div>${L('Goles/partido','Goals/match')}<br><span class="v">${c.avg_goals}</span></div>
        <div>${L('% empates','% draws')}<br><span class="v">${pf(c.pct_draws)}</span></div>
        <div>${L('Apuestas únicas','Unique bets')}<br><span class="v">${c.lobo}</span></div>
      </div>
      <div class="fline">${L('👯 Gemelo:','👯 Twin:')} <b>${esc(c.twin)}</b> (${pf(c.twin_pct)})</div>
      <div class="fline">${L('🎲 Su locura:','🎲 Wildest bet:')} <b>${big}</b></div>`;
    grid.appendChild(card);
  });
  s.appendChild(grid);
  inp.addEventListener('input', () => { const q = inp.value.toLowerCase().trim();
    grid.querySelectorAll('.ficha').forEach(f => { f.style.display = f.dataset.name.includes(q) ? '' : 'none'; }); });
}

/* ---- PREMIOS ---- */
function buildPremios(){
  const s = section('premios', L('09 · Palmarés','09 · Awards'),
    L('El palmarés de la porra 🏅',"The pool's hall of fame 🏅"),
    L('Los títulos honoríficos de esta edición.',"This edition's honorary titles."));
  const a = D.awards;
  const items = [
    ['🐺', L('El Rebelde','The Maverick'), a.rebelde.name, L('índice '+a.rebelde.index+'/100','index '+a.rebelde.index+'/100')],
    ['🐑', L('El Borrego','The Sheep'), a.borrego.name, L('el más previsible ('+a.borrego.index+'/100)','most predictable ('+a.borrego.index+'/100)')],
    ['💞', L('Almas Gemelas','Soulmates'), a.gemelas.a+' & '+a.gemelas.b, L(a.gemelas.pct+'% idénticos',a.gemelas.pct+'% identical')],
    ['🧊', L('Polos Opuestos','Opposites'), a.opuestos.a+' & '+a.opuestos.b, L('distancia '+a.opuestos.dist,'distance '+a.opuestos.dist)],
    ['🥅', L('El Goleador','The Goal Machine'), a.goleador.name, L(a.goleador.avg+' goles/partido',a.goleador.avg+' goals/match')],
    ['🔒', L('El Cerrojo','The Parked Bus'), a.cerrojo.name, L(a.cerrojo.avg+' goles/partido',a.cerrojo.avg+' goals/match')],
    ['🤝', L('Rey del Empate','Draw King'), a.empate.name, L(a.empate.pct+'% en tablas',a.empate.pct+'% drawn')],
    ['🎯', L('Lobo Solitario','Lone Wolf'), a.lobo.name, L(a.lobo.count+' apuestas únicas',a.lobo.count+' unique bets')],
  ];
  const grid = el('div','grid g4 reveal');
  items.forEach(([em,ti,wn,dt]) => grid.appendChild(el('div','award',
    `<div class="em">${em}</div><div class="ti">${ti}</div><div class="wn">${esc(wn)}</div><div class="dt">${dt}</div>`)));
  s.appendChild(grid);
}

/* ---- ACIERTOS ---- */
function buildAciertos(){
  const s = section('aciertos', L('📈 Ranking','📈 Ranking'),
    L('Aciertos en directo ⚽','Live scoring ⚽'),
    L('Cuando empiece a rodar el balón y rellenes los resultados reales en el Excel, aquí saldrá el ranking de aciertos.',
      'Once the ball starts rolling and you fill in the real results in the Excel, the scoring ranking appears here.'));
  if(!D.live){
    s.appendChild(el('div','card teaser reveal',
      `<div class="em">⚽️</div><h3 style="margin:10px 0 6px">${L('El Mundial aún no ha empezado',"The World Cup hasn't kicked off yet")}</h3>
       <p class="muted">${L('Rellena los marcadores reales en la pestaña <b>Real results</b> del Excel y vuelve a generar el dashboard: aparecerá aquí el ranking de aciertos (signo +2, marcador exacto +4).','Fill in the real scores in the <b>Real results</b> tab of the Excel and regenerate the dashboard: the scoring ranking shows up here (correct outcome +2, exact score +4).')}</p>`));
    return;
  }
  buildLiveRanking(s);
}

/* ---- FOOTER ---- */
function buildFooter(){
  const f = el('footer');
  f.innerHTML = `<span class="brand">${logo}</span> ${L('Porra Mundial 2026 · análisis de '+N+' quinielas · generado el __DATE__.','World Cup 2026 Pool · analysis of '+N+' predictions · generated on __DATE__.')}
    <br><span style="opacity:.6">${L('Hecho con cariño para la oficina, con el nuevo branding de Reveni: Shippori Mincho para titulares, DM Sans para texto y DM Mono para etiquetas.','Made with love for the office, in the new Reveni branding: Shippori Mincho for headlines, DM Sans for body and DM Mono for labels.')}</span>`;
  wrap.appendChild(f);
}

/* ---- OBSERVERS ---- */
function observeAll(){
  const io = new IntersectionObserver((es,o) => { es.forEach(e => { if(e.isIntersecting){
    e.target.classList.add('in');
    e.target.querySelectorAll && e.target.querySelectorAll('[data-count]').forEach(animateCount);
    e.target.querySelectorAll && e.target.querySelectorAll('.bar-fill').forEach(b => { b.style.width = (b.dataset.w||0)+'%'; });
    o.unobserve(e.target);
  }}); }, {threshold:.12});
  wrap.querySelectorAll('.reveal').forEach(n => io.observe(n));
}

/* ============================================================
   APP STRUCTURE — top tabs: En directo / Eliminatorias / Fase de grupos.
   View state lives in ?view=; default is the live (matchday) view.
   ============================================================ */
const VIEWS = [
  {key:'live',   es:'En directo',     en:'Live',        em:'⚽'},
  {key:'ko',     es:'Eliminatorias',  en:'Knockouts',   em:'🧩'},
  {key:'groups', es:'Fase de grupos', en:'Group stage', em:'📊'},
];
function currentView(){
  const r = new URLSearchParams(location.search).get('view');
  return VIEWS.some(v => v.key === r) ? r : 'live';
}
function setView(v){
  const u = new URL(location.href);
  u.searchParams.set('view', v);
  history.replaceState(null, '', u);
  rebuild(); window.scrollTo(0,0);
}
function viewBadge(key){
  if(key === 'live'){
    const n = (D.recent_results && D.recent_results.total) || (D.live && D.live.played) || 0;
    return n || '';
  }
  if(key === 'ko'){ const k = D.knockout || {}; return ((k.rounds||[]).reduce((a,r)=>a+(r.matches||[]).length,0) + (k.final_matches||[]).length) || ''; }
  if(key === 'groups') return D.hero.matches || '';
  return '';
}
function langBar(){
  return `<div class="proto-lang"><button data-l="es"${LANG==='es'?' class="on"':''}>ES</button><button data-l="en"${LANG==='en'?' class="on"':''}>EN</button></div>`;
}
function buildTopBar(activeView){
  const tabs = VIEWS.map(v => {
    const badge = viewBadge(v.key);
    return `<button class="proto-tab${v.key===activeView?' on':''}" data-view="${v.key}">
      <span class="pt-em">${v.em}</span>${L(v.es, v.en)}${badge!==''?`<span class="pt-badge">${badge}</span>`:''}</button>`;
  }).join('');
  const bar = el('div','proto-shell',
    `<div class="proto-shell-inner">
       <a class="brand" href="#" data-view="live">${logo}</a>
       <div class="proto-tabs">${tabs}</div>
       ${userBarHtml()}
       ${langBar()}
     </div>`);
  bar.addEventListener('click', e => {
    const ub = e.target.closest('[data-action="open-user-picker"]');
    if(ub){ showUserPicker(); return; }
    const lb = e.target.closest('button[data-l]');
    if(lb){ if(lb.dataset.l !== LANG){ LANG = lb.dataset.l; rebuild(); } return; }
    const t = e.target.closest('[data-view]');
    if(t){ e.preventDefault(); setView(t.dataset.view); }
  });
  wrap.appendChild(bar);
}
function renderView(view){
  if(view === 'ko') buildKnockouts();
  else if(view === 'groups'){
    buildHero(); buildRebeldia(); buildAfinidad(); buildEstilo();
    buildFavoritos(); buildPartidos(); buildLobo(); buildFichas(); buildPremios();
  }
  else {
    if(ME){
      const meSec = el('section','sec'); meSec.id = 'me-resumen';
      buildMeSummary(meSec);
      wrap.appendChild(meSec);
    }
    buildHoy(); buildAciertos(); buildUltimos();
  }
}

/* ---- BRACKET VISUALIZATION ----
   The FIFA 2026 bracket is a binary tree, but its ties are NOT sequential
   (R16-M1 = W73 vs W75, not W73 vs W74). We rebuild the tree from the W-number
   feeders so each tie sits next to its two real feeders, then draw the elbow
   connectors as an SVG overlay measured from the live DOM (pixel-accurate
   regardless of card heights / language). On mobile the tree is replaced by a
   round selector + a full-width list of that round's ties. */
const KO_WBASE = {R32:72, R16:88, QF:96, SF:100};
function koWOf(code){
  const m = /^(R32|R16|QF|SF)-M(\d+)$/.exec(code || '');
  return m ? KO_WBASE[m[1]] + (+m[2]) : null;
}
let _koWMap;
function koWToMatchMap(){
  if(_koWMap) return _koWMap;
  _koWMap = {};
  const k = D.knockout || {};
  (k.rounds || []).forEach(r => (r.matches || []).forEach(m => {
    const w = koWOf(m.code); if(w) _koWMap[w] = m;
  }));
  (k.final_matches || []).forEach(m => {
    const w = koWOf(m.code); if(w) _koWMap[w] = m;
  });
  return _koWMap;
}
function koMatchByCode(code){
  const k = D.knockout || {};
  for(const r of (k.rounds || [])){
    const m = (r.matches || []).find(x => x.code === code);
    if(m) return m;
  }
  return (k.final_matches || []).find(x => x.code === code) || null;
}
function bkSlotTipHtml(m){
  const played = !!(m.result && m.result.winner);
  if(played) return '';
  const k = D.knockout || {};
  if(!k.ready) return '';
  const wmap = koWToMatchMap();
  const slots = [];
  ['home','away'].forEach(side => {
    const fix = m['fixture_'+side] || '';
    const wm = /^W(\d+)$/.exec(fix);
    if(!wm) return;
    const feeder = wmap[+wm[1]];
    if(!feeder || !feeder.winner || !feeder.winner.dist || !feeder.winner.dist.length) return;
    const rows = feeder.winner.dist.map(d =>
      `<div class="bk-ht-row"><span>${d.flag || ''} ${esc(team(d.value))}</span><span>${d.count}/${N}</span></div>`
    ).join('');
    slots.push(`<div class="bk-ht-slot"><div class="bk-ht-lbl">${fix} · ${L('¿quién llega?','who advances?')}</div>${rows}</div>`);
  });
  if(!slots.length) return '';
  return `<div class="bk-ht-title">${esc(m.code || '')}</div>${slots.join('')}`;
}
let bktip;
function ensureBkTip(){
  if(bktip) return bktip;
  bktip = el('div'); bktip.id = 'bktip'; document.body.appendChild(bktip);
  return bktip;
}
function moveBkTip(e){
  if(!bktip) return;
  const pad = 14;
  let x = e.clientX + pad, y = e.clientY + pad;
  const r = bktip.getBoundingClientRect();
  if(x + r.width > window.innerWidth - 8) x = e.clientX - r.width - pad;
  if(y + r.height > window.innerHeight - 8) y = e.clientY - r.height - pad;
  bktip.style.left = x + 'px';
  bktip.style.top = y + 'px';
}
function attachBracketHovers(root){
  const tip = ensureBkTip();
  root.querySelectorAll('.bk-match, .ko-match').forEach(card => {
    const code = card.dataset.code;
    if(!code) return;
    const m = koMatchByCode(code);
    if(!m) return;
    const html = bkSlotTipHtml(m);
    if(!html) return;
    card.classList.add('bk-has-tip');
    card.addEventListener('mouseenter', e => {
      tip.innerHTML = html;
      tip.style.opacity = '1';
      moveBkTip(e);
    });
    card.addEventListener('mousemove', moveBkTip);
    card.addEventListener('mouseleave', () => { tip.style.opacity = '0'; });
  });
}
function bkMatchNode(m, opts){
  opts = opts || {};
  const k = D.knockout || {};
  const ready = k.ready && m.winner && m.winner.value;
  const played = !!(m.result && m.result.winner);
  const realScore = played && m.result.score
    ? `${m.result.score.home}-${m.result.score.away}` : '';
  const pct = ready ? Math.round(m.winner.agreement || 0) : 0;
  const norm = v => (v || '').toString().trim().toLowerCase();
  const distCount = value => ((m.winner && m.winner.dist) || [])
    .filter(x => norm(x.value) === norm(value))
    .reduce((a, x) => a + (x.count || 0), 0);
  const sideName = side => m['resolved_'+side] || m['fixture_'+side] || '–';
  const sideFlag = side => m['resolved_'+side+'_flag'] || m['fixture_'+side+'_flag'] || '';
  const realSide = side => {
    const name = sideName(side);
    return name && !/^W\d+$/.test(name) && !/^RU\d+$/.test(name) ? name : '';
  };
  const homeName = realSide('home'), awayName = realSide('away');
  const homeCount = ready && homeName ? distCount(homeName) : 0;
  const awayCount = ready && awayName ? distCount(awayName) : 0;
  const hasSplit = ready && homeName && awayName && (homeCount || awayCount);
  const homePct = N ? Math.round(homeCount / N * 100) : 0;
  const awayPct = N ? Math.round(awayCount / N * 100) : 0;
  const sideHtml = (side) => {
    const flag = sideFlag(side);
    const name = sideName(side);
    const count = side === 'home' ? homeCount : awayCount;
    const sidePct = side === 'home' ? homePct : awayPct;
    const isPick = ready && m.winner.value === name;
    const isRealWin = played && norm(m.result.winner) === norm(name);
    const isRealOut = played && homeName && awayName && norm(name) === norm(side === 'home' ? homeName : awayName) && !isRealWin;
    const goals = played && m.result.score
      ? `<span class="bk-goals">${m.result.score[side]}</span>` : '';
    const pctHtml = !played && count ? `<span class="bk-pct">${sidePct}%</span>` : '';
    const cls = [ready && !isPick ? ' dim' : '', isRealWin ? ' won' : '', isRealOut ? ' out' : ''].join('');
    return `<div class="bk-side${cls}">`
      + `<span class="bk-flag">${flag}</span><span class="bk-nm">${esc(team(name))}</span>`
      + `${goals || pctHtml}</div>`;
  };
  const consensus = played
    ? `<div class="bk-tip bk-real">${L('resultado','result')} ${realScore} · ${m.result.winner_flag || ''} ${esc(team(m.result.winner))}</div>`
    : (ready
      ? (hasSplit
        ? `<div class="bk-cbar split"><span class="home" style="width:${homePct}%"></span><span class="away" style="width:${awayPct}%"></span></div>`
        : `<div class="bk-cbar"><span style="width:${pct}%"></span></div>`)
        + `<div class="bk-tip">${L('consenso','consensus')} ${m.winner.count}/${N} · ${L('marcador','score')} ${m.score && m.score.value ? m.score.value : '–'}</div>`
      : `<div class="bk-tip">${L('pendiente de subir apuestas','picks not uploaded yet')}</div>`);
  const cls = (opts.cls ? ' ' + opts.cls : '') + (played ? ' played' : '');
  const feeders = opts.feeders ? ` data-feeders="${opts.feeders.join(',')}"` : '';
  return `<div class="bk-match${cls}" data-code="${m.code || ''}"${feeders}>`
    + `<div class="bk-code"><span>${m.code || ''}</span><span>${m.date || ''}</span></div>`
    + `${sideHtml('home')}${sideHtml('away')}${consensus}</div>`;
}
function drawBracketLines(tree){
  const svg = tree.querySelector('.bk-lines');
  if(!svg) return;
  const tr = tree.getBoundingClientRect();
  const w = tree.scrollWidth, h = tree.scrollHeight;
  svg.setAttribute('viewBox', `0 0 ${w} ${h}`);
  svg.setAttribute('width', w); svg.setAttribute('height', h);
  const box = node => { const r = node.getBoundingClientRect();
    return {l:r.left-tr.left, r:r.right-tr.left, cy:(r.top+r.bottom)/2-tr.top}; };
  let paths = '';
  tree.querySelectorAll('.bk-match[data-feeders]').forEach(pm => {
    const p = box(pm);
    pm.dataset.feeders.split(',').filter(Boolean).forEach(fc => {
      const cm = tree.querySelector('.bk-match[data-code="' + fc + '"]');
      if(!cm) return;
      const c = box(cm), midX = (c.r + p.l) / 2;
      paths += `<path d="M ${c.r.toFixed(1)} ${c.cy.toFixed(1)} H ${midX.toFixed(1)} V ${p.cy.toFixed(1)} H ${p.l.toFixed(1)}"/>`;
    });
  });
  svg.innerHTML = paths;
}
let _bkResizeBound = false;
function ensureBracketResize(){
  if(_bkResizeBound) return; _bkResizeBound = true;
  let t; window.addEventListener('resize', () => { clearTimeout(t); t = setTimeout(() => {
    const tr = document.querySelector('.bk-tree');
    if(tr && tr.offsetParent !== null) drawBracketLines(tr);
  }, 150); });
}
function koBracketBox(){
  const k = D.knockout || {};
  const rounds = k.rounds || [];
  const finals = k.final_matches || [];
  if(!rounds.length && !finals.length) return null;
  // index every tie by code and by the W-number its winner produces
  const byCode = {}, wToCode = {};
  const all = [];
  rounds.forEach(r => (r.matches || []).forEach(m => all.push(m)));
  finals.forEach(m => all.push(m));
  all.forEach(m => { byCode[m.code] = m; const w = koWOf(m.code); if(w) wToCode[w] = m.code; });
  const feederCodes = m => {
    const cs = ['fixture_home','fixture_away'].map(key => {
      const v = /^W(\d+)$/.exec(m[key] || ''); return v ? wToCode[+v[1]] : null;
    });
    return (cs[0] && cs[1]) ? cs : null;
  };
  // vertical order: rank each tie by the mean position of its R32 leaves so
  // every tie lines up between its two feeders
  const rankOf = {};
  const leaves = code => { const m = byCode[code]; const f = m && feederCodes(m);
    return f ? leaves(f[0]).concat(leaves(f[1])) : [code]; };
  const order = byCode['FINAL'] ? leaves('FINAL')
    : (rounds[0] ? rounds[0].matches.map(m => m.code) : []);
  order.forEach((c, i) => { rankOf[c] = i; });
  const rank = code => {
    if(rankOf[code] != null) return rankOf[code];
    const m = byCode[code], f = m && feederCodes(m);
    const r = f ? (rank(f[0]) + rank(f[1])) / 2 : 0;
    rankOf[code] = r; return r;
  };
  all.forEach(m => rank(m.code));
  const sortByRank = list => list.slice().sort((a, b) => rank(a.code) - rank(b.code));
  // columns: one per round, then a final column with Final + 3rd place
  const cols = rounds.map(r => ({ key:r.key, head:L(r.label_es, r.label_en), matches:sortByRank(r.matches || []) }));
  const finalMatch = byCode['FINAL'], thirdMatch = byCode['3P'];
  // ---- desktop tree ----
  let treeHtml = '<svg class="bk-lines" xmlns="http://www.w3.org/2000/svg"></svg>';
  cols.forEach(c => {
    treeHtml += `<div class="bk-col"><div class="bk-col-head">${c.head}</div>`
      + `<div class="bk-col-body">`
      + c.matches.map(m => bkMatchNode(m, {feeders:feederCodes(m)})).join('')
      + '</div></div>';
  });
  if(finalMatch || thirdMatch){
    let finalBlock = '', thirdBlock = '';
    if(finalMatch) finalBlock = `<div class="bk-col-head">${L('Final','Final')}</div>`
      + bkMatchNode(finalMatch, {cls:'final', feeders:feederCodes(finalMatch)});
    if(thirdMatch) thirdBlock = `<div class="bk-third-block"><div class="bk-col-head">${L('3.er puesto','3rd place')}</div>`
      + bkMatchNode(thirdMatch, {cls:'third'}) + '</div>';
    treeHtml += `<div class="bk-col bk-finalcol"><div class="bk-col-body">${finalBlock}${thirdBlock}</div></div>`;
  }
  const tree = el('div','bk-tree', treeHtml);
  // ---- mobile: round chips + list ----
  const mRounds = cols.slice();
  if(finalMatch || thirdMatch){
    mRounds.push({ key:'final', head:L('Final','Final'),
      matches:[finalMatch, thirdMatch].filter(Boolean) });
  }
  const mCls = m => m.code === 'FINAL' ? 'final' : (m.code === '3P' ? 'third' : '');
  const chips = mRounds.map((c, i) =>
    `<button class="bk-chip${i===0?' on':''}" data-i="${i}">${c.head}</button>`).join('');
  const lists = mRounds.map((c, i) =>
    `<div class="bk-list" data-i="${i}"${i===0?'':' hidden'}>`
    + c.matches.map(m => bkMatchNode(m, {cls:mCls(m)})).join('') + '</div>').join('');
  const mob = el('div','bk-mobile', `<div class="bk-chips">${chips}</div>${lists}`);
  mob.querySelector('.bk-chips').addEventListener('click', e => {
    const b = e.target.closest('.bk-chip'); if(!b) return;
    const i = b.dataset.i;
    mob.querySelectorAll('.bk-chip').forEach(x => x.classList.toggle('on', x.dataset.i === i));
    mob.querySelectorAll('.bk-list').forEach(x => { x.hidden = x.dataset.i !== i; });
  });
  const box = el('div','bracket-wrap reveal');
  box.appendChild(tree); box.appendChild(mob);
  requestAnimationFrame(() => drawBracketLines(tree));
  if(document.fonts && document.fonts.ready){
    document.fonts.ready.then(() => { if(tree.isConnected) drawBracketLines(tree); });
  }
  ensureBracketResize();
  attachBracketHovers(box);
  return box;
}
function koResultsStarted(){
  return !!(D.knockout && D.knockout.results_started);
}
function koMetricsData(){
  return (D.knockout && D.knockout.metrics) || null;
}
function koRoundLabel(r){
  return typeof r === 'string' ? r : L(r.es, r.en);
}
function koSection(){ const sec = el('section','sec'); wrap.appendChild(sec); return sec; }
function koBracketCellHtml(cell, roundLabel){
  const total = Math.max(1, cell.total || 1);
  const okW = 100 * (cell.hits || 0) / total;
  const badW = 100 * (cell.misses || 0) / total;
  const driftOp = Math.min(1, (cell.drift || 0) / total * 1.6);
  const layers = [];
  if(okW > 0){
    layers.push(`<span class="br-layer br-ok" style="width:${okW.toFixed(2)}%"></span>`);
  }
  if(badW > 0){
    layers.push(`<span class="br-layer br-bad" style="left:${okW.toFixed(2)}%;width:${badW.toFixed(2)}%"></span>`);
  }
  if((cell.drift || 0) > 0){
    layers.push(`<span class="br-layer br-drift" style="opacity:${driftOp.toFixed(2)}"></span>`);
  }
  return `<div class="surv-cell br-stack" title="${esc(koBracketCellTip(roundLabel, cell))}">${layers.join('')}</div>`;
}
function koBracketCellTip(roundLabel, cell){
  const total = cell.total || 0;
  const bits = [];
  if(cell.hits) bits.push(cell.hits + '/' + total + L(' aciertos',' hits'));
  if(cell.misses) bits.push(cell.misses + '/' + total + L(' fallos',' misses'));
  if(cell.drift) bits.push(cell.drift + L(' rama(s) desviada(s)',' off-branch tie(s)'));
  if(!bits.length) return roundLabel + ': ' + L('sin resultados aún','no results yet');
  return roundLabel + ': ' + bits.join(' · ');
}
function koBracketPrecisionCard(){
  const dz = koMetricsData();
  if(!dz || !dz.people || !dz.people.length || !dz.bracketRounds) return null;
  const rounds = dz.bracketRounds.map(koRoundLabel);
  const rows = dz.people.slice().sort((a,b) => a.name.localeCompare(b.name,'es'));
  const head = `<div class="surv-head"></div>` + rounds.map(r => `<div class="surv-head">${r}</div>`).join('');
  const body = rows.map(p => {
    const cells = (p.bracket || []).map((cell, i) => koBracketCellHtml(cell, rounds[i])).join('');
    return `<div class="surv-name${meClass(p.name)}">${esc(p.name)}</div>${cells}`;
  }).join('');
  const legend = `<div class="surv-legend">
    <span><i class="lg-grad"></i>${L('Proporcional a la fase (1 fallo en 16 ≠ todo rojo)','Scaled to the round (1 miss in 16 ≠ all red)')}</span>
    <span><i class="surv-cell"><span class="br-layer br-drift" style="opacity:.85;position:relative;display:block;height:100%"></span></i>${L('Rama muerta (tu equipo ya está fuera)','Dead branch (your team is already out)')}</span>
  </div>`;
  return el('div','survive reveal',
    `<h3>${L('Tu cuadro, fase a fase','Your bracket, round by round')}</h3>
     <p class="muted" style="margin-bottom:14px">${L('Verde y rojo son proporcionales a los cruces de cada fase (16 en 1/16, 8 en octavos…). Rojo = jugado y fallaste al que pasa. Rayado = cruce por jugar cuyo ganador que pusiste ya está eliminado.',
        'Green and red are proportional to ties in each round (16 in R32, 8 in R16…). Red = played and you missed who advanced. Stripes = an upcoming tie whose winner you picked is already out.')}</p>
     <div class="surv-scroll"><div class="surv-grid" style="--rounds:${rounds.length}">${head}${body}</div></div>${legend}`);
}
function buildKnockouts(){
  const k = D.knockout || {};
  if(!k.ready || !k.metrics){
    buildEliminatorias();
    return;
  }
  buildKoStoryBody();
  buildEliminatoriasSchedule();
  const coverage = koCoverageBlock(k);
  if(coverage){
    const s = el('section','sec');
    s.appendChild(coverage);
    wrap.appendChild(s);
  }
}
let _koTipEl = null, _koTipReady = false;
function ensureScatterTip(){
  if(_koTipReady) return;
  _koTipReady = true;
  _koTipEl = el('div','kp-tip'); document.body.appendChild(_koTipEl);
  const position = e => {
    const pad = 14; const r = _koTipEl.getBoundingClientRect();
    let x = e.clientX + pad, y = e.clientY + pad;
    if(x + r.width > innerWidth) x = e.clientX - r.width - pad;
    if(y + r.height > innerHeight) y = e.clientY - r.height - pad;
    _koTipEl.style.left = x + 'px'; _koTipEl.style.top = y + 'px';
  };
  document.addEventListener('mouseover', e => {
    const t = e.target;
    if(t && t.tagName === 'circle' && t.dataset && t.dataset.name){
      _koTipEl.innerHTML = `<b style="color:${t.dataset.color}">${t.dataset.name}</b><span class="muted">${t.dataset.info}</span>`;
      _koTipEl.style.opacity = '1'; position(e);
    }
  });
  document.addEventListener('mousemove', e => { if(_koTipEl.style.opacity === '1') position(e); });
  document.addEventListener('mouseout', e => { if(e.target && e.target.tagName === 'circle') _koTipEl.style.opacity = '0'; });
}
function koScatter(people){
  ensureScatterTip();
  const W = 560, H = 300, pad = 46;
  const xs = people.map(p => p.exp), ys = people.map(p => p.variance);
  const xmin = Math.min(...xs) - 3, xmax = Math.max(...xs) + 3, ymin = Math.min(...ys) - 3, ymax = Math.max(...ys) + 3;
  const X = v => pad + (v - xmin) / (xmax - xmin) * (W - pad - 12);
  const Y = v => H - pad - (v - ymin) / (ymax - ymin) * (H - pad - 24);
  const approx = people.some(p => p.expApprox);
  const dots = people.map(p => {
    const expLbl = p.expApprox ? L('pts esp. (aprox.)','exp pts (approx.)') : L('pts KO','KO pts');
    const info = `${p.exp} ${expLbl} · ${L('riesgo','risk')} ${p.variance}/100 · ${p.boldPct}% ${L('fuera del consenso','off consensus')}`;
    return `<circle class="${isMe(p.name)?'is-me':''}" data-name="${esc(p.name)}" data-info="${esc(info)}" data-color="${personColor(p.name)}" cx="${X(p.exp).toFixed(1)}" cy="${Y(p.variance).toFixed(1)}" r="${isMe(p.name)?7:6}" fill="${personColor(p.name)}" opacity=".85"></circle>`;
  }).join('');
  const xLbl = approx ? L('puntos esperados (aprox.) →','expected pts (approx.) →') : L('puntos KO →','KO points →');
  return `<svg class="kp-scatter" viewBox="0 0 ${W} ${H}">
    <line class="axis" x1="${pad}" y1="${H-pad}" x2="${W-6}" y2="${H-pad}"></line>
    <line class="axis" x1="${pad}" y1="14" x2="${pad}" y2="${H-pad}"></line>
    <text x="${W-6}" y="${H-pad+24}" text-anchor="end">${xLbl}</text>
    <text x="${pad}" y="12" text-anchor="middle">${L('↑ riesgo (0–100)','↑ risk (0–100)')}</text>
    ${dots}</svg>`;
}
function koChampionPathHtml(dz){
  if(!dz.championPath || !dz.championPath.length) return '';
  const steps = dz.championPath.map((s, i) => {
    const arrow = i < dz.championPath.length - 1 ? `<span class="kp-path-arrow">→</span>` : '';
    return `<div class="kp-path-step"><div class="lab">${L(s.round_es, s.round_en)}</div>
      <div class="opp">${s.opponent_flag} ${esc(team(s.opponent))}</div>
      <div class="muted" style="font-size:.76rem;margin-top:3px">${s.agreement}% ${L('consenso','consensus')}</div></div>${arrow}`;
  }).join('');
  return `<div class="kp-path">${steps}</div>`;
}
function koVsPuebloHtml(dz){
  if(!dz.vsPuebloRank || !dz.vsPuebloRank.length) return `<p class="muted">${L('Sin datos aún','No data yet')}</p>`;
  const mx = dz.vsPuebloRank[0].boldPct || 1;
  return dz.vsPuebloRank.slice(0,8).map((p,i) => `<div class="bar-row"><div class="bar-rank">${i+1}</div>
      <div class="bar-name" style="color:${personColor(p.name)}">${esc(p.name)}</div>
      <div class="bar-track"><div class="bar-fill" data-w="${mx ? (p.boldPct/mx*100).toFixed(0) : 0}"></div></div>
      <div class="bar-val">${p.boldPct}%</div></div>`).join('');
}
function koHonorsHtml(h){
  if(!h) return '';
  const items = [
    ['🔮', L('El Profeta','The Prophet'), h.profeta.name,
      h.profeta.approx ? L(h.profeta.pts + ' pts esp.',''+h.profeta.pts+' exp pts') : h.profeta.pts + ' ' + L('pts KO','KO pts')],
    ['📢', L('El Agorero','The Doomsayer'), h.agorero.name, L('riesgo '+h.agorero.risk+'/100','risk '+h.agorero.risk+'/100')],
    ['📋', L('El de manual','Mr. Chalk'), h.manual.name, L('riesgo '+h.manual.risk+'/100','risk '+h.manual.risk+'/100')],
    ['💥', L('El Reventador','The Upset Caller'), h.reventador.name,
      h.reventador.count ? L(h.reventador.count + ' sorpresa(s) acertada(s)', h.reventador.count + ' upset(s) called') : L('ninguna aún','none yet')],
    ['🏘️', L('Contra el pueblo','Against the crowd'), h.vsPueblo.name,
      L(h.vsPueblo.pct + '% fuera del consenso (' + h.vsPueblo.diff + ' cruces)',
        h.vsPueblo.pct + '% off consensus (' + h.vsPueblo.diff + ' ties)')],
  ];
  return items.map(([em,ti,wn,dt]) => `<div class="kp-honor"><div class="em">${em}</div><div class="ti">${ti}</div>
    <div class="wn">${esc(wn)}</div><div class="dt">${dt}</div></div>`).join('');
}
/* ---- métricas KO (fragmentos reutilizables del relato) ---- */
function koChampBarsHtml(dz){
  if(!dz.champRank || !dz.champRank.length) return `<p class="muted">${L('Sin datos aún','No data yet')}</p>`;
  const mx = dz.champRank[0].count || 1;
  return dz.champRank.map((c,i) => `<div class="bar-row"><div class="bar-rank">${i+1}</div>
      <div class="bar-name">${c.flag} ${esc(team(c.team))}</div>
      <div class="bar-track"><div class="bar-fill gold" data-w="${(c.count/mx*100).toFixed(0)}"></div></div>
      <div class="bar-val" data-count="${c.count}">0</div></div>`).join('');
}
function koBoldBarsHtml(dz){
  if(!dz.boldRank || !dz.boldRank.length) return `<p class="muted">${L('Sin datos aún','No data yet')}</p>`;
  const mx = dz.boldRank[0].boldPct || 1;
  return dz.boldRank.slice(0,8).map((p,i) => `<div class="bar-row${meClass(p.name)}"><div class="bar-rank">${i+1}</div>
      <div class="bar-name" style="color:${personColor(p.name)}">${esc(p.name)}</div>
      <div class="bar-track"><div class="bar-fill" data-w="${mx ? (p.boldPct/mx*100).toFixed(0) : 0}"></div></div>
      <div class="bar-val">${p.boldPct}%</div></div>`).join('');
}
function koTwinsHtml(dz){
  if(!dz.twins || !dz.twins.length) return `<p class="muted">${L('Ningún par supera el 45 % de similitud.','No pair clears 45% similarity.')}</p>`;
  return dz.twins.slice(0,6).map(t => `<div class="kp-twin-row"><span>${esc(t.a)} · ${esc(t.b)}</span>
      <div class="kp-twin-bar"><span style="width:${t.sim}%"></span></div><span class="pct">${t.sim}%</span></div>`).join('');
}
function koGraveHtml(dz){
  return (dz.grave && dz.grave.length)
    ? dz.grave.map(g => `<div class="kp-grave-card"><div class="x">⚰️</div><div class="nm">${esc(g.name)}</div><div class="ch">${g.flag} ${esc(team(g.champ))}</div></div>`).join('')
    : `<p class="muted">${L('Nadie enterrado… todavía.','Nobody buried… yet.')}</p>`;
}
function koPichichiHtml(dz){
  if(!dz.tsRank || !dz.tsRank.length) return `<p class="muted">${L('Sin datos aún','No data yet')}</p>`;
  const mx = dz.tsRank[0].count || 1;
  return dz.tsRank.slice(0,6).map((t,i) => `<div class="bar-row"><div class="bar-rank">${i+1}</div>
      <div class="bar-name">${esc(t.name)}</div>
      <div class="bar-track"><div class="bar-fill" data-w="${(t.count/mx*100).toFixed(0)}"></div></div>
      <div class="bar-val" data-count="${t.count}">0</div></div>`).join('');
}
function koFichasBody(dz){
  const twinOf = {};
  (dz.twins || []).forEach(t => { if(!twinOf[t.a]) twinOf[t.a] = {name:t.b, sim:t.sim}; if(!twinOf[t.b]) twinOf[t.b] = {name:t.a, sim:t.sim}; });
  const box = el('div','kp-fichas-wrap');
  const inp = el('input','search'); inp.placeholder = L('🔎 Busca tu nombre…','🔎 Search your name…');
  if(ME) inp.value = ME;
  const grid = el('div','grid g3');
  dz.people.slice().sort((a,b) => a.name.localeCompare(b.name,'es')).forEach(p => {
    const survives = p.fell >= dz.rounds.length;
    const fellTxt = survives ? L('aguanta hasta el final','lasts to the end') : koRoundLabel(dz.rounds[p.fell]);
    const tw = twinOf[p.name];
    const card = el('div','ficha'+meClass(p.name)); card.dataset.name = p.name.toLowerCase();
    card.innerHTML = `
      <div class="fh"><div><div class="fn">${esc(p.name)}</div><div class="lab">🏆 ${esc(team(p.champ[0]))}</div></div>
        <div class="rk">${p.exp}<br>${p.expApprox ? L('pts esp.','exp pts') : L('pts KO','KO pts')}</div></div>
      <div class="fstats">
        <div>${L('Campeón','Champion')}<br><span class="v">${p.champ[1]} ${esc(team(p.champ[0]))}</span></div>
        <div>${L('Subcampeón','Runner-up')}<br><span class="v">${p.runner[1]} ${esc(team(p.runner[0]))}</span></div>
        <div>${L('Pichichi','Top scorer')}<br><span class="v">${esc(p.ts)}</span></div>
        <div>${L('Riesgo','Risk')}<br><span class="v">${p.variance}/100</span></div>
      </div>
      <div class="fline">${L('📉 Tu campeón cae en:','📉 Your champion falls at:')} <b>${fellTxt}</b></div>
      ${p.vsPuebloTotal ? `<div class="fline">${L('🏘️ vs pueblo:','🏘️ vs crowd:')} <b>${p.boldPct}%</b> (${p.vsPueblo}/${p.vsPuebloTotal} ${L('cruces','ties')})</div>` : ''}
      ${tw ? `<div class="fline">${L('👯 Gemelo:','👯 Twin:')} <b>${esc(tw.name)}</b> (${tw.sim}%)</div>` : ''}`;
    grid.appendChild(card);
  });
  inp.addEventListener('input', () => { const q = inp.value.toLowerCase().trim();
    grid.querySelectorAll('.ficha').forEach(f => { f.style.display = f.dataset.name.includes(q) ? '' : 'none'; }); });
  box.appendChild(inp); box.appendChild(grid);
  return box;
}
function koAct(kicker, title, leadHtml, bodyEl){
  const a = el('div','kp-act reveal');
  a.innerHTML = `<div class="kp-act-kicker">${kicker}</div><h2>${title}</h2>${leadHtml || ''}`;
  if(bodyEl) a.appendChild(bodyEl);
  return a;
}
function buildKoStoryBody(){
  const dz = koMetricsData();
  const k = D.knockout || {};
  if(!dz) return;
  const top = dz.champRank[0] || {count:0, team:'–', flag:'🏳️'};
  const second = dz.champRank[1] || {count:0, team:'–', flag:'🏳️'};
  const bold = dz.boldRank[0] || {name:'–', boldPct:0, vsPueblo:0, vsPuebloTotal:0};
  const safe = dz.boldRank[dz.boldRank.length - 1] || bold;
  const prophetRow = k.scoring && k.scoring.table && k.scoring.table[0];
  const prophet = prophetRow
    ? {name: prophetRow.name, exp: prophetRow.pts, demo: false}
    : {name: (dz.people.slice().sort((a,b) => b.exp - a.exp)[0] || {name:'–', exp:0}).name,
       exp: (dz.people.slice().sort((a,b) => b.exp - a.exp)[0] || {exp:0}).exp,
       demo: dz.expApprox !== false};
  const sec = koSection();

  const bracketBox = koBracketBox();
  if(bracketBox){
    const act1Body = el('div');
    act1Body.appendChild(bracketBox);
    if(dz.championPath && dz.championPath.length){
      const pathViz = el('div','kp-act-viz');
      pathViz.innerHTML = `<span class="k">${L('🛤️ Camino del campeón del pueblo','🛤️ The crowd\'s champion path')}</span>
        <p class="muted" style="margin:8px 0 4px">${dz.championFlag} ${esc(team(dz.championTeam))} · ${L('cuadro consensual ronda a ronda','consensus bracket round by round')}</p>
        ${koChampionPathHtml(dz)}`;
      act1Body.appendChild(pathViz);
    }
    sec.appendChild(koAct(
      L('Acto 1 · El cuadro','Act 1 · The bracket'),
      L('El camino hacia la final','The road to the final'),
      `<p class="kp-act-lead">${k.ready
        ? L('El cuadro oficial de FIFA. Cada cruce muestra el % de consenso del ganador y el marcador más probable. Pasa el cursor sobre cruces con W## para ver quién puede llegar a ese hueco.',
            'The official FIFA bracket. Each tie shows the winner consensus % and the most likely scoreline. Hover ties with W## slots to see who might fill them.')
        : L('El cuadro oficial de FIFA. Cuando suban las apuestas, cada cruce mostrará el % de consenso del ganador y el marcador más probable.',
            'The official FIFA bracket. Once picks are uploaded, each tie shows the winner consensus % and the most likely scoreline.')}</p>`,
      act1Body));
  }

  if(top.count){
    sec.appendChild(koAct(
      L('Acto 2 · El consenso','Act 2 · The consensus'),
      L('El pueblo ha hablado','The people have spoken'),
      `<div class="kp-act-num">${top.flag} ${esc(team(top.team))}</div>
       <p class="kp-act-lead">${L(top.count + ' de ' + N + ' coronan a ' + team(top.team) + '. El siguiente, ' + team(second.team) + ', se queda en ' + second.count + '.',
                                 top.count + ' of ' + N + ' crown ' + team(top.team) + '. Next up, ' + team(second.team) + ', stalls at ' + second.count + '.')}</p>`,
      el('div','kp-act-viz', `<span class="k">${L('🏆 Reparto del título','🏆 Title split')}</span>${koChampBarsHtml(dz)}`)));
  }

  sec.appendChild(koAct(
    L('Acto 3 · Carácter','Act 3 · Character'),
    L('Valientes contra los de manual','The bold vs the by-the-book'),
    `<div class="kp-duo">
       <div class="b"><div class="muted">${L('🐺 El más atrevido','🐺 Boldest')}</div><div class="big" style="color:${personColor(bold.name)}">${esc(bold.name)}</div><div class="muted">${bold.boldPct}% ${L('fuera del consenso','off consensus')} (${bold.vsPueblo}/${bold.vsPuebloTotal})</div></div>
       <div class="b"><div class="muted">${L('🐑 El más de manual','🐑 Most chalk')}</div><div class="big" style="color:${personColor(safe.name)}">${esc(safe.name)}</div><div class="muted">${safe.boldPct}% ${L('fuera del consenso','off consensus')} (${safe.vsPueblo}/${safe.vsPuebloTotal})</div></div>
     </div>`,
    el('div','kp-act-grid',
      `<div class="kp-act-viz"><span class="k">${L('🎯 Atrevimiento','🎯 Boldness')}</span>
        <p class="muted" style="font-size:.82rem;margin:0 0 10px">${L('% de cruces KO donde tu ganador ≠ el favorito del pueblo.',
          '% of KO ties where your winner pick ≠ the crowd favourite.')}</p>${koBoldBarsHtml(dz)}</div>`
      + `<div class="kp-act-viz"><span class="k">${L('⚖️ Riesgo vs recompensa','⚖️ Risk vs reward')}</span>${koScatter(dz.people)}<p class="muted" style="margin-top:10px;font-size:.82rem">${dz.expApprox
        ? L('Eje X: pts esperados por alineación con el consenso. Eje Y: riesgo 0–100 (prob. implícita contraria).','X: expected pts from consensus alignment. Y: risk 0–100 (contrarian vs crowd).')
        : L('Eje X: pts KO reales. Eje Y: riesgo 0–100 (prob. implícita contraria).','X: actual KO pts. Y: risk 0–100 (contrarian vs crowd).')}</p></div>`)));

  if(koResultsStarted()){
    const bracketAct = koAct(
      L('Acto 4 · Tu cuadro','Act 4 · Your bracket'),
      L('Precisión fase a fase','Round-by-round accuracy'),
      `<p class="kp-act-lead">${L('Qué aciertas en cada ronda y dónde se desvía tu rama respecto al torneo real.',
                                'What you get right each round and where your branch drifts off from the real tournament.')}</p>`,
      koBracketPrecisionCard() || el('div','muted', L('Sin resultados KO todavía.','No knockout results yet.')));
    sec.appendChild(bracketAct);
  }

  sec.appendChild(koAct(
    L('Acto 5 · Almas gemelas','Act 5 · Soulmates'),
    L('¿Quién piensa como quién?','Who thinks like whom?'),
    `<p class="kp-act-lead">${L('Cuadros casi calcados: mismos pases, mismo campeón, mismo pichichi.',
                              'Near-identical brackets: same advances, champion and top scorer.')}</p>`,
    el('div','kp-act-viz', koTwinsHtml(dz))));

  const graveBody = el('div','kp-grave', koGraveHtml(dz));
  graveBody.style.marginTop = '18px';
  sec.appendChild(koAct(
    L('Acto 6 · El cementerio','Act 6 · The graveyard'),
    L('Campeones caídos','Fallen champions'),
    `<p class="kp-act-lead">${L('Su campeón ya está fuera. Un minuto de silencio.','Their champion is already out. A minute of silence.')}</p>`,
    graveBody));

  const prophetBody = el('div','kp-act-viz');
  prophetBody.innerHTML = `<span class="k">${L('👟 Pichichi del pueblo','👟 People\'s top scorer')}</span>${koPichichiHtml(dz)}`;
  const scoringBlock = koScoringBlock(k);
  if(scoringBlock){
    scoringBlock.style.marginTop = '18px';
    prophetBody.appendChild(scoringBlock);
  }
  sec.appendChild(koAct(
    L('Acto 7 · El profeta','Act 7 · The prophet'),
    L('El que más puntos lleva','The one leading on points'),
    `<div class="kp-act-num" style="color:${personColor(prophet.name)}">${esc(prophet.name)}</div>
     <p class="kp-act-lead">${prophet.demo
       ? L('Valor esperado ' + prophet.exp + ' pts en la fase final (aprox.).', 'Expected value ' + prophet.exp + ' pts in the final phase (approx.).')
       : L(prophet.exp + ' pts en eliminatorias tras ' + (k.scoring.played || 0) + ' partido(s).',
           prophet.exp + ' knockout pts after ' + (k.scoring.played || 0) + ' match(es).')}</p>`,
    prophetBody));

  if(dz.depth && dz.depth.length){
    sec.appendChild(koAct(
      L('Acto 8 · Profundidad','Act 8 · Depth'),
      L('¿Hasta dónde la ve la oficina?','How far does the office see them?'),
      `<p class="kp-act-lead">${L('Por selección: % de cuadros que la llevan a cada ronda.',
                                'Per team: % of brackets taking them to each round.')}</p>`,
      (() => {
        const chips = dz.depth.map((d,i) => `<button class="kp-teamchip${i===0?' on':''}" data-i="${i}">${d.flag} ${esc(team(d.team))}</button>`).join('');
        const dossier = el('div','kp-dossier', `<div class="kp-chips">${chips}</div><div class="kp-dossier-body"></div>`);
        const bodyEl = dossier.querySelector('.kp-dossier-body');
        const renderPanel = i => {
          const d = dz.depth[i];
          const rows = [['R16',d.r16],['QF',d.qf],['SF',d.sf],[L('Final','Final'),d.fin],[L('Campeón','Champ'),d.champ]];
          bodyEl.innerHTML = `<div class="kp-dossier-head">${d.flag} ${esc(team(d.team))} — ${L('% de cuadros que la llevan a…','% of brackets taking them to…')}</div>`
            + rows.map(([lab,v]) => `<div class="bar-row"><div class="bar-name">${lab}</div>
                <div class="bar-track"><div class="bar-fill" data-w="${v}" style="width:${v}%"></div></div>
                <div class="bar-val">${v}%</div></div>`).join('');
        };
        dossier.querySelector('.kp-chips').addEventListener('click', e => {
          const b = e.target.closest('.kp-teamchip'); if(!b) return;
          dossier.querySelectorAll('.kp-teamchip').forEach(x => x.classList.toggle('on', x === b));
          renderPanel(+b.dataset.i);
        });
        renderPanel(0);
        return dossier;
      })()));
  }

  const aCards = koAct(
    L('Acto 9 · Las fichas','Act 9 · The cards'),
    L('La ficha de cada uno 🪪',"Everyone's card 🪪"),
    `<p class="kp-act-lead">${L('Resumen por persona: su campeón, su pichichi, lo loco que va su cuadro y hasta dónde aguanta. Busca tu nombre.',
                              'A summary per person: their champion, top scorer, how wild their bracket is and how far it lasts. Search your name.')}</p>`,
    koFichasBody(dz));
  aCards.style.borderBottom = '0';
  sec.appendChild(aCards);

  if(dz.honors){
    const honorsAct = koAct(
      L('Acto 10 · Palmarés KO','Act 10 · KO honours'),
      L('Los títulos de la fase final','Knockout stage titles'),
      `<p class="kp-act-lead">${L('Premios honoríficos de eliminatorias: profeta, agorero, manual, reventador y más.',
                                'Knockout honorary titles: prophet, doomsayer, chalk, upset caller and more.')}</p>`,
      el('div','kp-honors reveal', koHonorsHtml(dz.honors)));
    honorsAct.style.borderBottom = '0';
    sec.appendChild(honorsAct);
  }
}

/* ---- BUILD / REBUILD ---- */
function rebuild(){
  clearRaceTimer();
  loadMe();
  if(wrap) wrap.remove();
  wrap = el('div','wrap'); document.body.appendChild(wrap);
  const view = currentView();
  buildTopBar(view);
  const body = el('div','proto-body'); wrap.appendChild(body);
  const prevWrap = wrap; wrap = body;     // route section() appends into the centred body
  renderView(view);
  wrap = prevWrap;
  buildFooter();
  observeAll();
  maybeShowUserPicker();
  afterRenderMeHooks(view);
}
ensureUserPicker();
loadMe();
rebuild();

/* ----- Fondo interactivo: red de puntos que reacciona al ratón (sello Reveni) ----- */
(function bgNet(){
  if(window.matchMedia && matchMedia('(prefers-reduced-motion:reduce)').matches) return;
  const cv = document.createElement('canvas'); cv.id = 'bgnet'; cv.setAttribute('aria-hidden','true');
  document.body.insertBefore(cv, document.body.firstChild);
  const ctx = cv.getContext('2d');
  const ORANGE = '242,133,54', MINT = '108,232,105';
  const LINK = 138, MOUSE_R = 200;
  let W, H, DPR, parts = [], raf = 0;
  const mouse = { x: -9999, y: -9999, active: false };
  function size(){
    DPR = Math.min(window.devicePixelRatio || 1, 2);
    W = cv.width = Math.floor(innerWidth * DPR); H = cv.height = Math.floor(innerHeight * DPR);
    cv.style.width = innerWidth + 'px'; cv.style.height = innerHeight + 'px';
    const target = Math.max(24, Math.min(72, Math.round(innerWidth * innerHeight / 22000)));
    parts = [];
    for(let i=0;i<target;i++) parts.push({
      x: Math.random()*W, y: Math.random()*H,
      vx: (Math.random()-.5)*.22*DPR, vy: (Math.random()-.5)*.22*DPR
    });
  }
  addEventListener('resize', size, {passive:true});
  addEventListener('mousemove', e => { mouse.x = e.clientX*DPR; mouse.y = e.clientY*DPR; mouse.active = true; }, {passive:true});
  addEventListener('mouseout', () => { mouse.active = false; mouse.x = mouse.y = -9999; }, {passive:true});
  size();
  function frame(){
    ctx.clearRect(0,0,W,H);
    const lk = LINK*DPR, mr = MOUSE_R*DPR, maxv = .68*DPR;
    for(const p of parts){
      p.x += p.vx; p.y += p.vy;
      if(p.x < 0 || p.x > W) p.vx *= -1;
      if(p.y < 0 || p.y > H) p.vy *= -1;
      if(mouse.active){
        const dx = mouse.x-p.x, dy = mouse.y-p.y, d = Math.hypot(dx,dy) || 1;
        if(d < mr){ const f = (1 - d/mr)*0.06; p.vx += dx/d*f; p.vy += dy/d*f; }
      }
      const sp = Math.hypot(p.vx,p.vy);
      if(sp > maxv){ p.vx = p.vx/sp*maxv; p.vy = p.vy/sp*maxv; }
    }
    for(let i=0;i<parts.length;i++){
      const a = parts[i];
      for(let j=i+1;j<parts.length;j++){
        const b = parts[j], dx = a.x-b.x, dy = a.y-b.y, d = Math.hypot(dx,dy);
        if(d < lk){
          ctx.strokeStyle = 'rgba('+MINT+','+(0.22*(1-d/lk)).toFixed(3)+')';
          ctx.lineWidth = DPR; ctx.beginPath(); ctx.moveTo(a.x,a.y); ctx.lineTo(b.x,b.y); ctx.stroke();
        }
      }
      if(mouse.active){
        const dx = a.x-mouse.x, dy = a.y-mouse.y, d = Math.hypot(dx,dy);
        if(d < mr){
          ctx.strokeStyle = 'rgba('+ORANGE+','+(0.55*(1-d/mr)).toFixed(3)+')';
          ctx.lineWidth = 1.2*DPR; ctx.beginPath(); ctx.moveTo(a.x,a.y); ctx.lineTo(mouse.x,mouse.y); ctx.stroke();
        }
      }
    }
    ctx.fillStyle = 'rgba('+ORANGE+',.9)';
    for(const p of parts){ ctx.beginPath(); ctx.arc(p.x,p.y,2.2*DPR,0,6.2832); ctx.fill(); }
    raf = requestAnimationFrame(frame);
  }
  raf = requestAnimationFrame(frame);
  document.addEventListener('visibilitychange', () => {
    if(document.hidden){ cancelAnimationFrame(raf); raf = 0; }
    else if(!raf){ raf = requestAnimationFrame(frame); }
  });
})();
"""

HTML_TEMPLATE = (
    '<!DOCTYPE html>\n<html lang="es">\n<head>\n<meta charset="utf-8">\n'
    '<meta name="viewport" content="width=device-width,initial-scale=1">\n'
    '<title>Porra Mundial 2026 · Reveni</title>\n'
    '<link rel="icon" href="favicon.svg" type="image/svg+xml">\n'
    '<link rel="preconnect" href="https://fonts.googleapis.com">\n'
    '<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>\n'
    '<link href="https://fonts.googleapis.com/css2?family=DM+Sans:opsz,wght@9..40,400;9..40,500;9..40,600;9..40,700&'
    'family=DM+Mono:wght@400;500&family=Shippori+Mincho:wght@500;600;700&display=swap" rel="stylesheet">\n'
    '<style>' + CSS + '</style>\n</head>\n<body>\n'
    '<div id="app"></div>\n'
    '<script>window.__PORRA__=__DATA__;</script>\n'
    '<script>' + JS + '</script>\n'
    '</body>\n</html>\n'
)


def load_logo():
    p = Path(LOGO_SVG_PATH)
    if not p.exists():
        return '<span class="wordmark">reveni</span>'
    svg = p.read_text(encoding="utf-8")
    # El wordmark ya usa fill="currentColor" (se tiñe vía CSS .brand);
    # el swoosh lleva el menta de marca horneado.
    return svg


def render_html(analytics, generated_on):
    data_json = json.dumps(analytics, ensure_ascii=False)
    logo = load_logo()
    return (
        HTML_TEMPLATE
        .replace("__DATA__", data_json)
        .replace("__LOGO__", logo)
        .replace("__DATE__", generated_on)
    )


if __name__ == "__main__":
    from datetime import date

    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    out = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUT
    data = parse_workbook(xlsx)
    analytics = compute(data)
    html = render_html(analytics, date.today().strftime("%d/%m/%Y"))
    Path(out).write_text(html, encoding="utf-8")
    ko_matches = sum(len(r["matches"]) for r in data["knockouts"]["rounds"]) + len(data["knockouts"]["final_matches"])
    ko_results = len(data["knockout_results"]["matches"])
    print(f"OK · {data['n']} participantes · grupos: {len(data['matches'])} partidos / {len(data['results'])} resultados · "
          f"KO: {ko_matches} cruces / {ko_results} resultados")
    print(f"Dashboard escrito en: {out}  ({len(html) // 1024} KB)")
